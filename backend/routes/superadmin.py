from flask import Blueprint, request, jsonify, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from bson import ObjectId
import bcrypt
import csv
import io
import json
from mongo import mongo_db
from config.constants import ROLES, MODULES, LEVELS, TEST_TYPES, WRITING_CONFIG
from datetime import datetime, timedelta
from routes.test_management import require_superadmin
from models import Test

superadmin_bp = Blueprint('superadmin', __name__)

# Define allowed admin roles
ALLOWED_ADMIN_ROLES = {ROLES['SUPER_ADMIN'], ROLES['CAMPUS_ADMIN'], ROLES['COURSE_ADMIN']}

def safe_isoformat(date_obj):
    """Safely convert a date object to ISO format string, handling various types."""
    if not date_obj:
        return None
    
    if hasattr(date_obj, 'isoformat'):
        # It's a datetime object
        return date_obj.isoformat()
    elif isinstance(date_obj, dict):
        # It's a MongoDB date dict, extract the date
        if '$date' in date_obj:
            try:
                from datetime import datetime
                date_str = date_obj['$date']
                # Handle different MongoDB date formats
                if 'T' in date_str:
                    # ISO format with T
                    date_obj_parsed = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                else:
                    # Just timestamp
                    date_obj_parsed = datetime.fromtimestamp(int(date_str) / 1000)
                return date_obj_parsed.isoformat()
            except (ValueError, KeyError, TypeError):
                return str(date_obj)
        else:
            return str(date_obj)
    else:
        # It's already a string or other type
        return str(date_obj)

@superadmin_bp.route('/dashboard', methods=['GET'])
@jwt_required()
def dashboard():
    """Super admin dashboard overview"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        # Debug logging
        current_app.logger.info(f"Dashboard access attempt - User ID: {current_user_id}")
        current_app.logger.info(f"User found: {user is not None}")
        if user:
            current_app.logger.info(f"User role: {user.get('role')}")
            current_app.logger.info(f"Allowed roles: {ALLOWED_ADMIN_ROLES}")
            current_app.logger.info(f"Role check: {user.get('role') in ALLOWED_ADMIN_ROLES}")
        
        # More permissive check for superadmin dashboard
        if not user:
            return jsonify({
                'success': False,
                'message': 'User not found'
            }), 403
        
        user_role = user.get('role')
        if user_role != 'superadmin':
            return jsonify({
                'success': False,
                'message': f'Access denied. Super admin privileges required. User role: {user_role}'
            }), 403
        
        total_users = mongo_db.users.count_documents({})
        total_students = mongo_db.users.count_documents({'role': 'student'})
        total_tests = mongo_db.tests.count_documents({})
        # Optionally, count admins (super, campus, course)
        total_admins = mongo_db.users.count_documents({'role': {'$in': [ROLES['SUPER_ADMIN'], ROLES['CAMPUS_ADMIN'], ROLES['COURSE_ADMIN']]}})
        # Optionally, count active courses
        total_courses = mongo_db.courses.count_documents({})

        dashboard_data = {
            'statistics': {
                'total_users': total_users,
                'total_students': total_students,
                'total_tests': total_tests,
                'total_admins': total_admins,
                'active_courses': total_courses
            }
        }
        
        return jsonify({
            'success': True,
            'message': 'Dashboard data retrieved successfully',
            'data': dashboard_data
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Failed to get dashboard data: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to get dashboard data: {str(e)}'
        }), 500

@superadmin_bp.route('/debug-roles', methods=['GET'])
@jwt_required()
def debug_roles():
    """Debug endpoint to check user roles"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        # Get all unique roles in the database
        all_roles = mongo_db.users.distinct('role')
        
        return jsonify({
            'success': True,
            'data': {
                'current_user': {
                    'id': str(current_user_id),
                    'role': user.get('role') if user else None,
                    'username': user.get('username') if user else None
                },
                'allowed_roles': list(ALLOWED_ADMIN_ROLES),
                'all_roles_in_db': all_roles,
                'role_constants': ROLES
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Debug error: {str(e)}'
        }), 500

@superadmin_bp.route('/users', methods=['POST'])
@jwt_required()
def create_user():
    """Create a new user"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') != 'superadmin':
            return jsonify({
                'success': False,
                'message': 'Access denied. Super admin privileges required.'
            }), 403
        
        data = request.get_json()
        
        if not all(key in data for key in ['username', 'email', 'password', 'role', 'name']):
            return jsonify({
                'success': False,
                'message': 'Missing required fields'
            }), 400
        
        # Check if username exists
        if mongo_db.find_user_by_username(data['username']):
            return jsonify({
                'success': False,
                'message': 'Username already exists'
            }), 400
        
        # Hash password
        password_hash = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt())
        
        # Create user
        user_data = {
            'username': data['username'],
            'email': data['email'],
            'password_hash': password_hash.decode('utf-8'),
            'role': data['role'],
            'name': data['name'],
            'mobile': data.get('mobile', ''),
            'is_active': True,
            'created_at': datetime.utcnow()
        }
        
        user_id = mongo_db.insert_user(user_data)
        
        return jsonify({
            'success': True,
            'message': 'User created successfully',
            'data': {'user_id': user_id}
        }), 201
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to create user: {str(e)}'
        }), 500

@superadmin_bp.route('/users', methods=['GET'])
@jwt_required()
def get_users():
    """Get all users"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') != 'superadmin':
            return jsonify({
                'success': False,
                'message': 'Access denied. Super admin privileges required.'
            }), 403
        
        users = list(mongo_db.users.find().sort('created_at', -1))
        
        users_data = []
        for user in users:
            users_data.append({
                'id': str(user['_id']),
                'username': user['username'],
                'email': user['email'],
                'name': user['name'],
                'role': user['role'],
                'is_active': user.get('is_active', True)
            })
        
        return jsonify({
            'success': True,
            'message': 'Users retrieved successfully',
            'data': users_data
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to get users: {str(e)}'
        }), 500

@superadmin_bp.route('/tests', methods=['POST'])
@jwt_required()
def create_test():
    """Create a new test"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') != 'superadmin':
            return jsonify({
                'success': False,
                'message': 'Access denied. Super admin privileges required.'
            }), 403
        
        data = request.get_json()
        required_fields = ['name', 'module_id', 'level_id', 'test_type']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'message': f'{field} is required'
                }), 400
        
        # Create test object
        test = Test(
            name=data['name'],
            module_id=data['module_id'],
            level_id=data['level_id'],
            created_by=current_user_id,
            test_type=data['test_type'],
            total_questions=data.get('total_questions', 0),
            time_limit=data.get('time_limit', 30),
            passing_score=data.get('passing_score', 70)
        )
        
        # Insert test
        test_id = mongo_db.insert_test(test.to_dict())
        
        return jsonify({
            'success': True,
            'message': 'Test created successfully',
            'data': {'test_id': test_id}
        }), 201
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to create test: {str(e)}'
        }), 500

@superadmin_bp.route('/tests', methods=['GET'])
@jwt_required()
def get_tests():
    """Get all tests with pagination"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') != 'superadmin':
            return jsonify({
                'success': False,
                'message': 'Access denied. Super admin privileges required.'
            }), 403
        
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        module_id = request.args.get('module_id')
        level_id = request.args.get('level_id')
        test_type = request.args.get('test_type')
        
        # Build query
        query = {}
        if module_id:
            query['module_id'] = ObjectId(module_id)
        if level_id:
            query['level_id'] = ObjectId(level_id)
        if test_type:
            query['test_type'] = test_type
        
        # Get total count
        total = mongo_db.tests.count_documents(query)
        
        # Get tests with pagination
        skip = (page - 1) * limit
        tests = list(mongo_db.tests.find(query).skip(skip).limit(limit).sort('created_at', -1))
        
        # Format response
        tests_data = []
        for test in tests:
            tests_data.append({
                'id': str(test['_id']),
                'name': test['name'],
                'module_id': str(test['module_id']),
                'level_id': str(test['level_id']),
                'test_type': test['test_type'],
                'status': test['status'],
                'total_questions': test['total_questions'],
                'time_limit': test['time_limit'],
                'passing_score': test['passing_score'],
                'created_at': test['created_at']
            })
        
        return jsonify({
            'success': True,
            'message': 'Tests retrieved successfully',
            'data': {
                'tests': tests_data,
                'pagination': {
                    'page': page,
                    'limit': limit,
                    'total': total,
                    'pages': (total + limit - 1) // limit
                }
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to get tests: {str(e)}'
        }), 500

@superadmin_bp.route('/online-exams', methods=['POST'])
@jwt_required()
def create_online_exam():
    """Create a new online exam"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') != 'superadmin':
            return jsonify({
                'success': False,
                'message': 'Access denied. Super admin privileges required.'
            }), 403
        
        data = request.get_json()
        required_fields = ['test_id', 'name', 'start_date', 'end_date', 'duration']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'message': f'{field} is required'
                }), 400
        
        # Create online exam object
        exam = OnlineExam(
            test_id=data['test_id'],
            name=data['name'],
            start_date=data['start_date'],
            end_date=data['end_date'],
            duration=data['duration'],
            campus_ids=data.get('campus_ids', []),
            course_ids=data.get('course_ids', []),
            batch_ids=data.get('batch_ids', []),
            created_by=current_user_id
        )
        
        # Insert exam
        exam_id = mongo_db.online_exams.insert_one(exam.to_dict()).inserted_id
        
        return jsonify({
            'success': True,
            'message': 'Online exam created successfully',
            'data': {'exam_id': str(exam_id)}
        }), 201
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to create online exam: {str(e)}'
        }), 500

@superadmin_bp.route('/student-practice-results', methods=['GET'])
@jwt_required()
def get_student_practice_results():
    """Get detailed practice results for all students"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        # Get query parameters
        module_filter = request.args.get('module')
        campus_filter = request.args.get('campus')
        student_filter = request.args.get('student')
        
        # Build match conditions
        match_conditions = {'test_type': 'practice'}
        if module_filter:
            match_conditions['module_id'] = module_filter
        
        # Get results from both collections
        all_results = []
        
        # Try to get from test_results collection
        try:
            current_app.logger.info(f"Checking test_results collection...")
            if hasattr(mongo_db, 'test_results'):
                current_app.logger.info(f"test_results collection exists")
                # Check if there are any practice results
                practice_count = mongo_db.test_results.count_documents({'test_type': 'practice'})
                current_app.logger.info(f"Found {practice_count} practice results in test_results")
                pipeline = [
                    {'$match': match_conditions},
                    {
                        '$lookup': {
                            'from': 'users',
                            'localField': 'student_id',
                            'foreignField': '_id',
                            'as': 'student_details'
                        }
                    },
                    {'$unwind': '$student_details'},
                    {
                        '$lookup': {
                            'from': 'tests',
                            'localField': 'test_id',
                            'foreignField': '_id',
                            'as': 'test_details'
                        }
                    },
                    {'$unwind': '$test_details'},
                    {
                        '$lookup': {
                            'from': 'students',
                            'localField': 'student_id',
                            'foreignField': 'user_id',
                            'as': 'student_profile'
                        }
                    },
                    {'$unwind': '$student_profile'},
                    {
                        '$lookup': {
                            'from': 'campuses',
                            'localField': 'student_profile.campus_id',
                            'foreignField': '_id',
                            'as': 'campus_details'
                        }
                    },
                    {'$unwind': '$campus_details'},
                    {
                        '$lookup': {
                            'from': 'courses',
                            'localField': 'student_profile.course_id',
                            'foreignField': '_id',
                            'as': 'course_details'
                        }
                    },
                    {'$unwind': '$course_details'},
                    {
                        '$lookup': {
                            'from': 'batches',
                            'localField': 'student_profile.batch_id',
                            'foreignField': '_id',
                            'as': 'batch_details'
                        }
                    },
                    {'$unwind': '$batch_details'},
                    {
                        '$project': {
                            '_id': 1,
                            'student_id': 1,
                            'student_name': '$student_details.name',
                            'student_email': '$student_details.email',
                            'campus_name': '$campus_details.name',
                            'course_name': '$course_details.name',
                            'batch_name': '$batch_details.name',
                            'test_name': '$test_details.name',
                            'module_name': '$test_details.module_id',
                            'test_type': 1,
                            'average_score': 1,
                            'score_percentage': 1,
                            'total_questions': 1,
                            'correct_answers': 1,
                            'submitted_at': 1,
                            'duration': 1,
                            'time_taken': 1,
                            'auto_submitted': 1,
                            'cheat_detected': 1
                        }
                    },
                    {'$sort': {'submitted_at': -1}}
                ]
                
                # Apply additional filters
                if campus_filter:
                    pipeline.insert(0, {'$match': {'campus_name': campus_filter}})
                if student_filter:
                    pipeline.insert(0, {'$match': {'$or': [
                        {'student_name': {'$regex': student_filter, '$options': 'i'}},
                        {'student_email': {'$regex': student_filter, '$options': 'i'}}
                    ]}})
                
                results = list(mongo_db.test_results.aggregate(pipeline))
                all_results.extend(results)
                current_app.logger.info(f"Found {len(results)} practice results in test_results collection")
            else:
                current_app.logger.warning("test_results collection not found")
        except Exception as e:
            current_app.logger.warning(f"Error aggregating from test_results: {e}")
        
        # Also get from student_test_attempts collection
        try:
            current_app.logger.info(f"Checking student_test_attempts collection...")
            if hasattr(mongo_db, 'student_test_attempts'):
                current_app.logger.info(f"student_test_attempts collection exists")
                # Check if there are any practice results
                practice_attempts_count = mongo_db.student_test_attempts.count_documents({'test_type': 'practice'})
                current_app.logger.info(f"Found {practice_attempts_count} practice results in student_test_attempts")
                pipeline = [
                    {'$match': match_conditions},
                    {
                        '$lookup': {
                            'from': 'users',
                            'localField': 'student_id',
                            'foreignField': '_id',
                            'as': 'student_details'
                        }
                    },
                    {'$unwind': '$student_details'},
                    {
                        '$lookup': {
                            'from': 'tests',
                            'localField': 'test_id',
                            'foreignField': '_id',
                            'as': 'test_details'
                        }
                    },
                    {'$unwind': '$test_details'},
                    {
                        '$lookup': {
                            'from': 'students',
                            'localField': 'student_id',
                            'foreignField': 'user_id',
                            'as': 'student_profile'
                        }
                    },
                    {'$unwind': '$student_profile'},
                    {
                        '$lookup': {
                            'from': 'campuses',
                            'localField': 'student_profile.campus_id',
                            'foreignField': '_id',
                            'as': 'campus_details'
                        }
                    },
                    {'$unwind': '$campus_details'},
                    {
                        '$lookup': {
                            'from': 'courses',
                            'localField': 'student_profile.course_id',
                            'foreignField': '_id',
                            'as': 'course_details'
                        }
                    },
                    {'$unwind': '$course_details'},
                    {
                        '$lookup': {
                            'from': 'batches',
                            'localField': 'student_profile.batch_id',
                            'foreignField': '_id',
                            'as': 'batch_details'
                        }
                    },
                    {'$unwind': '$batch_details'},
                    {
                        '$project': {
                            '_id': 1,
                            'student_id': 1,
                            'student_name': '$student_details.name',
                            'student_email': '$student_details.email',
                            'campus_name': '$campus_details.name',
                            'course_name': '$course_details.name',
                            'batch_name': '$batch_details.name',
                            'test_name': '$test_details.name',
                            'module_name': '$test_details.module_id',
                            'test_type': 1,
                            'average_score': 1,
                            'score_percentage': 1,
                            'total_questions': 1,
                            'correct_answers': 1,
                            'submitted_at': 1,
                            'duration': 1,
                            'time_taken': 1,
                            'auto_submitted': 1,
                            'cheat_detected': 1
                        }
                    },
                    {'$sort': {'submitted_at': -1}}
                ]
                
                # Apply additional filters
                if campus_filter:
                    pipeline.insert(0, {'$match': {'campus_name': campus_filter}})
                if student_filter:
                    pipeline.insert(0, {'$match': {'$or': [
                        {'student_name': {'$regex': student_filter, '$options': 'i'}},
                        {'student_email': {'$regex': student_filter, '$options': 'i'}}
                    ]}})
                
                attempt_results = list(mongo_db.student_test_attempts.aggregate(pipeline))
                all_results.extend(attempt_results)
                current_app.logger.info(f"Found {len(attempt_results)} practice results in student_test_attempts collection")
            else:
                current_app.logger.warning("student_test_attempts collection not found")
        except Exception as e:
            current_app.logger.warning(f"Error aggregating from student_test_attempts: {e}")
        
        # Remove duplicates based on test_id, student_id, and submitted_at
        seen = set()
        unique_results = []
        for result in all_results:
            key = (str(result.get('test_id')), str(result.get('student_id')), str(result.get('submitted_at')))
            if key not in seen:
                seen.add(key)
                unique_results.append(result)
        
        results = unique_results
        current_app.logger.info(f"Total unique practice results: {len(results)}")
        
        # Process results
        for result in results:
            result['_id'] = str(result['_id'])
            result['student_id'] = str(result['student_id'])
            result['module_name'] = MODULES.get(result.get('module_name', ''), result.get('module_name', 'Unknown'))
            if result.get('submitted_at'):
                result['submitted_at'] = safe_isoformat(result['submitted_at'])

        return jsonify({
            'success': True,
            'data': results
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching practice results: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to get student practice results: {str(e)}'
        }), 500

@superadmin_bp.route('/student-online-results', methods=['GET'])
@jwt_required()
def get_student_online_results():
    """Get detailed online exam results for all students"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        # Get query parameters
        module_filter = request.args.get('module')
        campus_filter = request.args.get('campus')
        student_filter = request.args.get('student')
        
        # Build match conditions - exclude practice tests
        match_conditions = {'test_type': {'$ne': 'practice'}}
        if module_filter:
            match_conditions['module_id'] = module_filter
        
        pipeline = [
            {'$match': match_conditions},
            {
                '$lookup': {
                    'from': 'users',
                    'localField': 'student_id',
                    'foreignField': '_id',
                    'as': 'student_details'
                }
            },
            {'$unwind': '$student_details'},
            {
                '$lookup': {
                    'from': 'tests',
                    'localField': 'test_id',
                    'foreignField': '_id',
                    'as': 'test_details'
                }
            },
            {'$unwind': '$test_details'},
            {
                '$lookup': {
                    'from': 'students',
                    'localField': 'student_id',
                    'foreignField': 'user_id',
                    'as': 'student_profile'
                }
            },
            {'$unwind': '$student_profile'},
            {
                '$lookup': {
                    'from': 'campuses',
                    'localField': 'student_profile.campus_id',
                    'foreignField': '_id',
                    'as': 'campus_details'
                }
            },
            {'$unwind': '$campus_details'},
            {
                '$lookup': {
                    'from': 'courses',
                    'localField': 'student_profile.course_id',
                    'foreignField': '_id',
                    'as': 'course_details'
                }
            },
            {'$unwind': '$course_details'},
            {
                '$lookup': {
                    'from': 'batches',
                    'localField': 'student_profile.batch_id',
                    'foreignField': '_id',
                    'as': 'batch_details'
                }
            },
            {'$unwind': '$batch_details'},
            {
                '$project': {
                    '_id': 1,
                    'student_id': 1,
                    'student_name': '$student_details.name',
                    'student_email': '$student_details.email',
                    'campus_name': '$campus_details.name',
                    'course_name': '$course_details.name',
                    'batch_name': '$batch_details.name',
                    'test_name': '$test_details.name',
                    'module_name': '$test_details.module_id',
                    'test_type': 1,
                    'average_score': 1,
                    'total_questions': 1,
                    'correct_answers': 1,
                    'submitted_at': 1,
                    'duration': 1,
                    'time_taken': 1,
                    'auto_submitted': 1,
                    'cheat_detected': 1
                }
            },
            {'$sort': {'submitted_at': -1}}
        ]
        
        # Apply additional filters
        if campus_filter:
            pipeline.insert(0, {'$match': {'campus_name': campus_filter}})
        if student_filter:
            pipeline.insert(0, {'$match': {'$or': [
                {'student_name': {'$regex': student_filter, '$options': 'i'}},
                {'student_email': {'$regex': student_filter, '$options': 'i'}}
            ]}})
        
        # Try all collections for online results
        all_results = []
        
        # Try student_test_attempts collection
        try:
            if hasattr(mongo_db, 'student_test_attempts'):
                attempt_results = list(mongo_db.student_test_attempts.aggregate(pipeline))
                all_results.extend(attempt_results)
                current_app.logger.info(f"Found {len(attempt_results)} online results in student_test_attempts")
        except Exception as e:
            current_app.logger.warning(f"Error aggregating from student_test_attempts: {e}")
        
        # Try test_results collection
        try:
            if hasattr(mongo_db, 'test_results'):
                test_results = list(mongo_db.test_results.aggregate(pipeline))
                all_results.extend(test_results)
                current_app.logger.info(f"Found {len(test_results)} online results in test_results")
        except Exception as e:
            current_app.logger.warning(f"Error aggregating from test_results: {e}")
        
        # Try student_test_assignments collection (where online exam results are stored)
        try:
            current_app.logger.info(f"Checking student_test_assignments collection...")
            if hasattr(mongo_db, 'student_test_assignments'):
                current_app.logger.info(f"student_test_assignments collection exists")
                # Check if there are any attempted assignments
                attempted_count = mongo_db.student_test_assignments.count_documents({'attempted': True})
                current_app.logger.info(f"Found {attempted_count} attempted assignments")
                # Create a different pipeline for student_test_assignments
                assignment_pipeline = [
                    {'$match': {'attempted': True}},  # Only get attempted assignments
                    {
                        '$lookup': {
                            'from': 'tests',
                            'localField': 'test_id',
                            'foreignField': '_id',
                            'as': 'test_details'
                        }
                    },
                    {'$unwind': '$test_details'},
                    {
                        '$lookup': {
                            'from': 'users',
                            'localField': 'student_id',
                            'foreignField': '_id',
                            'as': 'student_details'
                        }
                    },
                    {'$unwind': '$student_details'},
                    {
                        '$lookup': {
                            'from': 'students',
                            'localField': 'student_id',
                            'foreignField': 'user_id',
                            'as': 'student_profile'
                        }
                    },
                    {'$unwind': '$student_profile'},
                    {
                        '$lookup': {
                            'from': 'campuses',
                            'localField': 'student_profile.campus_id',
                            'foreignField': '_id',
                            'as': 'campus_details'
                        }
                    },
                    {'$unwind': '$campus_details'},
                    {
                        '$lookup': {
                            'from': 'courses',
                            'localField': 'student_profile.course_id',
                            'foreignField': '_id',
                            'as': 'course_details'
                        }
                    },
                    {'$unwind': '$course_details'},
                    {
                        '$lookup': {
                            'from': 'batches',
                            'localField': 'student_profile.batch_id',
                            'foreignField': '_id',
                            'as': 'batch_details'
                        }
                    },
                    {'$unwind': '$batch_details'},
                    {
                        '$project': {
                            '_id': 1,
                            'student_id': 1,
                            'student_name': '$student_details.name',
                            'student_email': '$student_details.email',
                            'campus_name': '$campus_details.name',
                            'course_name': '$course_details.name',
                            'batch_name': '$batch_details.name',
                            'test_name': '$test_details.name',
                            'module_name': '$test_details.module_id',
                            'test_type': '$test_details.test_type',
                            'average_score': '$percentage',
                            'total_questions': {'$size': '$questions'},
                            'correct_answers': '$score',
                            'submitted_at': '$completed_at',
                            'duration': '$test_details.duration',
                            'time_taken': 0,
                            'auto_submitted': False,
                            'cheat_detected': False
                        }
                    },
                    {'$sort': {'submitted_at': -1}}
                ]
                
                # Apply additional filters to assignment pipeline
                if campus_filter:
                    assignment_pipeline.insert(0, {'$match': {'campus_name': campus_filter}})
                if student_filter:
                    assignment_pipeline.insert(0, {'$match': {'$or': [
                        {'student_name': {'$regex': student_filter, '$options': 'i'}},
                        {'student_email': {'$regex': student_filter, '$options': 'i'}}
                    ]}})
                if module_filter:
                    assignment_pipeline.insert(0, {'$match': {'module_name': module_filter}})
                
                assignment_results = list(mongo_db.student_test_assignments.aggregate(assignment_pipeline))
                all_results.extend(assignment_results)
                current_app.logger.info(f"Found {len(assignment_results)} online results in student_test_assignments")
        except Exception as e:
            current_app.logger.warning(f"Error aggregating from student_test_assignments: {e}")
        
        # Remove duplicates based on test_id, student_id, and submitted_at
        seen = set()
        results = []
        for result in all_results:
            key = (str(result.get('test_id')), str(result.get('student_id')), str(result.get('submitted_at')))
            if key not in seen:
                seen.add(key)
                results.append(result)
        
        # Process results
        for result in results:
            result['_id'] = str(result['_id'])
            result['student_id'] = str(result['student_id'])
            result['module_name'] = MODULES.get(result.get('module_name', ''), result.get('module_name', 'Unknown'))
            if result.get('submitted_at'):
                result['submitted_at'] = safe_isoformat(result['submitted_at'])

        return jsonify({
            'success': True,
            'data': results
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching online results: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to get student online results: {str(e)}'
        }), 500

@superadmin_bp.route('/grammar-analytics', methods=['GET'])
@jwt_required()
def get_grammar_analytics():
    """Get detailed grammar practice analytics"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        pipeline = [
            {
                '$match': {
                    'module_id': 'GRAMMAR',
                    'test_type': 'practice'
                }
            },
            {
                '$lookup': {
                    'from': 'tests',
                    'localField': 'test_id',
                    'foreignField': '_id',
                    'as': 'test_details'
                }
            },
            {'$unwind': '$test_details'},
            {
                '$group': {
                    '_id': '$subcategory',
                    'subcategory_name': {'$first': '$subcategory'},
                    'total_attempts': {'$sum': 1},
                    'total_students': {'$addToSet': '$student_id'},
                    'average_score': {'$avg': '$average_score'},
                    'highest_score': {'$max': '$average_score'},
                    'lowest_score': {'$min': '$average_score'},
                    'total_questions': {'$sum': '$total_questions'},
                    'total_correct': {'$sum': '$correct_answers'},
                    'completion_rate': {
                        '$avg': {
                            '$cond': [
                                {'$gte': ['$average_score', 60]},
                                1,
                                0
                            ]
                        }
                    }
                }
            },
            {'$sort': {'subcategory_name': 1}}
        ]
        
        results = list(mongo_db.db.test_results.aggregate(pipeline))
        
        # Process results
        for result in results:
            result['subcategory_display_name'] = GRAMMAR_CATEGORIES.get(result['subcategory_name'], result['subcategory_name'])
            result['total_students'] = len(result['total_students'])
            result['accuracy'] = (result['total_correct'] / result['total_questions'] * 100) if result['total_questions'] > 0 else 0
            result['completion_rate'] = result['completion_rate'] * 100
        
        return jsonify({
            'success': True,
            'data': results
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to get grammar analytics: {str(e)}'
        }), 500

@superadmin_bp.route('/vocabulary-analytics', methods=['GET'])
@jwt_required()
def get_vocabulary_analytics():
    """Get detailed vocabulary practice analytics"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        pipeline = [
            {
                '$match': {
                    'module_id': 'VOCABULARY',
                    'test_type': 'practice'
                }
            },
            {
                '$lookup': {
                    'from': 'tests',
                    'localField': 'test_id',
                    'foreignField': '_id',
                    'as': 'test_details'
                }
            },
            {'$unwind': '$test_details'},
            {
                '$group': {
                    '_id': '$test_details.level_id',
                    'level_name': {'$first': '$test_details.level_id'},
                    'total_attempts': {'$sum': 1},
                    'total_students': {'$addToSet': '$student_id'},
                    'average_score': {'$avg': '$average_score'},
                    'highest_score': {'$max': '$average_score'},
                    'lowest_score': {'$min': '$average_score'},
                    'total_questions': {'$sum': '$total_questions'},
                    'total_correct': {'$sum': '$correct_answers'},
                    'completion_rate': {
                        '$avg': {
                            '$cond': [
                                {'$gte': ['$average_score', 60]},
                                1,
                                0
                            ]
                        }
                    }
                }
            },
            {'$sort': {'level_name': 1}}
        ]
        
        results = list(mongo_db.db.test_results.aggregate(pipeline))
        
        # Process results
        for result in results:
            result['level_display_name'] = LEVELS.get(result['level_name'], {}).get('name', result['level_name'])
            result['total_students'] = len(result['total_students'])
            result['accuracy'] = (result['total_correct'] / result['total_questions'] * 100) if result['total_questions'] > 0 else 0
            result['completion_rate'] = result['completion_rate'] * 100
        
        return jsonify({
            'success': True,
            'data': results
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to get vocabulary analytics: {str(e)}'
        }), 500

@superadmin_bp.route('/practice-overview', methods=['GET'])
@jwt_required()
def get_practice_overview():
    """Get overview of all practice module usage"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        # Overall statistics
        total_practice_tests = mongo_db.db.test_results.count_documents({'test_type': 'practice'})
        total_students_practicing = len(mongo_db.db.test_results.distinct('student_id', {'test_type': 'practice'}))
        
        # Module-wise statistics
        pipeline = [
            {'$match': {'test_type': 'practice'}},
            {
                '$lookup': {
                    'from': 'tests',
                    'localField': 'test_id',
                    'foreignField': '_id',
                    'as': 'test_details'
                }
            },
            {'$unwind': '$test_details'},
            {
                '$group': {
                    '_id': '$test_details.module_id',
                    'module_name': {'$first': '$test_details.module_id'},
                    'total_attempts': {'$sum': 1},
                    'unique_students': {'$addToSet': '$student_id'},
                    'average_score': {'$avg': '$average_score'},
                    'completion_rate': {
                        '$avg': {
                            '$cond': [
                                {'$gte': ['$average_score', 60]},
                                1,
                                0
                            ]
                        }
                    }
                }
            }
        ]
        
        module_stats = list(mongo_db.db.test_results.aggregate(pipeline))
        
        # Process module statistics
        for stat in module_stats:
            stat['module_display_name'] = MODULES.get(stat['module_name'], 'Unknown')
            stat['unique_students'] = len(stat['unique_students'])
            stat['completion_rate'] = stat['completion_rate'] * 100
        
        overview = {
            'total_practice_tests': total_practice_tests,
            'total_students_practicing': total_students_practicing,
            'modules': module_stats
        }
        
        return jsonify({
            'success': True,
            'data': overview
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Failed to get practice overview: {str(e)}'
        }), 500

@superadmin_bp.route('/student-assigned-modules', methods=['GET'])
@jwt_required()
def get_student_assigned_modules():
    """Get all tests assigned to a student, with attempt data if available, or status 'Pending' if not attempted. Each test is a separate row."""
    try:
        student_email = request.args.get('student')
        batch_id = request.args.get('batch')
        if not student_email or not batch_id:
            return jsonify({'success': False, 'message': 'Missing student email or batch id'}), 400

        # Find the student user and student profile
        user = mongo_db.users.find_one({'email': student_email})
        student = mongo_db.students.find_one({'user_id': user['_id']}) if user else None
        if not user or not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        # Only include tests where the student is explicitly assigned (assigned_student_ids)
        assigned_tests = list(mongo_db.tests.find({
            'test_type': 'practice',
            'assigned_student_ids': {'$in': [student['_id']]}
        }))

        # Get all attempts/results for this student
        attempts = list(mongo_db.db.test_results.find({
            'student_id': user['_id'],
            'test_type': 'practice'
        }))
        # Map attempts by test_id
        attempts_by_test = {}
        for att in attempts:
            tid = att.get('test_id')
            if tid not in attempts_by_test:
                attempts_by_test[tid] = []
            attempts_by_test[tid].append(att)

        # Merge: for each assigned test, show attempt data if exists, else Pending
        result = []
        for idx, test in enumerate(assigned_tests):
            tid = test['_id']
            test_name = test.get('name') or f"Test {idx+1}"
            test_info = {
                'test_id': str(tid),
                'test_name': test_name,
                'module_id': str(test['module_id']),
                'module_display_name': MODULES.get(test['module_id'], 'Unknown'),
                'level_id': test.get('level_id'),
                'level_display_name': LEVELS.get(test.get('level_id'), {}).get('name', 'Unknown'),
                'assigned': True
            }
            if tid in attempts_by_test:
                # There are attempts for this test
                test_attempts = attempts_by_test[tid]
                scores = [att.get('average_score', 0) for att in test_attempts]
                test_info['status'] = 'completed'
                test_info['attempts'] = [
                    {
                        'score': att.get('average_score', 0),
                        'correct_answers': att.get('correct_answers', 0),
                        'total_questions': att.get('total_questions', 0),
                        'submitted_at': safe_isoformat(att.get('submitted_at')) if att.get('submitted_at') else None,
                        'result_id': str(att.get('_id'))
                    }
                    for att in test_attempts
                ]
                test_info['total_attempts'] = len(test_attempts)
                test_info['best_score'] = max(scores) if scores else 0
                test_info['avg_score'] = sum(scores)/len(scores) if scores else 0
            else:
                test_info['status'] = 'pending'
                test_info['attempts'] = []
                test_info['total_attempts'] = 0
                test_info['best_score'] = 0
                test_info['avg_score'] = 0
            result.append(test_info)

        return jsonify({'success': True, 'data': result}), 200
    except Exception as e:
        return jsonify({'success': False, 'message': f'Failed to get assigned modules: {str(e)}'}), 500

@superadmin_bp.route('/student-modules', methods=['GET'])
@jwt_required()
def get_student_modules():
    """Get all available practice modules for a student (by email or roll number, admin roles only)."""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        allowed_roles = ['super_admin', 'campus_admin', 'course_admin']
        if not user or user.get('role') not in allowed_roles:
            return jsonify({'success': False, 'message': 'Access denied. Admin privileges required.'}), 403
        roll_number = request.args.get('roll_number')
        student_email = request.args.get('student_email') or request.headers.get('X-Student-Email')
        student_user = None
        if roll_number:
            student = mongo_db.students.find_one({'roll_number': roll_number})
            if student:
                student_user = mongo_db.users.find_one({'_id': student['user_id']})
        if not student_user and student_email:
            student_user = mongo_db.users.find_one({'email': student_email})
        if not student_user:
            return jsonify({'success': False, 'message': 'Student not found.'}), 404
        # Use the same logic as /student/modules
        priority_order = ['GRAMMAR', 'VOCABULARY']
        all_modules = [{'id': key, 'name': value} for key, value in MODULES.items()]
        priority_modules = [m for m in all_modules if m['id'] in priority_order]
        other_modules = [m for m in all_modules if m['id'] not in priority_order]
        priority_modules.sort(key=lambda m: priority_order.index(m['id']))
        final_module_list = []
        for m in priority_modules:
            final_module_list.append({**m, 'locked': False})
        for m in other_modules:
            final_module_list.append({**m, 'locked': True})
        return jsonify({'success': True, 'data': final_module_list})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Could not load modules: {str(e)}'}), 500

@superadmin_bp.route('/batch/<batch_id>/course/<course_id>/module/upload', methods=['POST'])
@jwt_required()
def upload_module_to_batch_course(batch_id, course_id):
    # Only admin roles
    current_user_id = get_jwt_identity()
    user = mongo_db.find_user_by_id(current_user_id)
    if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
        return jsonify({'success': False, 'message': 'Access denied. Admin privileges required.'}), 403
    # Find or create batch_course_instance
    instance_id = mongo_db.find_or_create_batch_course_instance(ObjectId(batch_id), ObjectId(course_id))
    # Get module data from request
    module_data = request.get_json()
    module_data['batch_course_instance_id'] = instance_id
    # Insert module (assume modules collection)
    result = mongo_db.modules.insert_one(module_data)
    return jsonify({'success': True, 'message': 'Module uploaded', 'module_id': str(result.inserted_id), 'batch_course_instance_id': str(instance_id)}), 201

# Update result fetching endpoint to filter by batch_course_instance_id
@superadmin_bp.route('/batch-course/<instance_id>/module/results', methods=['GET'])
@jwt_required()
def get_module_results_by_instance(instance_id):
    # Only admin roles
    current_user_id = get_jwt_identity()
    user = mongo_db.find_user_by_id(current_user_id)
    if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
        return jsonify({'success': False, 'message': 'Access denied. Admin privileges required.'}), 403
    # Fetch results for this batch_course_instance_id
    results = list(mongo_db.db.test_results.find({'batch_course_instance_id': ObjectId(instance_id)}))
    for r in results:
        r['_id'] = str(r['_id'])
    return jsonify({'success': True, 'data': results}), 200

@superadmin_bp.route('/superadmin/batch-course-instance', methods=['POST'])
@jwt_required()
def get_or_create_batch_course_instance():
    data = request.get_json()
    batch_id = data.get('batch_id')
    course_id = data.get('course_id')
    if not batch_id or not course_id:
        return jsonify({'success': False, 'message': 'Missing batch_id or course_id'}), 400
    instance_id = mongo_db.find_or_create_batch_course_instance(ObjectId(batch_id), ObjectId(course_id))
    return jsonify({'success': True, 'instance_id': str(instance_id)}), 200

@superadmin_bp.route('/writing-upload', methods=['POST'])
@jwt_required()
@require_superadmin
def writing_upload():
    """Upload writing paragraphs with validation"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not file.filename.endswith('.csv'):
            return jsonify({'success': False, 'message': 'Please upload a CSV file'}), 400
        
        # Read CSV content
        csv_content = file.read().decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(csv_content))
        
        # Validate required columns
        required_columns = ['level', 'topic', 'paragraph', 'instructions']
        headers = csv_reader.fieldnames
        
        if not headers:
            return jsonify({'success': False, 'message': 'CSV file is empty'}), 400
        
        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            return jsonify({
                'success': False, 
                'message': f'Missing required columns: {", ".join(missing_columns)}'
            }), 400
        
        # Process and validate paragraphs
        paragraphs = []
        errors = []
        valid_levels = ['Beginner', 'Intermediate', 'Advanced']
        
        for row_num, row in enumerate(csv_reader, start=2):  # Start from 2 because row 1 is headers
            try:
                level = row['level'].strip()
                topic = row['topic'].strip()
                paragraph = row['paragraph'].strip()
                instructions = row.get('instructions', '').strip()
                
                # Validate level
                if level not in valid_levels:
                    errors.append(f"Row {row_num}: Invalid level '{level}'. Must be one of {valid_levels}")
                    continue
                
                # Validate paragraph content
                if not paragraph:
                    errors.append(f"Row {row_num}: Paragraph is empty")
                    continue
                
                # Validate character count
                char_count = len(paragraph)
                if char_count < WRITING_CONFIG['MIN_CHARACTERS']:
                    errors.append(f"Row {row_num}: Character count ({char_count}) is below minimum ({WRITING_CONFIG['MIN_CHARACTERS']})")
                    continue
                if char_count > WRITING_CONFIG['MAX_CHARACTERS']:
                    errors.append(f"Row {row_num}: Character count ({char_count}) exceeds maximum ({WRITING_CONFIG['MAX_CHARACTERS']})")
                    continue
                
                # Validate word count
                word_count = len(paragraph.split())
                if word_count < WRITING_CONFIG['MIN_WORDS']:
                    errors.append(f"Row {row_num}: Word count ({word_count}) is below minimum ({WRITING_CONFIG['MIN_WORDS']})")
                    continue
                if word_count > WRITING_CONFIG['MAX_WORDS']:
                    errors.append(f"Row {row_num}: Word count ({word_count}) exceeds maximum ({WRITING_CONFIG['MAX_WORDS']})")
                    continue
                
                # Validate sentence count
                sentence_count = len([s for s in paragraph.split('.') if s.strip()])
                if sentence_count < WRITING_CONFIG['MIN_SENTENCES']:
                    errors.append(f"Row {row_num}: Sentence count ({sentence_count}) is below minimum ({WRITING_CONFIG['MIN_SENTENCES']})")
                    continue
                if sentence_count > WRITING_CONFIG['MAX_SENTENCES']:
                    errors.append(f"Row {row_num}: Sentence count ({sentence_count}) exceeds maximum ({WRITING_CONFIG['MAX_SENTENCES']})")
                    continue
                
                # Store paragraph data
                paragraph_data = {
                    'module_id': 'WRITING',
                    'level_id': f'WRITING_{level.upper()}',
                    'level': level,
                    'topic': topic,
                    'paragraph': paragraph,
                    'instructions': instructions,
                    'character_count': char_count,
                    'word_count': word_count,
                    'sentence_count': sentence_count,
                    'question_type': 'paragraph',
                    'created_by': current_user_id,
                    'created_at': datetime.utcnow()
                }
                
                paragraphs.append(paragraph_data)
                
            except Exception as e:
                errors.append(f"Row {row_num}: Error processing row - {str(e)}")
        
        if errors:
            return jsonify({
                'success': False,
                'message': f'Validation failed with {len(errors)} errors',
                'errors': errors
            }), 400
        
        if not paragraphs:
            return jsonify({'success': False, 'message': 'No valid paragraphs found'}), 400
        
        # Insert paragraphs into database
        inserted_count = 0
        for paragraph_data in paragraphs:
            try:
                mongo_db.question_bank.insert_one(paragraph_data)
                inserted_count += 1
            except Exception as e:
                errors.append(f"Failed to insert paragraph '{paragraph_data['topic']}': {str(e)}")
        
        if errors:
            return jsonify({
                'success': bool(inserted_count),
                'message': f'Upload completed with {len(errors)} errors',
                'data': {
                    'inserted_count': inserted_count,
                    'total_count': len(paragraphs)
                },
                'errors': errors
            }), 207 if inserted_count > 0 else 500
        
        return jsonify({
            'success': True,
            'message': f'Successfully uploaded {inserted_count} writing paragraphs',
            'data': {
                'inserted_count': inserted_count,
                'total_count': len(paragraphs)
            }
        }), 201
        
    except Exception as e:
        current_app.logger.error(f"Error in writing upload: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Upload failed: {str(e)}'
        }), 500

@superadmin_bp.route('/sentence-upload', methods=['POST'])
@jwt_required()
def sentence_upload():
    """Upload sentences for listening and speaking modules with audio support"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        module_id = request.form.get('module_id')
        level_id = request.form.get('level_id')
        level = request.form.get('level')
        
        # Audio file handling for Listening module
        audio_file = request.files.get('audio_file')
        audio_config = request.form.get('audio_config')
        transcript_validation = request.form.get('transcript_validation')
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not file.filename.endswith(('.csv', '.txt')):
            return jsonify({'success': False, 'message': 'Please upload a CSV or TXT file'}), 400
        
        if not module_id or not level_id or not level:
            return jsonify({'success': False, 'message': 'Module ID, Level ID, and Level are required'}), 400
        
        # Validate module and level
        valid_modules = ['LISTENING', 'SPEAKING']
        valid_levels = ['Beginner', 'Intermediate', 'Advanced']
        
        if module_id not in valid_modules:
            return jsonify({'success': False, 'message': f'Invalid module. Must be one of {valid_modules}'}), 400
        
        if level not in valid_levels:
            return jsonify({'success': False, 'message': f'Invalid level. Must be one of {valid_levels}'}), 400
        
        # For Listening module, require audio file
        if module_id == 'LISTENING' and not audio_file:
            return jsonify({'success': False, 'message': 'Audio file is required for Listening module'}), 400
        
        # Process audio file for Listening module
        audio_url = None
        if module_id == 'LISTENING' and audio_file:
            try:
                # Validate audio file
                allowed_audio_types = ['audio/mp3', 'audio/wav', 'audio/m4a', 'audio/ogg', 'audio/mpeg']
                if audio_file.content_type not in allowed_audio_types:
                    return jsonify({'success': False, 'message': 'Invalid audio file type. Please upload MP3, WAV, M4A, or OGG'}), 400
                
                # Upload audio to S3
                import uuid
                audio_filename = f"listening_audio/{uuid.uuid4()}_{audio_file.filename}"
                
                # Upload to S3
                current_s3_client = get_s3_client_safe()
                if current_s3_client is None:
                    return jsonify({'success': False, 'message': 'S3 client not available for audio upload. Please check AWS configuration.'}), 500
                
                current_s3_client.upload_fileobj(
                    audio_file,
                    S3_BUCKET_NAME,
                    audio_filename,
                    ExtraArgs={'ContentType': audio_file.content_type}
                )
                
                audio_url = f"https://{S3_BUCKET_NAME}.s3.amazonaws.com/{audio_filename}"
                
            except Exception as e:
                current_app.logger.error(f"Error uploading audio file: {str(e)}")
                return jsonify({'success': False, 'message': f'Failed to upload audio file: {str(e)}'}), 500
        
        # Read file content
        content = file.read().decode('utf-8')
        
        # Parse sentences based on file type
        sentences = []
        if file.filename.endswith('.csv'):
            # Parse CSV content
            csv_reader = csv.reader(io.StringIO(content))
            for row in csv_reader:
                if row and row[0].strip():
                    sentences.extend([s.strip() for s in row[0].split('.') if s.strip()])
        else:
            # Parse TXT content
            sentences = [s.strip() for s in content.split('\n') if s.strip()]
        
        # Validate sentences
        valid_sentences = []
        errors = []
        
        for i, sentence in enumerate(sentences, start=1):
            try:
                # Basic validation
                if not sentence:
                    continue
                
                if len(sentence) < 10:
                    errors.append(f"Sentence {i}: Too short (minimum 10 characters)")
                    continue
                
                if len(sentence) > 200:
                    errors.append(f"Sentence {i}: Too long (maximum 200 characters)")
                    continue
                
                # Check punctuation
                if not sentence.rstrip().endswith(('.', '!', '?')):
                    errors.append(f"Sentence {i}: Must end with proper punctuation (.!?)")
                    continue
                
                valid_sentences.append(sentence)
                
            except Exception as e:
                errors.append(f"Sentence {i}: Error processing - {str(e)}")
        
        if errors:
            return jsonify({
                'success': False,
                'message': f'Validation failed with {len(errors)} errors',
                'errors': errors[:10]  # Limit to first 10 errors
            }), 400
        
        if not valid_sentences:
            return jsonify({'success': False, 'message': 'No valid sentences found'}), 400
        
        # Parse configuration
        parsed_audio_config = {}
        parsed_transcript_validation = {}
        
        if audio_config:
            try:
                parsed_audio_config = json.loads(audio_config)
            except:
                parsed_audio_config = {'speed': 1.0, 'accent': 'en-US', 'volume': 1.0}
        
        if transcript_validation:
            try:
                parsed_transcript_validation = json.loads(transcript_validation)
            except:
                parsed_transcript_validation = {'enabled': True, 'tolerance': 0.8, 'checkMismatchedWords': True, 'allowPartialMatches': True}
        
        # Insert sentences into database
        inserted_count = 0
        for sentence in valid_sentences:
            try:
                sentence_data = {
                    'module_id': module_id,
                    'level_id': level_id,
                    'level': level,
                    'sentence': sentence,
                    'question_type': 'sentence',
                    'created_by': current_user_id,
                    'created_at': datetime.utcnow()
                }
                
                # Add audio-specific data for Listening module
                if module_id == 'LISTENING':
                    sentence_data.update({
                        'audio_url': audio_url,
                        'audio_config': parsed_audio_config,
                        'transcript_validation': parsed_transcript_validation,
                        'has_audio': True
                    })
                
                # Add transcript validation for Speaking module
                if module_id == 'SPEAKING':
                    sentence_data.update({
                        'transcript_validation': parsed_transcript_validation,
                        'question_type': 'speaking'
                    })
                
                mongo_db.question_bank.insert_one(sentence_data)
                inserted_count += 1
                
            except Exception as e:
                errors.append(f"Failed to insert sentence '{sentence[:50]}...': {str(e)}")
        
        if errors:
            return jsonify({
                'success': bool(inserted_count),
                'message': f'Upload completed with {len(errors)} errors',
                'data': {
                    'inserted_count': inserted_count,
                    'total_count': len(valid_sentences),
                    'audio_url': audio_url if module_id == 'LISTENING' else None
                },
                'errors': errors[:10]
            }), 207 if inserted_count > 0 else 500
        
        return jsonify({
            'success': True,
            'message': f'Successfully uploaded {inserted_count} sentences for {module_id} - {level}',
            'data': {
                'inserted_count': inserted_count,
                'total_count': len(valid_sentences),
                'module_id': module_id,
                'level': level,
                'audio_url': audio_url if module_id == 'LISTENING' else None,
                'audio_config': parsed_audio_config if module_id == 'LISTENING' else None,
                'transcript_validation': parsed_transcript_validation if module_id == 'LISTENING' else None
            }
        }), 201
        
    except Exception as e:
        current_app.logger.error(f"Error in sentence upload: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Upload failed: {str(e)}'
        }), 500

@superadmin_bp.route('/student-attempts/<student_id>/<test_id>', methods=['GET'])
@jwt_required()
def get_student_attempts(student_id, test_id):
    """Get detailed attempts for a specific student and test"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user:
            return jsonify({
                'success': False,
                'message': 'Access denied. Authentication required.'
            }), 401
        
        # Convert string IDs to ObjectId
        student_object_id = ObjectId(student_id)
        test_object_id = ObjectId(test_id)
        
        # Get the student attempt directly from student_test_attempts
        attempt = mongo_db.student_test_attempts.find_one({
            'student_id': student_object_id,
            'test_id': test_object_id,
            'status': 'completed'
        })
        
        if not attempt:
            return jsonify({
                'success': True,
                'data': [],
                'message': 'No completed attempts found for this student and test'
            })
        
        # Get test details
        test = mongo_db.tests.find_one({'_id': test_object_id})
        if not test:
            return jsonify({
                'success': False,
                'message': 'Test not found'
            }), 404
        
        # Get student details
        student = mongo_db.students.find_one({'user_id': student_object_id})
        if not student:
            return jsonify({
                'success': False,
                'message': 'Student not found'
            }), 404
        
        # Get campus, course, and batch details
        campus = mongo_db.campuses.find_one({'_id': student.get('campus_id')}) if student.get('campus_id') else None
        course = mongo_db.courses.find_one({'_id': student.get('course_id')}) if student.get('course_id') else None
        batch = mongo_db.batches.find_one({'_id': student.get('batch_id')}) if student.get('batch_id') else None
        
        # Transform detailed_results to match frontend expectations
        detailed_results = attempt.get('detailed_results', [])
        transformed_results = []
        for result in detailed_results:
            transformed_result = {
                'question_index': result.get('question_index', 0),
                'question_id': result.get('question_id', ''),
                'question_text': result.get('question', 'Question not available'),
                'student_answer': result.get('student_answer', 'No answer provided'),
                'selected_answer': result.get('student_answer', 'No answer provided'),
                'correct_answer_letter': result.get('correct_answer_letter', ''),
                'correct_answer_text': result.get('correct_answer_text', ''),
                'is_correct': result.get('is_correct', False),
                'score': result.get('score', 0)
            }
            transformed_results.append(transformed_result)
        
        # Format the response data
        response_data = {
            'attempt_id': str(attempt['_id']),
            'student_id': str(attempt['student_id']),
            'test_id': str(attempt['test_id']),
            'test_name': test.get('name', 'Unknown Test'),
            'student_name': student.get('name', 'Unknown Student'),
            'student_email': student.get('email', ''),
            'campus_name': campus.get('name', 'Unknown Campus') if campus else 'Unknown Campus',
            'course_name': course.get('name', 'Unknown Course') if course else 'Unknown Course',
            'batch_name': batch.get('name', 'Unknown Batch') if batch else 'Unknown Batch',
            'roll_number': student.get('roll_number', ''),
            'mobile_number': student.get('mobile_number', ''),
            'score': attempt.get('score', 0),
            'correct_answers': attempt.get('correct_answers', 0),
            'total_questions': attempt.get('total_questions', 0),
            'percentage': attempt.get('percentage', 0),
            'score_percentage': attempt.get('percentage', 0),  # Add this for frontend compatibility
            'duration_seconds': attempt.get('duration_seconds', 0),
            'time_taken_ms': attempt.get('time_taken_ms', 0),
            'time_taken': attempt.get('duration_seconds', 0) / 60,  # Convert to minutes for frontend
            'start_time': safe_isoformat(attempt.get('start_time')),
            'end_time': safe_isoformat(attempt.get('end_time')),
            'submitted_at': safe_isoformat(attempt.get('submitted_at')),
            'status': attempt.get('status', 'unknown'),
            'test_type': attempt.get('test_type', 'unknown'),
            'detailed_results': transformed_results,
            'answers': attempt.get('answers', {})
        }
        
        # Format duration
        duration_seconds = response_data.get('duration_seconds', 0)
        time_taken_ms = response_data.get('time_taken_ms', 0)
        
        if duration_seconds > 0:
            minutes = int(duration_seconds // 60)
            seconds = int(duration_seconds % 60)
            response_data['formatted_duration'] = f"{minutes}m {seconds}s"
        elif time_taken_ms > 0:
            duration_seconds = time_taken_ms / 1000
            minutes = int(duration_seconds // 60)
            seconds = int(duration_seconds % 60)
            response_data['formatted_duration'] = f"{minutes}m {seconds}s"
        else:
            response_data['formatted_duration'] = "0m 0s"
        
        return jsonify({
            'success': True,
            'data': [response_data],
            'message': f'Found attempt details for student {student_id} and test {test_id}'
        })
        
    except Exception as e:
        current_app.logger.error(f"Error in get_student_attempts: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Failed to fetch student attempts',
            'error': str(e)
        }), 500

@superadmin_bp.route('/export-test-attempts/<test_id>', methods=['GET'])
@jwt_required()
def export_test_attempts(test_id):
    """Export test attempts for a specific test as Excel"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user:
            return jsonify({
                'success': False,
                'message': 'Access denied. Authentication required.'
            }), 401
        
        # Get all attempts for this test
        attempts = list(mongo_db.student_test_attempts.find({
            'test_id': ObjectId(test_id),
            'test_type': 'online'
        }))
        
        if not attempts:
            return jsonify({
                'success': False,
                'message': 'No attempts found for this test'
            }), 404
        
        # Get test details
        test = mongo_db.tests.find_one({'_id': ObjectId(test_id)})
        if not test:
            return jsonify({
                'success': False,
                'message': 'Test not found'
            }), 404
        
        # Process attempts and get student details
        excel_data = []
        for attempt in attempts:
            student_id = str(attempt['student_id'])
            
            # Get student details
            student = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
            if not student:
                continue
            
            # Get campus, course, and batch details
            campus = mongo_db.campuses.find_one({'_id': student.get('campus_id')}) if student.get('campus_id') else None
            course = mongo_db.courses.find_one({'_id': student.get('course_id')}) if student.get('course_id') else None
            batch = mongo_db.batches.find_one({'_id': student.get('batch_id')}) if student.get('batch_id') else None
            
            # Calculate score - ensure we get clean numeric values
            total_questions = attempt.get('total_questions', 0)
            correct_answers = attempt.get('correct_answers', 0)
            
            # Handle MongoDB number types
            if isinstance(total_questions, dict):
                total_questions = total_questions.get('numberInt', total_questions.get('numberDouble', 0))
            if isinstance(correct_answers, dict):
                correct_answers = correct_answers.get('numberInt', correct_answers.get('numberDouble', 0))
            
            # Convert to integers
            total_questions = int(total_questions) if total_questions else 0
            correct_answers = int(correct_answers) if correct_answers else 0
            
            # Calculate score percentage
            score = 0
            if total_questions > 0:
                score = (correct_answers / total_questions) * 100
            
            # Format dates
            submitted_at = attempt.get('submitted_at', attempt.get('created_at'))
            if submitted_at:
                if isinstance(submitted_at, dict):
                    # Handle MongoDB date objects
                    submitted_at = safe_isoformat(submitted_at.get('$date', submitted_at))
                else:
                    submitted_at = safe_isoformat(submitted_at)
            else:
                submitted_at = ''
            
            # Get clean numeric values for duration
            duration_seconds = attempt.get('duration_seconds', 0)
            time_taken_ms = attempt.get('time_taken_ms', 0)
            
            if isinstance(duration_seconds, dict):
                duration_seconds = duration_seconds.get('numberDouble', duration_seconds.get('numberInt', 0))
            if isinstance(time_taken_ms, dict):
                time_taken_ms = time_taken_ms.get('numberDouble', time_taken_ms.get('numberInt', 0))
            
            excel_data.append({
                'Student Name': str(student.get('name', 'Unknown')),
                'Student Email': str(student.get('email', '')),
                'Roll Number': str(student.get('roll_number', '')),
                'Mobile Number': str(student.get('mobile_number', '')),
                'Campus': str(campus.get('name', 'Unknown Campus')) if campus else 'Unknown Campus',
                'Course': str(course.get('name', 'Unknown Course')) if course else 'Unknown Course',
                'Batch': str(batch.get('name', 'Unknown Batch')) if batch else 'Unknown Batch',
                'Test Name': str(test.get('name', 'Unknown Test')),
                'Total Questions': total_questions,
                'Correct Answers': correct_answers,
                'Total Score': round(score, 2),
                'Duration (seconds)': round(float(duration_seconds), 2) if duration_seconds else 0,
                'Time Taken (ms)': int(time_taken_ms) if time_taken_ms else 0,
                'Submitted At': submitted_at,
                'Status': str(attempt.get('status', 'unknown'))
            })
        
        # Create Excel file using openpyxl with enhanced styling
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Attempts"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        title_font = Font(bold=True, size=16, color="2F4F4F")
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal='left', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add test name as title (row 1)
        test_name = test.get('name', 'Unknown Test')
        ws.merge_cells('A1:P1')
        title_cell = ws.cell(row=1, column=1, value=f"Test Results: {test_name}")
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        ws.row_dimensions[1].height = 30
        
        # Add empty row for spacing
        ws.row_dimensions[2].height = 10
        
        # Add headers (row 3)
        headers = [
            'Student Name', 'Student Email', 'Roll Number', 'Mobile Number',
            'Campus', 'Course', 'Batch', 'Test Name',
            'Total Questions', 'Correct Answers', 'Total Score (%)',
            'Duration (seconds)', 'Time Taken (ms)', 'Submitted At', 'Status'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data starting from row 4
        for row_idx, data in enumerate(excel_data, 4):
            row_data = [
                data['Student Name'],
                data['Student Email'],
                data['Roll Number'],
                data['Mobile Number'],
                data['Campus'],
                data['Course'],
                data['Batch'],
                data['Test Name'],
                data['Total Questions'],
                data['Correct Answers'],
                data['Total Score'],
                data['Duration (seconds)'],
                data['Time Taken (ms)'],
                data['Submitted At'],
                data['Status']
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Add alternating row colors
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        # Set column widths for better display
        column_widths = {
            'A': 20,  # Student Name
            'B': 25,  # Student Email
            'C': 15,  # Roll Number
            'D': 15,  # Mobile Number
            'E': 20,  # Campus
            'F': 20,  # Course
            'G': 15,  # Batch
            'H': 25,  # Test Name
            'I': 15,  # Total Questions
            'J': 15,  # Correct Answers
            'K': 15,  # Total Score
            'L': 15,  # Duration
            'M': 15,  # Time Taken
            'N': 20,  # Submitted At
            'O': 12   # Status
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Set row heights
        for row in range(4, len(excel_data) + 4):
            ws.row_dimensions[row].height = 25
        
        # Add summary information at the bottom
        summary_row = len(excel_data) + 6
        ws.cell(row=summary_row, column=1, value="Summary:").font = Font(bold=True, size=12)
        ws.cell(row=summary_row + 1, column=1, value=f"Total Students: {len(excel_data)}")
        ws.cell(row=summary_row + 2, column=1, value=f"Test Name: {test_name}")
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return Excel file
        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=test_attempts_{test_id}.xlsx'
            }
        )
        
    except Exception as e:
        current_app.logger.error(f"Error in export_test_attempts: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Failed to export test attempts',
            'error': str(e)
        }), 500

@superadmin_bp.route('/export-test-attempts-csv/<test_id>', methods=['GET'])
@jwt_required()
def export_test_attempts_csv(test_id):
    """Export test attempts for a specific test as CSV"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user:
            return jsonify({
                'success': False,
                'message': 'Access denied. Authentication required.'
            }), 401
        
        # Get all attempts for this test
        attempts = list(mongo_db.student_test_attempts.find({
            'test_id': ObjectId(test_id),
            'test_type': 'online'
        }))
        
        if not attempts:
            return jsonify({
                'success': False,
                'message': 'No attempts found for this test'
            }), 404
        
        # Get test details
        test = mongo_db.tests.find_one({'_id': ObjectId(test_id)})
        if not test:
            return jsonify({
                'success': False,
                'message': 'Test not found'
            }), 404
        
        # Process attempts and get student details
        csv_data = []
        for attempt in attempts:
            student_id = str(attempt['student_id'])
            
            # Get student details
            student = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
            if not student:
                continue
            
            # Get campus, course, and batch details
            campus = mongo_db.campuses.find_one({'_id': student.get('campus_id')}) if student.get('campus_id') else None
            course = mongo_db.courses.find_one({'_id': student.get('course_id')}) if student.get('course_id') else None
            batch = mongo_db.batches.find_one({'_id': student.get('batch_id')}) if student.get('batch_id') else None
            
            # Calculate score - ensure we get clean numeric values
            total_questions = attempt.get('total_questions', 0)
            correct_answers = attempt.get('correct_answers', 0)
            
            # Handle MongoDB number types
            if isinstance(total_questions, dict):
                total_questions = total_questions.get('numberInt', total_questions.get('numberDouble', 0))
            if isinstance(correct_answers, dict):
                correct_answers = correct_answers.get('numberInt', correct_answers.get('numberDouble', 0))
            
            # Convert to integers
            total_questions = int(total_questions) if total_questions else 0
            correct_answers = int(correct_answers) if correct_answers else 0
            
            # Calculate score percentage
            score = 0
            if total_questions > 0:
                score = (correct_answers / total_questions) * 100
            
            # Format dates
            submitted_at = attempt.get('submitted_at', attempt.get('created_at'))
            if submitted_at:
                if isinstance(submitted_at, dict):
                    submitted_at = safe_isoformat(submitted_at.get('$date', submitted_at))
                else:
                    submitted_at = safe_isoformat(submitted_at)
            else:
                submitted_at = ''
            
            # Get clean numeric values for duration
            duration_seconds = attempt.get('duration_seconds', 0)
            time_taken_ms = attempt.get('time_taken_ms', 0)
            
            if isinstance(duration_seconds, dict):
                duration_seconds = duration_seconds.get('numberDouble', duration_seconds.get('numberInt', 0))
            if isinstance(time_taken_ms, dict):
                time_taken_ms = time_taken_ms.get('numberDouble', time_taken_ms.get('numberInt', 0))
            
            csv_data.append({
                'Student Name': str(student.get('name', 'Unknown')),
                'Student Email': str(student.get('email', '')),
                'Roll Number': str(student.get('roll_number', '')),
                'Mobile Number': str(student.get('mobile_number', '')),
                'Campus': str(campus.get('name', 'Unknown Campus')) if campus else 'Unknown Campus',
                'Course': str(course.get('name', 'Unknown Course')) if course else 'Unknown Course',
                'Batch': str(batch.get('name', 'Unknown Batch')) if batch else 'Unknown Batch',
                'Test Name': str(test.get('name', 'Unknown Test')),
                'Total Questions': total_questions,
                'Correct Answers': correct_answers,
                'Total Score (%)': round(score, 2),
                'Duration (seconds)': round(float(duration_seconds), 2) if duration_seconds else 0,
                'Time Taken (ms)': int(time_taken_ms) if time_taken_ms else 0,
                'Submitted At': submitted_at,
                'Status': str(attempt.get('status', 'unknown'))
            })
        
        # Create CSV file
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            'Student Name', 'Student Email', 'Roll Number', 'Mobile Number',
            'Campus', 'Course', 'Batch', 'Test Name',
            'Total Questions', 'Correct Answers', 'Total Score (%)',
            'Duration (seconds)', 'Time Taken (ms)', 'Submitted At', 'Status'
        ])
        
        # Write header
        writer.writeheader()
        
        # Write data
        for row in csv_data:
            writer.writerow(row)
        
        output.seek(0)
        
        # Return CSV file
        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename=test_attempts_{test_id}.csv'
            }
        )
        
    except Exception as e:
        current_app.logger.error(f"Error in export_test_attempts_csv: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Failed to export test attempts as CSV',
            'error': str(e)
        }), 500

@superadmin_bp.route('/online-tests-overview', methods=['GET'])
@jwt_required()
def get_online_tests_overview():
    """Get overview of all online tests with statistics"""
    try:
        # Get all online tests
        tests = list(mongo_db.tests.find({'test_type': 'online'}, {
            '_id': 1, 'name': 1, 'module_id': 1, 'created_at': 1
        }))
        
        if not tests:
            return jsonify({
                'success': True,
                'data': [],
                'message': 'No online tests found'
            })
        
        # Get all test results to calculate statistics
        test_results = list(mongo_db.student_test_attempts.find({
            'test_type': 'online'
        }))
        
        # Group results by test_id
        test_stats = {}
        for result in test_results:
            test_id = str(result['test_id'])
            if test_id not in test_stats:
                test_stats[test_id] = {
                    'total_attempts': 0,
                    'unique_students': set(),
                    'student_scores': {},  # student_id -> highest_score
                    'all_scores': []
                }
            
            test_stats[test_id]['total_attempts'] += 1
            test_stats[test_id]['unique_students'].add(str(result['student_id']))
            
            # Calculate score
            score = 0
            if 'correct_answers' in result and 'total_questions' in result:
                if result['total_questions'] > 0:
                    score = (result['correct_answers'] / result['total_questions']) * 100
            
            test_stats[test_id]['all_scores'].append(score)
            
            # Track highest score per student
            student_id = str(result['student_id'])
            if student_id not in test_stats[test_id]['student_scores']:
                test_stats[test_id]['student_scores'][student_id] = 0
            test_stats[test_id]['student_scores'][student_id] = max(
                test_stats[test_id]['student_scores'][student_id], 
                score
            )
        
        # Build response with statistics
        tests_overview = []
        for test in tests:
            test_id = str(test['_id'])
            stats = test_stats.get(test_id, {
                'total_attempts': 0,
                'unique_students': set(),
                'student_scores': {},
                'all_scores': []
            })
            
            # Calculate statistics
            unique_students_count = len(stats['unique_students'])
            highest_score = max(stats['all_scores']) if stats['all_scores'] else 0
            
            # Average score from unique students' highest scores
            student_highest_scores = list(stats['student_scores'].values())
            average_score = sum(student_highest_scores) / len(student_highest_scores) if student_highest_scores else 0
            
            tests_overview.append({
                'test_id': test_id,
                'test_name': test.get('name', 'Unknown Test'),
                'category': test.get('module_id', 'Unknown Category'),
                'total_attempts': stats['total_attempts'],
                'unique_students': unique_students_count,
                'highest_score': round(highest_score, 2),
                'average_score': round(average_score, 2),
                'created_at': safe_isoformat(test.get('created_at'))
            })
        
        # Sort by test name
        tests_overview.sort(key=lambda x: x['test_name'])
        
        return jsonify({
            'success': True,
            'data': tests_overview,
            'message': f'Found {len(tests_overview)} online tests'
        })
        
    except Exception as e:
        current_app.logger.error(f"Error in get_online_tests_overview: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Failed to fetch online tests overview',
            'error': str(e)
        }), 500

@superadmin_bp.route('/test-attempts/<test_id>', methods=['GET'])
@jwt_required()
def get_test_attempts(test_id):
    """Get all student attempts for a specific test"""
    try:
        # Get all attempts for this test
        attempts = list(mongo_db.student_test_attempts.find({
            'test_id': ObjectId(test_id),
            'test_type': 'online'
        }))
        
        if not attempts:
            return jsonify({
                'success': True,
                'data': [],
                'message': 'No attempts found for this test'
            })
        
        # Group attempts by student to get highest score per student
        student_attempts = {}
        for attempt in attempts:
            student_id = str(attempt['student_id'])
            
            # Calculate score
            score = 0
            if 'correct_answers' in attempt and 'total_questions' in attempt:
                if attempt['total_questions'] > 0:
                    score = (attempt['correct_answers'] / attempt['total_questions']) * 100
            
            if student_id not in student_attempts:
                # Get student details from students collection
                student = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
                if not student:
                    continue
                
                # Get campus, course, and batch details
                campus = mongo_db.campuses.find_one({'_id': student.get('campus_id')}) if student.get('campus_id') else None
                course = mongo_db.courses.find_one({'_id': student.get('course_id')}) if student.get('course_id') else None
                batch = mongo_db.batches.find_one({'_id': student.get('batch_id')}) if student.get('batch_id') else None
                
                student_attempts[student_id] = {
                    'student_id': student_id,
                    'student_name': student.get('name', 'Unknown'),
                    'student_email': student.get('email', ''),
                    'campus_name': campus.get('name', 'Unknown Campus') if campus else 'Unknown Campus',
                    'course_name': course.get('name', 'Unknown Course') if course else 'Unknown Course',
                    'batch_name': batch.get('name', 'Unknown Batch') if batch else 'Unknown Batch',
                    'total_questions': attempt.get('total_questions', 0),
                    'correct_answers': attempt.get('correct_answers', 0),
                    'highest_score': score,
                    'attempts_count': 0,
                    'latest_attempt': attempt.get('submitted_at', attempt.get('created_at'))
                }
            
            # Update highest score and attempts count
            student_attempts[student_id]['highest_score'] = max(
                student_attempts[student_id]['highest_score'], 
                score
            )
            student_attempts[student_id]['attempts_count'] += 1
            
            # Update latest attempt date
            attempt_date = attempt.get('submitted_at', attempt.get('created_at'))
            if attempt_date and attempt_date > student_attempts[student_id]['latest_attempt']:
                student_attempts[student_id]['latest_attempt'] = attempt_date
        
        # Convert to list and sort by highest score (descending)
        attempts_list = list(student_attempts.values())
        attempts_list.sort(key=lambda x: x['highest_score'], reverse=True)
        
        # Format dates
        for attempt in attempts_list:
            attempt['latest_attempt'] = safe_isoformat(attempt['latest_attempt'])
        
        return jsonify({
            'success': True,
            'data': attempts_list,
            'message': f'Found {len(attempts_list)} students who attempted this test'
        })
        
    except Exception as e:
        current_app.logger.error(f"Error in get_test_attempts: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'Failed to fetch test attempts',
            'error': str(e)
        }), 500

@superadmin_bp.route('/export-results', methods=['GET'])
@jwt_required()
def export_results():
    """Export test results as CSV"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        # Get query parameters
        module_filter = request.args.get('module')
        test_type_filter = request.args.get('test_type', 'online')  # Default to online tests
        campus_filter = request.args.get('campus')
        course_filter = request.args.get('course')
        batch_filter = request.args.get('batch')
        date_range = request.args.get('dateRange', 'all')
        score_range = request.args.get('scoreRange', 'all')
        export_format = request.args.get('format', 'csv')  # 'csv' or 'excel'
        
        # Build match conditions
        match_conditions = {}
        if module_filter:
            match_conditions['module_id'] = module_filter
        if test_type_filter:
            match_conditions['test_type'] = test_type_filter
        
        # Date range filter
        if date_range != 'all':
            now = datetime.utcnow()
            if date_range == 'today':
                start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            elif date_range == 'week':
                start_date = now - timedelta(days=7)
            elif date_range == 'month':
                start_date = now - timedelta(days=30)
            else:
                start_date = None
            
            if start_date:
                match_conditions['submitted_at'] = {'$gte': start_date}
        
        # Score range filter
        if score_range != 'all':
            if score_range == 'excellent':
                match_conditions['average_score'] = {'$gte': 90}
            elif score_range == 'good':
                match_conditions['average_score'] = {'$gte': 70, '$lt': 90}
            elif score_range == 'average':
                match_conditions['average_score'] = {'$gte': 50, '$lt': 70}
            elif score_range == 'poor':
                match_conditions['average_score'] = {'$lt': 50}
        
        pipeline = [
            {'$match': match_conditions},
            {
                '$lookup': {
                    'from': 'users',
                    'localField': 'student_id',
                    'foreignField': '_id',
                    'as': 'student_details'
                }
            },
            {'$unwind': '$student_details'},
            {
                '$lookup': {
                    'from': 'tests',
                    'localField': 'test_id',
                    'foreignField': '_id',
                    'as': 'test_details'
                }
            },
            {'$unwind': '$test_details'},
            {
                '$lookup': {
                    'from': 'students',
                    'localField': 'student_id',
                    'foreignField': 'user_id',
                    'as': 'student_profile'
                }
            },
            {'$unwind': '$student_profile'},
            {
                '$lookup': {
                    'from': 'campuses',
                    'localField': 'student_profile.campus_id',
                    'foreignField': '_id',
                    'as': 'campus_details'
                }
            },
            {'$unwind': '$campus_details'},
            {
                '$lookup': {
                    'from': 'courses',
                    'localField': 'student_profile.course_id',
                    'foreignField': '_id',
                    'as': 'course_details'
                }
            },
            {'$unwind': '$course_details'},
            {
                '$lookup': {
                    'from': 'batches',
                    'localField': 'student_profile.batch_id',
                    'foreignField': '_id',
                    'as': 'batch_details'
                }
            },
            {'$unwind': '$batch_details'},
            {
                '$project': {
                    'student_name': '$student_details.name',
                    'student_email': '$student_details.email',
                    'campus_name': '$campus_details.name',
                    'course_name': '$course_details.name',
                    'batch_name': '$batch_details.name',
                    'test_name': '$test_details.name',
                    'module_name': '$test_details.module_id',
                    'test_type': '$test_type',
                    'average_score': '$average_score',
                    'total_questions': '$total_questions',
                    'correct_answers': '$correct_answers',
                    'submitted_at': '$submitted_at',
                    'duration': '$duration',
                    'time_taken': '$time_taken'
                }
            },
            {'$sort': {'submitted_at': -1}}
        ]
        
        # Apply additional filters
        if campus_filter:
            pipeline.insert(0, {'$match': {'campus_name': campus_filter}})
        if course_filter:
            pipeline.insert(0, {'$match': {'course_name': course_filter}})
        if batch_filter:
            pipeline.insert(0, {'$match': {'batch_name': batch_filter}})
        
        results = list(mongo_db.db.test_results.aggregate(pipeline))
        
        if export_format == 'excel':
            # Export as Excel
            try:
                import pandas as pd
                from io import BytesIO
                
                # Prepare data for Excel
                excel_data = []
                for result in results:
                    # Extract actual values from MongoDB structures
                    score = result.get('average_score', 0)
                    if isinstance(score, dict) and 'numberDouble' in score:
                        score = float(score['numberDouble'])
                    elif isinstance(score, dict) and 'numberInt' in score:
                        score = float(score['numberInt'])
                    
                    total_questions = result.get('total_questions', 0)
                    if isinstance(total_questions, dict) and 'numberInt' in total_questions:
                        total_questions = int(total_questions['numberInt'])
                    
                    correct_answers = result.get('correct_answers', 0)
                    if isinstance(correct_answers, dict) and 'numberInt' in correct_answers:
                        correct_answers = int(correct_answers['numberInt'])
                    
                    # Format submitted_at as proper datetime
                    submitted_at = result.get('submitted_at', '')
                    if isinstance(submitted_at, dict) and 'date' in submitted_at:
                        # Convert MongoDB date to readable format
                        timestamp = int(submitted_at['date']['numberLong']) / 1000
                        submitted_at = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        submitted_at = safe_isoformat(submitted_at) if submitted_at else ''
                    
                    excel_data.append({
                        'Student Name': result.get('student_name', ''),
                        'Student Email': result.get('student_email', ''),
                        'Campus': result.get('campus_name', ''),
                        'Course': result.get('course_name', ''),
                        'Batch': result.get('batch_name', ''),
                        'Test Name': result.get('test_name', ''),
                        'Module': result.get('module_name', ''),
                        'Test Type': result.get('test_type', ''),
                        'Score (%)': score,
                        'Total Questions': total_questions,
                        'Correct Answers': correct_answers,
                        'Submitted At': submitted_at
                    })
                
                # Create Excel file
                df = pd.DataFrame(excel_data)
                output = BytesIO()
                
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Main results sheet
                    df.to_excel(writer, sheet_name='Test Results', index=False)
                    
                    # Summary sheet
                    summary_data = {
                        'Metric': ['Total Tests', 'Total Students', 'Average Score', 'Pass Rate'],
                        'Value': [
                            len(results),
                            len(set(r.get('student_id') for r in results)),
                            sum(r.get('average_score', 0) for r in results) / len(results) if results else 0,
                            len([r for r in results if r.get('average_score', 0) >= 60]) / len(results) * 100 if results else 0
                        ]
                    }
                    pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
                
                output.seek(0)
                
                from flask import Response
                return Response(
                    output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename=online-test-results-{datetime.now().strftime("%Y%m%d")}.xlsx'}
                )
                
            except ImportError:
                # Fallback to CSV if pandas is not available
                current_app.logger.warning("pandas not available, falling back to CSV export")
                export_format = 'csv'
        
        if export_format == 'csv':
            # Export as CSV
            import csv
            import io
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow([
                'Student Name',
                'Student Email',
                'Campus',
                'Course',
                'Batch',
                'Test Name',
                'Module',
                'Test Type',
                'Score (%)',
                'Total Questions',
                'Correct Answers',
                'Submitted At'
            ])
            
            # Write data
            for result in results:
                # Extract actual values from MongoDB structures
                score = result.get('average_score', 0)
                if isinstance(score, dict) and 'numberDouble' in score:
                    score = float(score['numberDouble'])
                elif isinstance(score, dict) and 'numberInt' in score:
                    score = float(score['numberInt'])
                
                total_questions = result.get('total_questions', 0)
                if isinstance(total_questions, dict) and 'numberInt' in total_questions:
                    total_questions = int(total_questions['numberInt'])
                
                correct_answers = result.get('correct_answers', 0)
                if isinstance(correct_answers, dict) and 'numberInt' in correct_answers:
                    correct_answers = int(correct_answers['numberInt'])
                
                # Format submitted_at as proper datetime
                submitted_at = result.get('submitted_at', '')
                if isinstance(submitted_at, dict) and 'date' in submitted_at:
                    # Convert MongoDB date to readable format
                    timestamp = int(submitted_at['date']['numberLong']) / 1000
                    submitted_at = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
                else:
                    submitted_at = safe_isoformat(submitted_at) if submitted_at else ''
                
                writer.writerow([
                    result.get('student_name', ''),
                    result.get('student_email', ''),
                    result.get('campus_name', ''),
                    result.get('course_name', ''),
                    result.get('batch_name', ''),
                    result.get('test_name', ''),
                    result.get('module_name', ''),
                    result.get('test_type', ''),
                    score,
                    total_questions,
                    correct_answers,
                    submitted_at
                ])
            
            output.seek(0)
            csv_content = output.getvalue()
            
            from flask import Response
            return Response(
                csv_content,
                mimetype='text/csv',
                headers={'Content-Disposition': f'attachment; filename=online-test-results-{datetime.now().strftime("%Y%m%d")}.csv'}
            )
        
    except Exception as e:
        current_app.logger.error(f"Error exporting results: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to export results: {str(e)}'
        }), 500


@superadmin_bp.route('/migrate-batch-course-instances', methods=['POST'])
@jwt_required()
def migrate_batch_course_instances():
    # Only super admin
    current_user_id = get_jwt_identity()
    user = mongo_db.find_user_by_id(current_user_id)
    if not user or user.get('role') != 'superadmin':
        return jsonify({'success': False, 'message': 'Access denied. Super admin privileges required.'}), 403
    # Migrate students
    students = list(mongo_db.students.find())
    updated_students = 0
    for s in students:
        batch_id = s.get('batch_id')
        course_id = s.get('course_id')
        if batch_id and course_id:
            instance_id = mongo_db.find_or_create_batch_course_instance(batch_id, course_id)
            mongo_db.students.update_one({'_id': s['_id']}, {'$set': {'batch_course_instance_id': instance_id}})
            updated_students += 1
    # Migrate modules
    modules = list(mongo_db.modules.find())
    updated_modules = 0
    for m in modules:
        batch_id = m.get('batch_id')
        course_id = m.get('course_id')
        if batch_id and course_id:
            instance_id = mongo_db.find_or_create_batch_course_instance(batch_id, course_id)
            mongo_db.modules.update_one({'_id': m['_id']}, {'$set': {'batch_course_instance_id': instance_id}})
            updated_modules += 1
    # Migrate test_results
    results = list(mongo_db.db.test_results.find())
    updated_results = 0
    for r in results:
        batch_id = r.get('batch_id')
        course_id = r.get('course_id')
        if batch_id and course_id:
            instance_id = mongo_db.find_or_create_batch_course_instance(batch_id, course_id)
            mongo_db.db.test_results.update_one({'_id': r['_id']}, {'$set': {'batch_course_instance_id': instance_id}})
            updated_results += 1
    return jsonify({'success': True, 'students_updated': updated_students, 'modules_updated': updated_modules, 'results_updated': updated_results}), 200 

@superadmin_bp.route('/debug-test-results', methods=['GET'])
@jwt_required()
def debug_test_results():
    """Debug endpoint to check test results data"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        # Check all collections
        debug_info = {
            'student_test_attempts_count': 0,
            'test_results_count': 0,
            'student_test_assignments_count': 0,
            'sample_student_attempts': [],
            'sample_test_results': [],
            'sample_assignments': [],
            'collections_exist': {
                'student_test_attempts': False,
                'test_results': False,
                'student_test_assignments': False
            }
        }
        
        try:
            # Check student_test_attempts collection
            debug_info['student_test_attempts_count'] = mongo_db.student_test_attempts.count_documents({})
            debug_info['collections_exist']['student_test_attempts'] = True
            if debug_info['student_test_attempts_count'] > 0:
                debug_info['sample_student_attempts'] = list(mongo_db.student_test_attempts.find().limit(3))
        except Exception as e:
            debug_info['student_test_attempts_error'] = str(e)
        
        try:
            # Check test_results collection
            if hasattr(mongo_db, 'db') and hasattr(mongo_db.db, 'test_results'):
                debug_info['test_results_count'] = mongo_db.db.test_results.count_documents({})
                debug_info['collections_exist']['test_results'] = True
                if debug_info['test_results_count'] > 0:
                    debug_info['sample_test_results'] = list(mongo_db.db.test_results.find().limit(3))
        except Exception as e:
            debug_info['test_results_error'] = str(e)
        
        try:
            # Check student_test_assignments collection
            debug_info['student_test_assignments_count'] = mongo_db.student_test_assignments.count_documents({'attempted': True})
            debug_info['collections_exist']['student_test_assignments'] = True
            if debug_info['student_test_assignments_count'] > 0:
                debug_info['sample_assignments'] = list(mongo_db.student_test_assignments.find({'attempted': True}).limit(3))
        except Exception as e:
            debug_info['student_test_assignments_error'] = str(e)
        
        # Convert ObjectIds to strings for JSON serialization
        def convert_objectids(obj):
            if isinstance(obj, dict):
                for key, value in obj.items():
                    if isinstance(value, ObjectId):
                        obj[key] = str(value)
                    elif isinstance(value, (list, dict)):
                        convert_objectids(value)
            elif isinstance(obj, list):
                for item in obj:
                    convert_objectids(item)
        
        convert_objectids(debug_info)
        
        return jsonify({
            'success': True,
            'data': debug_info
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error in debug_test_results: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to debug test results: {str(e)}'
        }), 500

@superadmin_bp.route('/test-results-simple', methods=['GET'])
@jwt_required()
def get_test_results_simple():
    """Simple endpoint to get test results without complex aggregation"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        all_results = []
        
        # Get simple results from student_test_attempts
        try:
            if hasattr(mongo_db, 'student_test_attempts'):
                attempts = list(mongo_db.student_test_attempts.find().limit(50))
                for attempt in attempts:
                    attempt['_id'] = str(attempt['_id'])
                    attempt['student_id'] = str(attempt['student_id'])
                    attempt['test_id'] = str(attempt['test_id'])
                    if attempt.get('submitted_at'):
                        attempt['submitted_at'] = safe_isoformat(attempt['submitted_at'])
                    attempt['source_collection'] = 'student_test_attempts'
                all_results.extend(attempts)
                current_app.logger.info(f"Found {len(attempts)} simple results in student_test_attempts")
        except Exception as e:
            current_app.logger.warning(f"Error getting simple results from student_test_attempts: {e}")
        
        # Get simple results from test_results
        try:
            if hasattr(mongo_db, 'db') and hasattr(mongo_db.db, 'test_results'):
                results = list(mongo_db.db.test_results.find().limit(50))
                for result in results:
                    result['_id'] = str(result['_id'])
                    result['student_id'] = str(result['student_id'])
                    result['test_id'] = str(result['test_id'])
                    if result.get('submitted_at'):
                        result['submitted_at'] = safe_isoformat(result['submitted_at'])
                    result['source_collection'] = 'test_results'
                all_results.extend(results)
                current_app.logger.info(f"Found {len(results)} simple results in test_results")
        except Exception as e:
            current_app.logger.warning(f"Error getting simple results from test_results: {e}")
        
        # Get simple results from student_test_assignments
        try:
            if hasattr(mongo_db, 'student_test_assignments'):
                assignments = list(mongo_db.student_test_assignments.find({'attempted': True}).limit(50))
                for assignment in assignments:
                    assignment['_id'] = str(assignment['_id'])
                    assignment['student_id'] = str(assignment['student_id'])
                    assignment['test_id'] = str(assignment['test_id'])
                    if assignment.get('completed_at'):
                        assignment['submitted_at'] = safe_isoformat(assignment['completed_at'])
                    assignment['source_collection'] = 'student_test_assignments'
                all_results.extend(assignments)
                current_app.logger.info(f"Found {len(assignments)} simple results in student_test_assignments")
        except Exception as e:
            current_app.logger.warning(f"Error getting simple results from student_test_assignments: {e}")
        
        return jsonify({
            'success': True,
            'data': all_results,
            'total_count': len(all_results)
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error getting simple test results: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to get simple test results: {str(e)}'
        }), 500 

@superadmin_bp.route('/debug-collections', methods=['GET'])
@jwt_required()
def debug_collections():
    """Debug endpoint to check what collections exist and what data they contain"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        debug_info = {
            'collections': {},
            'sample_data': {}
        }
        
        # Check student_test_attempts
        try:
            if hasattr(mongo_db, 'student_test_attempts'):
                count = mongo_db.student_test_attempts.count_documents({})
                debug_info['collections']['student_test_attempts'] = count
                if count > 0:
                    sample = list(mongo_db.student_test_attempts.find().limit(1))
                    debug_info['sample_data']['student_test_attempts'] = sample
        except Exception as e:
            debug_info['collections']['student_test_attempts'] = f"Error: {str(e)}"
        
        # Check test_results
        try:
            if hasattr(mongo_db, 'test_results'):
                count = mongo_db.test_results.count_documents({})
                debug_info['collections']['test_results'] = count
                if count > 0:
                    sample = list(mongo_db.test_results.find().limit(1))
                    debug_info['sample_data']['test_results'] = sample
        except Exception as e:
            debug_info['collections']['test_results'] = f"Error: {str(e)}"
        
        # Check student_test_assignments
        try:
            if hasattr(mongo_db, 'student_test_assignments'):
                count = mongo_db.student_test_assignments.count_documents({})
                attempted_count = mongo_db.student_test_assignments.count_documents({'attempted': True})
                debug_info['collections']['student_test_assignments'] = {
                    'total': count,
                    'attempted': attempted_count
                }
                if attempted_count > 0:
                    sample = list(mongo_db.student_test_assignments.find({'attempted': True}).limit(1))
                    debug_info['sample_data']['student_test_assignments'] = sample
        except Exception as e:
            debug_info['collections']['student_test_assignments'] = f"Error: {str(e)}"
        
        # Check tests collection
        try:
            if hasattr(mongo_db, 'tests'):
                count = mongo_db.tests.count_documents({})
                debug_info['collections']['tests'] = count
        except Exception as e:
            debug_info['collections']['tests'] = f"Error: {str(e)}"
        
        return jsonify({
            'success': True,
            'data': debug_info
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error in debug endpoint: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Debug failed: {str(e)}'
        }), 500

@superadmin_bp.route('/verify-database', methods=['GET'])
@jwt_required()
def verify_database():
    """Comprehensive database verification endpoint"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        verification_info = {
            'database_connection': {},
            'collections': {},
            'sample_data': {},
            'connection_test': {}
        }
        
        # Test database connection
        try:
            # Test basic connection
            client = mongo_db.db.client
            db_name = mongo_db.db.name
            verification_info['database_connection'] = {
                'status': 'connected',
                'database_name': db_name,
                'client_info': str(client)
            }
            
            # Test ping
            ping_result = client.admin.command('ping')
            verification_info['connection_test']['ping'] = ping_result
            
        except Exception as e:
            verification_info['database_connection'] = {
                'status': 'error',
                'error': str(e)
            }
        
        # Check all collections
        collections_to_check = [
            'users', 'students', 'tests', 'test_results', 'student_test_attempts',
            'student_test_assignments', 'campuses', 'courses', 'batches', 'question_bank'
        ]
        
        for collection_name in collections_to_check:
            try:
                collection = getattr(mongo_db, collection_name, None)
                if collection:
                    count = collection.count_documents({})
                    verification_info['collections'][collection_name] = {
                        'exists': True,
                        'count': count,
                        'type': str(type(collection))
                    }
                    
                    # Get sample data if count > 0
                    if count > 0:
                        sample = list(collection.find().limit(1))
                        if sample:
                            # Remove sensitive data
                            sample_doc = sample[0]
                            if '_id' in sample_doc:
                                sample_doc['_id'] = str(sample_doc['_id'])
                            verification_info['sample_data'][collection_name] = sample_doc
                else:
                    verification_info['collections'][collection_name] = {
                        'exists': False,
                        'error': 'Collection not found in mongo_db object'
                    }
            except Exception as e:
                verification_info['collections'][collection_name] = {
                    'exists': False,
                    'error': str(e)
                }
        
        # Test specific test result queries
        test_queries = {
            'practice_results_count': 0,
            'online_results_count': 0,
            'technical_results_count': 0,
            'total_student_attempts': 0,
            'total_test_assignments': 0
        }
        
        try:
            # Count practice results
            test_queries['practice_results_count'] = mongo_db.student_test_attempts.count_documents({'test_type': 'practice'})
        except Exception as e:
            test_queries['practice_results_count'] = f"Error: {str(e)}"
        
        try:
            # Count online results
            test_queries['online_results_count'] = mongo_db.student_test_assignments.count_documents({'attempted': True})
        except Exception as e:
            test_queries['online_results_count'] = f"Error: {str(e)}"
        
        try:
            # Count technical results
            test_queries['technical_results_count'] = mongo_db.test_results.count_documents({'test_type': 'technical'})
        except Exception as e:
            test_queries['technical_results_count'] = f"Error: {str(e)}"
        
        try:
            # Total student attempts
            test_queries['total_student_attempts'] = mongo_db.student_test_attempts.count_documents({})
        except Exception as e:
            test_queries['total_student_attempts'] = f"Error: {str(e)}"
        
        try:
            # Total test assignments
            test_queries['total_test_assignments'] = mongo_db.student_test_assignments.count_documents({})
        except Exception as e:
            test_queries['total_test_assignments'] = f"Error: {str(e)}"
        
        verification_info['test_queries'] = test_queries
        
        return jsonify({
            'success': True,
            'data': verification_info
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error in database verification: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Database verification failed: {str(e)}'
        }), 500

@superadmin_bp.route('/database-status', methods=['GET'])
@jwt_required()
def get_database_status():
    """Get comprehensive database status and connection information"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        status_info = {
            'connection_info': {},
            'database_info': {},
            'collections_status': {},
            'test_data_summary': {},
            'sample_data': {}
        }
        
        # Get connection information
        try:
            client = mongo_db.db.client
            db = mongo_db.db
            
            status_info['connection_info'] = {
                'database_name': db.name,
                'client_address': str(client.address),
                'client_host': str(client.HOST),
                'client_port': str(client.PORT),
                'connection_status': 'connected'
            }
            
            # Test ping
            ping_result = client.admin.command('ping')
            status_info['connection_info']['ping_result'] = ping_result
            
        except Exception as e:
            status_info['connection_info'] = {
                'status': 'error',
                'error': str(e)
            }
        
        # Get database information
        try:
            db_stats = db.command('dbStats')
            status_info['database_info'] = {
                'collections': db_stats.get('collections', 0),
                'data_size': db_stats.get('dataSize', 0),
                'storage_size': db_stats.get('storageSize', 0),
                'indexes': db_stats.get('indexes', 0),
                'index_size': db_stats.get('indexSize', 0)
            }
        except Exception as e:
            status_info['database_info'] = {'error': str(e)}
        
        # Check all collections and their status
        collections_to_check = [
            'users', 'students', 'tests', 'test_results', 'student_test_attempts',
            'student_test_assignments', 'campuses', 'courses', 'batches', 'question_bank'
        ]
        
        for collection_name in collections_to_check:
            try:
                collection = getattr(mongo_db, collection_name, None)
                if collection:
                    count = collection.count_documents({})
                    status_info['collections_status'][collection_name] = {
                        'exists': True,
                        'count': count,
                        'status': 'accessible'
                    }
                    
                    # Get sample data for key collections
                    if collection_name in ['tests', 'student_test_attempts', 'test_results', 'student_test_assignments'] and count > 0:
                        sample = list(collection.find().limit(1))
                        if sample:
                            sample_doc = sample[0]
                            if '_id' in sample_doc:
                                sample_doc['_id'] = str(sample_doc['_id'])
                            status_info['sample_data'][collection_name] = sample_doc
                else:
                    status_info['collections_status'][collection_name] = {
                        'exists': False,
                        'status': 'not_accessible'
                    }
            except Exception as e:
                status_info['collections_status'][collection_name] = {
                    'exists': False,
                    'status': 'error',
                    'error': str(e)
                }
        
        # Get test data summary
        try:
            # Count tests by type
            test_types = {}
            if hasattr(mongo_db, 'tests'):
                pipeline = [
                    {'$group': {'_id': '$test_type', 'count': {'$sum': 1}}}
                ]
                test_type_results = list(mongo_db.tests.aggregate(pipeline))
                for result in test_type_results:
                    test_types[result['_id']] = result['count']
            
            # Count test results by type
            result_types = {}
            if hasattr(mongo_db, 'student_test_attempts'):
                pipeline = [
                    {'$group': {'_id': '$test_type', 'count': {'$sum': 1}}}
                ]
                result_type_results = list(mongo_db.student_test_attempts.aggregate(pipeline))
                for result in result_type_results:
                    result_types[result['_id']] = result['count']
            
            # Count online exam results
            online_results_count = 0
            if hasattr(mongo_db, 'student_test_assignments'):
                online_results_count = mongo_db.student_test_assignments.count_documents({'attempted': True})
            
            status_info['test_data_summary'] = {
                'test_types': test_types,
                'result_types': result_types,
                'online_exam_results': online_results_count,
                'total_tests': sum(test_types.values()) if test_types else 0,
                'total_results': sum(result_types.values()) if result_types else 0
            }
            
        except Exception as e:
            status_info['test_data_summary'] = {'error': str(e)}
        
        return jsonify({
            'success': True,
            'data': status_info
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error getting database status: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to get database status: {str(e)}'
        }), 500

@superadmin_bp.route('/test-database-connection', methods=['GET'])
@jwt_required()
def test_database_connection():
    """Simple test endpoint to verify database connection and show sample data"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        # Test basic connection
        db_name = mongo_db.db.name
        client = mongo_db.db.client
        
        # Get collection counts
        collections_info = {}
        key_collections = ['tests', 'student_test_attempts', 'test_results', 'student_test_assignments', 'users', 'students']
        
        for collection_name in key_collections:
            try:
                collection = getattr(mongo_db, collection_name, None)
                if collection:
                    count = collection.count_documents({})
                    collections_info[collection_name] = {
                        'count': count,
                        'status': 'accessible'
                    }
                else:
                    collections_info[collection_name] = {
                        'count': 0,
                        'status': 'not_found'
                    }
            except Exception as e:
                collections_info[collection_name] = {
                    'count': 0,
                    'status': 'error',
                    'error': str(e)
                }
        
        # Get sample test data
        sample_tests = []
        if hasattr(mongo_db, 'tests'):
            try:
                sample_tests = list(mongo_db.tests.find().limit(3))
                for test in sample_tests:
                    if '_id' in test:
                        test['_id'] = str(test['_id'])
            except Exception as e:
                sample_tests = [{'error': str(e)}]
        
        # Get sample test results
        sample_results = []
        if hasattr(mongo_db, 'student_test_attempts'):
            try:
                sample_results = list(mongo_db.student_test_attempts.find().limit(3))
                for result in sample_results:
                    if '_id' in result:
                        result['_id'] = str(result['_id'])
            except Exception as e:
                sample_results = [{'error': str(e)}]
        
        return jsonify({
            'success': True,
            'data': {
                'database_name': db_name,
                'connection_status': 'connected',
                'collections_info': collections_info,
                'sample_tests': sample_tests,
                'sample_results': sample_results,
                'message': f'Successfully connected to database: {db_name}'
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error testing database connection: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Database connection test failed: {str(e)}'
        }), 500

@superadmin_bp.route('/debug-database-connection', methods=['GET'])
@jwt_required()
def debug_database_connection():
    """Debug endpoint to show exact database connection details"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        # Get detailed connection information
        db = mongo_db.db
        client = db.client
        
        # Get environment variables
        import os
        mongodb_uri = os.getenv('MONGODB_URI', 'NOT_SET')
        
        # Get database name from URI parsing
        from urllib.parse import urlparse
        parsed_uri = urlparse(mongodb_uri)
        uri_db_name = parsed_uri.path.strip('/').split('?')[0] if parsed_uri.path else 'NO_DB_IN_URI'
        
        # Get actual database name being used
        actual_db_name = db.name
        
        # Test collections with detailed info
        collections_debug = {}
        key_collections = ['tests', 'student_test_attempts', 'test_results', 'student_test_assignments']
        
        for collection_name in key_collections:
            try:
                collection = getattr(mongo_db, collection_name, None)
                if collection:
                    count = collection.count_documents({})
                    # Get a sample document
                    sample = list(collection.find().limit(1))
                    collections_debug[collection_name] = {
                        'count': count,
                        'sample_exists': len(sample) > 0,
                        'sample_doc': sample[0] if sample else None
                    }
                else:
                    collections_debug[collection_name] = {
                        'count': 0,
                        'error': 'Collection not found'
                    }
            except Exception as e:
                collections_debug[collection_name] = {
                    'count': 0,
                    'error': str(e)
                }
        
        # List all collections in the database
        all_collections = []
        try:
            all_collections = db.list_collection_names()
        except Exception as e:
            all_collections = [f'Error: {str(e)}']
        
        return jsonify({
            'success': True,
            'debug_info': {
                'environment_mongodb_uri': mongodb_uri,
                'parsed_db_name_from_uri': uri_db_name,
                'actual_database_name': actual_db_name,
                'client_address': str(client.address),
                'all_collections_in_db': all_collections,
                'collections_debug': collections_debug,
                'connection_status': 'connected'
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error in database debug: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Database debug failed: {str(e)}'
        }), 500

@superadmin_bp.route('/force-connect-suma-madam', methods=['GET'])
@jwt_required()
def force_connect_suma_madam():
    """Force connection to suma_madam database and test collections"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        # Force connection to suma_madam database
        import os
        from pymongo import MongoClient
        from urllib.parse import urlparse
        
        # Get the base URI without database name
        mongodb_uri = os.getenv('MONGODB_URI', 'mongodb+srv://teja:teja0000@versant.ia46v3i.mongodb.net/suma_madam?retryWrites=true&w=majority&appName=Versant&connectTimeoutMS=30000&socketTimeoutMS=30000&serverSelectionTimeoutMS=30000')
        
        # Parse URI and replace database name with suma_madam
        parsed_uri = urlparse(mongodb_uri)
        base_uri = f"{parsed_uri.scheme}://{parsed_uri.netloc}"
        params = parsed_uri.query
        
        # Create new URI with suma_madam database
        new_uri = f"{base_uri}/suma_madam"
        if params:
            new_uri += f"?{params}"
        
        print(f"🔗 Forcing connection to: {new_uri}")
        
        # Create client with suma_madam database
        client = MongoClient(new_uri, connectTimeoutMS=30000, socketTimeoutMS=30000, serverSelectionTimeoutMS=30000)
        db = client['suma_madam']
        
        # Test connection
        client.admin.command('ping')
        
        # Get collection information
        collections_info = {}
        key_collections = ['tests', 'student_test_attempts', 'test_results', 'student_test_assignments', 'users', 'students']
        
        for collection_name in key_collections:
            try:
                collection = db[collection_name]
                count = collection.count_documents({})
                sample = list(collection.find().limit(1))
                collections_info[collection_name] = {
                    'count': count,
                    'sample_exists': len(sample) > 0,
                    'sample_doc': sample[0] if sample else None
                }
            except Exception as e:
                collections_info[collection_name] = {
                    'count': 0,
                    'error': str(e)
                }
        
        # List all collections
        all_collections = db.list_collection_names()
        
        return jsonify({
            'success': True,
            'data': {
                'database_name': 'suma_madam',
                'connection_uri': new_uri,
                'connection_status': 'connected',
                'all_collections': all_collections,
                'collections_info': collections_info,
                'message': 'Successfully connected to suma_madam database'
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error forcing connection to suma_madam: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to connect to suma_madam database: {str(e)}'
        }), 500




@superadmin_bp.route('/all-test-results', methods=['GET'])
@jwt_required()
def get_all_test_results():
    """Get all test results from all collections with detailed information"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        all_results = []
        debug_info = {}
        
        # Add debug information about the first student attempt
        try:
            if hasattr(mongo_db, 'student_test_attempts'):
                sample_attempt = mongo_db.student_test_attempts.find_one()
                if sample_attempt:
                    debug_info['sample_student_attempt'] = {
                        'student_id': str(sample_attempt.get('student_id')),
                        'test_id': str(sample_attempt.get('test_id')),
                        'test_type': sample_attempt.get('test_type'),
                        'average_score': sample_attempt.get('average_score')
                    }
                    
                    # Get the student profile
                    student_profile = mongo_db.students.find_one({'user_id': sample_attempt['student_id']})
                    if student_profile:
                        debug_info['sample_student_profile'] = {
                            'user_id': str(student_profile.get('user_id')),
                            'campus_id': str(student_profile.get('campus_id')) if student_profile.get('campus_id') else None,
                            'course_id': str(student_profile.get('course_id')) if student_profile.get('course_id') else None,
                            'batch_id': str(student_profile.get('batch_id')) if student_profile.get('batch_id') else None,
                            'roll_number': student_profile.get('roll_number')
                        }
                        
                        # Get campus details
                        if student_profile.get('campus_id'):
                            campus = mongo_db.campuses.find_one({'_id': student_profile['campus_id']})
                            if campus:
                                debug_info['sample_campus'] = {
                                    'id': str(campus['_id']),
                                    'name': campus.get('name')
                                }
                        
                        # Get course details
                        if student_profile.get('course_id'):
                            course = mongo_db.courses.find_one({'_id': student_profile['course_id']})
                            if course:
                                debug_info['sample_course'] = {
                                    'id': str(course['_id']),
                                    'name': course.get('name')
                                }
                        
                        # Get batch details
                        if student_profile.get('batch_id'):
                            batch = mongo_db.batches.find_one({'_id': student_profile['batch_id']})
                            if batch:
                                debug_info['sample_batch'] = {
                                    'id': str(batch['_id']),
                                    'name': batch.get('name')
                                }
                    
                    # Get user details
                    user_details = mongo_db.users.find_one({'_id': sample_attempt['student_id']})
                    if user_details:
                        debug_info['sample_user'] = {
                            'id': str(user_details['_id']),
                            'name': user_details.get('name'),
                            'email': user_details.get('email'),
                            'role': user_details.get('role')
                        }
        except Exception as e:
            current_app.logger.error(f"Error getting debug info: {str(e)}")
        
        # Get results from student_test_attempts collection
        try:
            if hasattr(mongo_db, 'student_test_attempts'):
                # First check if there are any documents in the collection
                total_attempts = mongo_db.student_test_attempts.count_documents({})
                current_app.logger.info(f"Total documents in student_test_attempts: {total_attempts}")
                
                if total_attempts > 0:
                    pipeline = [
                    {
                        '$lookup': {
                            'from': 'users',
                            'localField': 'student_id',
                            'foreignField': '_id',
                            'as': 'student_details'
                        }
                    },
                        {'$unwind': {'path': '$student_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'tests',
                            'localField': 'test_id',
                            'foreignField': '_id',
                            'as': 'test_details'
                        }
                    },
                        {'$unwind': {'path': '$test_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'students',
                            'localField': 'student_id',
                            'foreignField': 'user_id',
                            'as': 'student_profile'
                        }
                    },
                        {'$unwind': {'path': '$student_profile', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'campuses',
                            'localField': 'student_profile.campus_id',
                            'foreignField': '_id',
                            'as': 'campus_details'
                        }
                    },
                        {'$unwind': {'path': '$campus_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'courses',
                            'localField': 'student_profile.course_id',
                            'foreignField': '_id',
                            'as': 'course_details'
                        }
                    },
                        {'$unwind': {'path': '$course_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'batches',
                            'localField': 'student_profile.batch_id',
                            'foreignField': '_id',
                            'as': 'batch_details'
                        }
                    },
                        {'$unwind': {'path': '$batch_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$project': {
                            '_id': 1,
                            'student_id': 1,
                            'test_id': 1,
                            'student_name': {'$ifNull': ['$student_details.name', 'Unknown Student']},
                            'student_email': {'$ifNull': ['$student_details.email', 'unknown@example.com']},
                            'campus_name': {'$ifNull': ['$campus_details.name', 'Unknown Campus']},
                            'course_name': {'$ifNull': ['$course_details.name', 'Unknown Course']},
                            'batch_name': {'$ifNull': ['$batch_details.name', 'Unknown Batch']},
                            'test_name': {'$ifNull': ['$test_details.name', 'Unknown Test']},
                            'module_name': {'$ifNull': ['$test_details.module_id', 'Unknown']},
                            'test_type': 1,
                            'average_score': 1,
                            'score_percentage': 1,
                            'total_questions': 1,
                            'correct_answers': 1,
                            'submitted_at': 1,
                            'results': 1,  # Include detailed results with transcripts
                            'source_collection': 'student_test_attempts',
                            # Add debug fields to see what's actually in the data
                            'debug_student_profile': '$student_profile',
                            'debug_campus_details': '$campus_details',
                            'debug_course_details': '$course_details',
                            'debug_batch_details': '$batch_details'
                        }
                    },
                    {'$sort': {'submitted_at': -1}}
                ]
                
                    results = list(mongo_db.student_test_attempts.aggregate(pipeline))
                    all_results.extend(results)
                    current_app.logger.info(f"Found {len(results)} results in student_test_attempts")
                    
                    # Debug: Check a sample result to see what data we're getting
                    if results:
                        sample_result = results[0]
                        current_app.logger.info(f"Sample result from student_test_attempts: {sample_result}")
                        current_app.logger.info(f"Sample student_id type: {type(sample_result.get('student_id'))}")
                        current_app.logger.info(f"Sample student_id value: {sample_result.get('student_id')}")
                        current_app.logger.info(f"Sample student_name: {sample_result.get('student_name')}")
                        current_app.logger.info(f"Sample campus_name: {sample_result.get('campus_name')}")
                else:
                    current_app.logger.info("No documents found in student_test_attempts collection")
        except Exception as e:
            current_app.logger.warning(f"Error getting results from student_test_attempts: {e}")
        
        # Get results from test_results collection
        try:
            if hasattr(mongo_db, 'db') and hasattr(mongo_db.db, 'test_results'):
                total_test_results = mongo_db.db.test_results.count_documents({})
                current_app.logger.info(f"Total documents in test_results: {total_test_results}")
                
                if total_test_results > 0:
                    pipeline = [
                    {
                        '$lookup': {
                            'from': 'users',
                            'localField': 'student_id',
                            'foreignField': '_id',
                            'as': 'student_details'
                        }
                    },
                        {'$unwind': {'path': '$student_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'tests',
                            'localField': 'test_id',
                            'foreignField': '_id',
                            'as': 'test_details'
                        }
                    },
                        {'$unwind': {'path': '$test_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'students',
                            'localField': 'student_id',
                            'foreignField': 'user_id',
                            'as': 'student_profile'
                        }
                    },
                        {'$unwind': {'path': '$student_profile', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'campuses',
                            'localField': 'student_profile.campus_id',
                            'foreignField': '_id',
                            'as': 'campus_details'
                        }
                    },
                        {'$unwind': {'path': '$campus_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'courses',
                            'localField': 'student_profile.course_id',
                            'foreignField': '_id',
                            'as': 'course_details'
                        }
                    },
                        {'$unwind': {'path': '$course_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'batches',
                            'localField': 'student_profile.batch_id',
                            'foreignField': '_id',
                            'as': 'batch_details'
                        }
                    },
                        {'$unwind': {'path': '$batch_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$project': {
                            '_id': 1,
                            'student_id': 1,
                                'test_id': 1,
                                'student_name': {'$ifNull': ['$student_details.name', 'Unknown Student']},
                                'student_email': {'$ifNull': ['$student_details.email', 'unknown@example.com']},
                                'campus_name': {'$ifNull': ['$campus_details.name', 'Unknown Campus']},
                                'course_name': {'$ifNull': ['$course_details.name', 'Unknown Course']},
                                'batch_name': {'$ifNull': ['$batch_details.name', 'Unknown Batch']},
                                'test_name': {'$ifNull': ['$test_details.name', 'Unknown Test']},
                                'module_name': {'$ifNull': ['$test_details.module_id', 'Unknown']},
                            'test_type': 1,
                            'average_score': 1,
                            'score_percentage': 1,
                            'total_questions': 1,
                            'correct_answers': 1,
                            'submitted_at': 1,
                            'results': 1,  # Include detailed results with transcripts
                            'source_collection': 'test_results'
                        }
                    },
                    {'$sort': {'submitted_at': -1}}
                ]
                
                    results = list(mongo_db.db.test_results.aggregate(pipeline))
                    all_results.extend(results)
                    current_app.logger.info(f"Found {len(results)} results in test_results")
                else:
                    current_app.logger.info("No documents found in test_results collection")
        except Exception as e:
            current_app.logger.warning(f"Error getting results from test_results: {e}")
        
        # Get results from student_test_assignments collection
        try:
            if hasattr(mongo_db, 'student_test_assignments'):
                total_assignments = mongo_db.student_test_assignments.count_documents({'attempted': True})
                current_app.logger.info(f"Total attempted assignments in student_test_assignments: {total_assignments}")
                
                if total_assignments > 0:
                    pipeline = [
                    {'$match': {'attempted': True}},
                    {
                        '$lookup': {
                            'from': 'tests',
                            'localField': 'test_id',
                            'foreignField': '_id',
                            'as': 'test_details'
                        }
                    },
                        {'$unwind': {'path': '$test_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'users',
                            'localField': 'student_id',
                            'foreignField': '_id',
                            'as': 'student_details'
                        }
                    },
                        {'$unwind': {'path': '$student_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'students',
                            'localField': 'student_id',
                            'foreignField': 'user_id',
                            'as': 'student_profile'
                        }
                    },
                        {'$unwind': {'path': '$student_profile', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'campuses',
                            'localField': 'student_profile.campus_id',
                            'foreignField': '_id',
                            'as': 'campus_details'
                        }
                    },
                        {'$unwind': {'path': '$campus_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'courses',
                            'localField': 'student_profile.course_id',
                            'foreignField': '_id',
                            'as': 'course_details'
                        }
                    },
                        {'$unwind': {'path': '$course_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$lookup': {
                            'from': 'batches',
                            'localField': 'student_profile.batch_id',
                            'foreignField': '_id',
                            'as': 'batch_details'
                        }
                    },
                        {'$unwind': {'path': '$batch_details', 'preserveNullAndEmptyArrays': True}},
                    {
                        '$project': {
                            '_id': 1,
                            'student_id': 1,
                                'test_id': 1,
                                'student_name': {'$ifNull': ['$student_details.name', 'Unknown Student']},
                                'student_email': {'$ifNull': ['$student_details.email', 'unknown@example.com']},
                                'campus_name': {'$ifNull': ['$campus_details.name', 'Unknown Campus']},
                                'course_name': {'$ifNull': ['$course_details.name', 'Unknown Course']},
                                'batch_name': {'$ifNull': ['$batch_details.name', 'Unknown Batch']},
                                'test_name': {'$ifNull': ['$test_details.name', 'Unknown Test']},
                                'module_name': {'$ifNull': ['$test_details.module_id', 'Unknown']},
                                'test_type': {'$ifNull': ['$test_details.test_type', 'practice']},
                            'average_score': '$percentage',
                            'score_percentage': '$percentage',
                                'total_questions': {'$size': {'$ifNull': ['$questions', []]}},
                            'correct_answers': '$score',
                            'submitted_at': '$completed_at',
                            'results': '$detailed_results',  # Include detailed results
                            'source_collection': 'student_test_assignments'
                        }
                    },
                    {'$sort': {'submitted_at': -1}}
                ]
                
                    results = list(mongo_db.student_test_assignments.aggregate(pipeline))
                    all_results.extend(results)
                    current_app.logger.info(f"Found {len(results)} results in student_test_assignments")
                else:
                    current_app.logger.info("No attempted assignments found in student_test_assignments collection")
        except Exception as e:
            current_app.logger.warning(f"Error getting results from student_test_assignments: {e}")
        
        # Remove duplicates and process results
        seen = set()
        unique_results = []
        for result in all_results:
            key = (str(result.get('test_id')), str(result.get('student_id')), str(result.get('submitted_at')))
            if key not in seen:
                seen.add(key)
                # Process the result - convert ObjectIds to strings
                result['_id'] = str(result['_id'])
                result['student_id'] = str(result['student_id'])
                if result.get('test_id'):
                    result['test_id'] = str(result['test_id'])
                
                # Convert any other ObjectId fields that might exist recursively
                def convert_objectids_recursive(obj):
                    if isinstance(obj, dict):
                        for k, v in obj.items():
                            if isinstance(v, ObjectId):
                                obj[k] = str(v)
                            elif isinstance(v, (dict, list)):
                                convert_objectids_recursive(v)
                    elif isinstance(obj, list):
                        for i, item in enumerate(obj):
                            if isinstance(item, ObjectId):
                                obj[i] = str(item)
                            elif isinstance(item, (dict, list)):
                                convert_objectids_recursive(item)
                
                convert_objectids_recursive(result)
                
                # If student details are missing, try direct lookup
                if result.get('student_name') == 'Unknown Student' and result.get('student_id'):
                    try:
                        # Direct lookup for user details
                        user_details = mongo_db.users.find_one({'_id': ObjectId(result['student_id'])})
                        if user_details:
                            result['student_name'] = user_details.get('name', 'Unknown Student')
                            result['student_email'] = user_details.get('email', 'unknown@example.com')
                            
                            # Direct lookup for student profile
                            student_profile = mongo_db.students.find_one({'user_id': ObjectId(result['student_id'])})
                            if student_profile:
                                # Direct lookup for campus
                                if student_profile.get('campus_id'):
                                    campus = mongo_db.campuses.find_one({'_id': student_profile['campus_id']})
                                    if campus:
                                        result['campus_name'] = campus.get('name', 'Unknown Campus')
                                
                                # Direct lookup for course
                                if student_profile.get('course_id'):
                                    course = mongo_db.courses.find_one({'_id': student_profile['course_id']})
                                    if course:
                                        result['course_name'] = course.get('name', 'Unknown Course')
                                
                                # Direct lookup for batch
                                if student_profile.get('batch_id'):
                                    batch = mongo_db.batches.find_one({'_id': student_profile['batch_id']})
                                    if batch:
                                        result['batch_name'] = batch.get('name', 'Unknown Batch')
                                
                                current_app.logger.info(f"Enhanced result for student {result['student_id']}: {result['student_name']}, {result['campus_name']}, {result['course_name']}, {result['batch_name']}")
                    except Exception as e:
                        current_app.logger.error(f"Error enhancing result for student {result.get('student_id')}: {str(e)}")
                
                result['module_name'] = MODULES.get(result.get('module_name', ''), result.get('module_name', 'Unknown'))
                if result.get('submitted_at'):
                    result['submitted_at'] = safe_isoformat(result['submitted_at'])
                unique_results.append(result)
        
        # Sort results by submitted_at in descending order to show latest attempts first
        unique_results.sort(key=lambda x: x.get('submitted_at', ''), reverse=True)
        
        current_app.logger.info(f"Total unique results found: {len(unique_results)}")
        
        # Final comprehensive ObjectId conversion for all results
        def final_convert_objectids(obj):
            if isinstance(obj, dict):
                for k, v in obj.items():
                    if isinstance(v, ObjectId):
                        obj[k] = str(v)
                    elif isinstance(v, (dict, list)):
                        final_convert_objectids(v)
            elif isinstance(obj, list):
                for i, item in enumerate(obj):
                    if isinstance(item, ObjectId):
                        obj[i] = str(item)
                    elif isinstance(item, (dict, list)):
                        final_convert_objectids(item)
        
        # Convert ObjectIds in all results
        for result in unique_results:
            final_convert_objectids(result)
        
        # Convert ObjectIds in debug_info as well
        final_convert_objectids(debug_info)
        
        return jsonify({
            'success': True,
            'data': unique_results,
            'total_count': len(unique_results),
            'debug_info': debug_info
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error getting all test results: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to get test results: {str(e)}'
        }), 500

@superadmin_bp.route('/test-result-details/<result_id>', methods=['GET'])
@jwt_required()
def get_test_result_details(result_id):
    """Get detailed test result with transcripts and audio URLs"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        # Try to find the result in all collections
        result = None
        source_collection = None
        
        # Try student_test_attempts first
        try:
            if hasattr(mongo_db, 'student_test_attempts'):
                result = mongo_db.student_test_attempts.find_one({'_id': ObjectId(result_id)})
                if result:
                    source_collection = 'student_test_attempts'
        except Exception as e:
            current_app.logger.warning(f"Error searching student_test_attempts: {e}")
        
        # Try test_results if not found
        if not result:
            try:
                if hasattr(mongo_db, 'db') and hasattr(mongo_db.db, 'test_results'):
                    result = mongo_db.db.test_results.find_one({'_id': ObjectId(result_id)})
                    if result:
                        source_collection = 'test_results'
                elif hasattr(mongo_db, 'test_results'):
                    result = mongo_db.test_results.find_one({'_id': ObjectId(result_id)})
                    if result:
                        source_collection = 'test_results'
            except Exception as e:
                current_app.logger.warning(f"Error searching test_results: {e}")
        
        # Try student_test_assignments if not found
        if not result:
            try:
                if hasattr(mongo_db, 'student_test_assignments'):
                    result = mongo_db.student_test_assignments.find_one({'_id': ObjectId(result_id)})
                    if result:
                        source_collection = 'student_test_assignments'
            except Exception as e:
                current_app.logger.warning(f"Error searching student_test_assignments: {e}")
        
        if not result:
            return jsonify({
                'success': False,
                'message': 'Test result not found'
            }), 404
        
        # Get test details
        test = mongo_db.tests.find_one({'_id': result['test_id']})
        if not test:
            return jsonify({
                'success': False,
                'message': 'Test not found'
            }), 404
        
        # Get student details
        student = mongo_db.users.find_one({'_id': result['student_id']})
        if not student:
            return jsonify({
                'success': False,
                'message': 'Student not found'
            }), 404
        
        # Process detailed results
        detailed_results = []
        if result.get('results'):
            for i, question_result in enumerate(result['results']):
                detailed_result = {
                    'question_index': i,
                    'question': question_result.get('question', ''),
                    'question_type': question_result.get('question_type', 'unknown'),
                    'is_correct': question_result.get('is_correct', False),
                    'score': question_result.get('score', 0),
                    'similarity_score': question_result.get('similarity_score', 0)
                }
                
                # Add audio-specific details for listening/speaking tests
                if question_result.get('question_type') in ['audio', 'listening', 'speaking']:
                    detailed_result.update({
                        'student_audio_url': question_result.get('student_audio_url', ''),
                        'student_text': question_result.get('student_text', ''),
                        'original_text': question_result.get('original_text', ''),
                        'missing_words': question_result.get('missing_words', []),
                        'extra_words': question_result.get('extra_words', [])
                    })
                
                # Add MCQ-specific details
                elif question_result.get('question_type') == 'mcq':
                    detailed_result.update({
                        'selected_answer': question_result.get('selected_answer', ''),
                        'correct_answer': question_result.get('correct_answer', ''),
                        'options': question_result.get('options', [])
                    })
                
                detailed_results.append(detailed_result)
        
        # Prepare response
        response_data = {
            'result_id': str(result['_id']),
            'test_id': str(result['test_id']),
            'student_id': str(result['student_id']),
            'test_name': test.get('name', 'Unknown Test'),
            'module_name': MODULES.get(test.get('module_id', ''), test.get('module_id', 'Unknown')),
            'test_type': result.get('test_type', 'unknown'),
            'student_name': student.get('name', 'Unknown Student'),
            'student_email': student.get('email', ''),
            'average_score': result.get('average_score', 0),
            'score_percentage': result.get('score_percentage', 0),
            'total_questions': result.get('total_questions', 0),
            'correct_answers': result.get('correct_answers', 0),
            'submitted_at': safe_isoformat(result.get('submitted_at')) if result.get('submitted_at') else '',
            'source_collection': source_collection,
            'detailed_results': detailed_results
        }
        
        return jsonify({
            'success': True,
            'data': response_data
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error getting test result details: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to get test result details: {str(e)}'
        }), 500 

@superadmin_bp.route('/filter-options', methods=['GET'])
@jwt_required()
def get_filter_options():
    """Get filter options for results page"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        # Get modules
        modules = []
        for module_id, module_info in MODULES.items():
            modules.append({
                'id': module_id,
                'name': module_info
            })
        
        # Get test types
        test_types = [
            {'id': 'practice', 'name': 'Practice'},
            {'id': 'online', 'name': 'Online Exam'},
            {'id': 'technical', 'name': 'Technical'},
            {'id': 'mcq', 'name': 'MCQ'}
        ]
        
        # Get campuses
        campuses = []
        try:
            campus_cursor = mongo_db.campuses.find({}, {'name': 1})
            for campus in campus_cursor:
                campuses.append({
                    'id': str(campus['_id']),
                    'name': campus['name']
                })
        except Exception as e:
            current_app.logger.error(f"Error fetching campuses: {e}")
        
        # Get courses
        courses = []
        try:
            course_cursor = mongo_db.courses.find({}, {'name': 1})
            for course in course_cursor:
                courses.append({
                    'id': str(course['_id']),
                    'name': course['name']
                })
        except Exception as e:
            current_app.logger.error(f"Error fetching courses: {e}")
        
        # Get batches
        batches = []
        try:
            batch_cursor = mongo_db.batches.find({}, {'name': 1})
            for batch in batch_cursor:
                batches.append({
                    'id': str(batch['_id']),
                    'name': batch['name']
                })
        except Exception as e:
            current_app.logger.error(f"Error fetching batches: {e}")
        
        # Get levels
        levels = []
        for level_id, level_info in LEVELS.items():
            levels.append({
                'id': level_id,
                'name': level_info.get('name', level_id)
            })
        
        return jsonify({
            'success': True,
            'data': {
                'modules': modules,
                'test_types': test_types,
                'campuses': campuses,
                'courses': courses,
                'batches': batches,
                'levels': levels
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching filter options: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to fetch filter options: {str(e)}'
        }), 500

@superadmin_bp.route('/all-practice-results', methods=['GET'])
@jwt_required()
def get_all_practice_results():
    """Get all practice test results from both collections for super admin"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        all_results = []
        
        # Get from student_test_attempts collection
        try:
            if hasattr(mongo_db, 'student_test_attempts'):
                attempts = list(mongo_db.student_test_attempts.find({'test_type': 'practice'}))
                for attempt in attempts:
                    attempt['_id'] = str(attempt['_id'])
                    attempt['student_id'] = str(attempt['student_id'])
                    attempt['test_id'] = str(attempt['test_id'])
                    if attempt.get('submitted_at'):
                        attempt['submitted_at'] = safe_isoformat(attempt['submitted_at'])
                    attempt['source_collection'] = 'student_test_attempts'
                all_results.extend(attempts)
                current_app.logger.info(f"Found {len(attempts)} results in student_test_attempts")
            else:
                current_app.logger.warning("student_test_attempts collection not found")
        except Exception as e:
            current_app.logger.warning(f"Error reading from student_test_attempts: {e}")
        
        # Get from test_results collection
        try:
            if hasattr(mongo_db, 'test_results'):
                results = list(mongo_db.test_results.find({'test_type': 'practice'}))
                for result in results:
                    result['_id'] = str(result['_id'])
                    result['student_id'] = str(result['student_id'])
                    result['test_id'] = str(result['test_id'])
                    if result.get('submitted_at'):
                        result['submitted_at'] = safe_isoformat(result['submitted_at'])
                    result['source_collection'] = 'test_results'
                all_results.extend(results)
                current_app.logger.info(f"Found {len(results)} results in test_results")
            else:
                current_app.logger.warning("test_results collection not found")
        except Exception as e:
            current_app.logger.warning(f"Error reading from test_results: {e}")
        
        # Sort by submitted_at (most recent first)
        all_results.sort(key=lambda x: x.get('submitted_at', ''), reverse=True)
        
        # Final comprehensive ObjectId conversion for all results
        def final_convert_objectids(obj):
            if isinstance(obj, dict):
                for k, v in obj.items():
                    if isinstance(v, ObjectId):
                        obj[k] = str(v)
                    elif isinstance(v, (dict, list)):
                        final_convert_objectids(v)
            elif isinstance(obj, list):
                for i, item in enumerate(obj):
                    if isinstance(item, ObjectId):
                        obj[i] = str(item)
                    elif isinstance(item, (dict, list)):
                        final_convert_objectids(item)
        
        # Convert ObjectIds in all results
        for result in all_results:
            final_convert_objectids(result)
        
        return jsonify({
            'success': True,
            'data': all_results,
            'total_count': len(all_results)
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching all practice results: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to fetch practice results: {str(e)}'
        }), 500

@superadmin_bp.route('/student-progress/<student_id>', methods=['GET'])
@jwt_required()
def get_student_progress(student_id):
    """Get detailed progress for a specific student"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        if not user or user.get('role') not in ALLOWED_ADMIN_ROLES:
            return jsonify({
                'success': False,
                'message': 'Access denied. Admin privileges required.'
            }), 403
        
        # Get student details
        student_user = mongo_db.users.find_one({'_id': ObjectId(student_id)})
        if not student_user:
            return jsonify({
                'success': False,
                'message': 'Student not found'
            }), 404
        
        student_profile = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
        
        all_results = []
        
        # Get from student_test_attempts collection
        try:
            if hasattr(mongo_db, 'student_test_attempts'):
                attempts = list(mongo_db.student_test_attempts.find({'student_id': student_id, 'test_type': 'practice'}))
                for attempt in attempts:
                    attempt['_id'] = str(attempt['_id'])
                    attempt['test_id'] = str(attempt['test_id'])
                    if attempt.get('submitted_at'):
                        attempt['submitted_at'] = safe_isoformat(attempt['submitted_at'])
                    attempt['source_collection'] = 'student_test_attempts'
                all_results.extend(attempts)
        except Exception as e:
            current_app.logger.warning(f"Error reading from student_test_attempts: {e}")
        
        # Get from test_results collection
        try:
            if hasattr(mongo_db, 'test_results'):
                results = list(mongo_db.test_results.find({'student_id': ObjectId(student_id), 'test_type': 'practice'}))
                for result in results:
                    result['_id'] = str(result['_id'])
                    result['test_id'] = str(result['test_id'])
                    if result.get('submitted_at'):
                        result['submitted_at'] = safe_isoformat(result['submitted_at'])
                    result['source_collection'] = 'test_results'
                all_results.extend(results)
        except Exception as e:
            current_app.logger.warning(f"Error reading from test_results: {e}")
        
        # Calculate progress summary
        total_attempts = len(all_results)
        total_correct = sum(result.get('correct_answers', 0) for result in all_results)
        total_questions = sum(result.get('total_questions', 0) for result in all_results)
        average_score = sum(result.get('score_percentage', 0) for result in all_results) / total_attempts if total_attempts > 0 else 0
        
        # Group by module
        modules = {}
        for result in all_results:
            module_id = result.get('module_id', 'Unknown')
            if module_id not in modules:
                modules[module_id] = {
                    'module_id': module_id,
                    'module_name': MODULES.get(module_id, module_id),
                    'attempts': 0,
                    'total_score': 0,
                    'highest_score': 0,
                    'total_correct': 0,
                    'total_questions': 0
                }
            
            modules[module_id]['attempts'] += 1
            score = result.get('score_percentage', 0)
            modules[module_id]['total_score'] += score
            modules[module_id]['highest_score'] = max(modules[module_id]['highest_score'], score)
            modules[module_id]['total_correct'] += result.get('correct_answers', 0)
            modules[module_id]['total_questions'] += result.get('total_questions', 0)
        
        # Calculate averages for each module
        for module in modules.values():
            module['average_score'] = module['total_score'] / module['attempts'] if module['attempts'] > 0 else 0
            module['accuracy'] = (module['total_correct'] / module['total_questions'] * 100) if module['total_questions'] > 0 else 0
        
        progress_summary = {
            'student_info': {
                'name': student_user.get('name', 'Unknown'),
                'email': student_user.get('email', 'Unknown'),
                'roll_number': student_profile.get('roll_number', 'Unknown') if student_profile else 'Unknown'
            },
            'overall_stats': {
                'total_attempts': total_attempts,
                'total_correct_answers': total_correct,
                'total_questions': total_questions,
                'overall_accuracy': (total_correct / total_questions * 100) if total_questions > 0 else 0,
                'average_score': average_score
            },
            'modules': list(modules.values()),
            'recent_attempts': sorted(all_results, key=lambda x: x.get('submitted_at', ''), reverse=True)[:10]
        }
        
        return jsonify({
            'success': True,
            'data': progress_summary
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching student progress: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Failed to fetch student progress: {str(e)}'
        }), 500 