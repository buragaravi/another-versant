from flask import Blueprint, request, jsonify, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from mongo import mongo_db
from bson import ObjectId
import csv
import openpyxl
from werkzeug.utils import secure_filename
from config.constants import ROLES
from datetime import datetime
import pytz
import io
from utils.email_service import send_email, render_template
from config.shared import bcrypt
from socketio_instance import socketio
from routes.access_control import require_permission

batch_management_bp = Blueprint('batch_management', __name__)

def safe_isoformat(date_obj):
    """Safely convert a date object to ISO format string, handling various types."""
    if not date_obj:
        return None
    
    if hasattr(date_obj, 'isoformat'):
        # It's a datetime object
        return date_obj.isoformat()
    elif isinstance(date_obj, dict):
        # It's a MongoDB date dict, extract the date
        if '$date' in date_obj:
            try:
                from datetime import datetime
                date_str = date_obj['$date']
                # Handle different MongoDB date formats
                if 'T' in date_str:
                    # ISO format with T
                    date_obj_parsed = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                else:
                    # Just timestamp
                    date_obj_parsed = datetime.fromtimestamp(int(date_str) / 1000)
                return date_obj_parsed.isoformat()
            except (ValueError, KeyError, TypeError):
                return str(date_obj)
        else:
            return str(date_obj)
    else:
        # It's already a string or other type
        return str(date_obj)

@batch_management_bp.route('/', methods=['GET'])
@jwt_required()
@require_permission(module='batch_management')
def get_batches():
    """Get all batches"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        # Super admin can see all batches
        if user.get('role') == 'superadmin':
            batches = list(mongo_db.batches.find())
        else:
            # Campus and course admins can only see batches in their campus
            campus_id = user.get('campus_id')
            if not campus_id:
                return jsonify({'success': False, 'message': 'No campus assigned'}), 400
            batches = list(mongo_db.batches.find({'campus_ids': ObjectId(campus_id)}))
        
        batch_list = []
        for batch in batches:
            # Get course details
            course_objs = list(mongo_db.courses.find({'_id': {'$in': batch.get('course_ids', [])}}))
            
            # Handle both old and new batch structures for campus data
            campus_ids = batch.get('campus_ids', [])
            if not campus_ids and batch.get('campus_id'):
                # Handle old structure with single campus_id
                campus_ids = [batch.get('campus_id')]
            
            # Get campus details
            campus_objs = list(mongo_db.campuses.find({'_id': {'$in': campus_ids}}))
            student_count = mongo_db.students.count_documents({'batch_id': batch['_id']})
            
            batch_list.append({
                'id': str(batch['_id']),
                'name': batch.get('name'),
                'campuses': [{'id': str(c['_id']), 'name': c['name']} for c in campus_objs],
                'courses': [{'id': str(c['_id']), 'name': c['name']} for c in course_objs],
                'student_count': student_count,
                'created_at': batch.get('created_at')
            })
        
        return jsonify({'success': True, 'data': batch_list}), 200
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/', methods=['POST'])
@jwt_required()
@require_permission(module='batch_management', action='create_batch')
def create_batch_from_selection():
    """Create a new batch from selected campuses and courses"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        # Check if user has permission to create batches
        if not user or user.get('role') not in ['superadmin', 'campus_admin', 'course_admin']:
            return jsonify({
                'success': False,
                'message': 'Access denied. You do not have permission to create batches.'
            }), 403
        
        data = request.get_json()
        name = data.get('name')
        campus_ids = [ObjectId(cid) for cid in data.get('campus_ids', [])]
        course_ids = [ObjectId(cid) for cid in data.get('course_ids', [])]

        if not name or not campus_ids or not course_ids:
            return jsonify({'success': False, 'message': 'Name, campuses, and courses are required'}), 400

        if mongo_db.batches.find_one({'name': name}):
            return jsonify({'success': False, 'message': 'Batch name already exists'}), 409

        # Campus and course admins can only create batches in their own campus
        if user.get('role') in ['campus_admin', 'course_admin']:
            user_campus_id = user.get('campus_id')
            if not user_campus_id or str(user_campus_id) not in [str(cid) for cid in campus_ids]:
                return jsonify({
                    'success': False,
                    'message': 'Access denied. You can only create batches in your own campus.'
                }), 403

        # Create the batch
        batch_id = mongo_db.batches.insert_one({
            'name': name,
            'campus_ids': campus_ids,
            'course_ids': course_ids,
            'created_at': datetime.now(pytz.utc)
        }).inserted_id

        # Create batch-course instances for each course
        created_instances = []
        for course_id in course_ids:
            instance_id = mongo_db.find_or_create_batch_course_instance(batch_id, course_id)
            created_instances.append({
                'batch_id': str(batch_id),
                'course_id': str(course_id),
                'instance_id': str(instance_id)
            })

        return jsonify({
            'success': True,
            'message': 'Batch created successfully',
            'data': {
                'batch_id': str(batch_id),
                'instances': created_instances
            }
        }), 201
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/course/<course_id>/batches', methods=['GET'])
@jwt_required()
def get_batches_for_course(course_id):
    try:
        batches = list(mongo_db.batches.find({'course_ids': ObjectId(course_id)}))
        batch_list = []
        for batch in batches:
            campus_objs = list(mongo_db.campuses.find({'_id': {'$in': batch.get('campus_ids', [])}}))
            # Get student count for this batch
            student_count = mongo_db.students.count_documents({'batch_id': batch['_id']})
            batch_list.append({
                'id': str(batch['_id']),
                'name': batch['name'],
                'campuses': [{'id': str(c['_id']), 'name': c['name']} for c in campus_objs],
                'student_count': student_count
            })
        return jsonify({'success': True, 'data': batch_list}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching batches for course: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/campus/<campus_id>/batches', methods=['GET'])
@jwt_required()
def get_batches_for_campus(campus_id):
    try:
        batches = list(mongo_db.batches.find({'campus_ids': ObjectId(campus_id)}))
        batch_list = []
        for batch in batches:
            campus_objs = list(mongo_db.campuses.find({'_id': {'$in': batch.get('campus_ids', [])}}))
            course_objs = list(mongo_db.courses.find({'_id': {'$in': batch.get('course_ids', [])}}))
            # Get student count for this batch
            student_count = mongo_db.students.count_documents({'batch_id': batch['_id']})
            batch_list.append({
                'id': str(batch['_id']),
                'name': batch['name'],
                'campuses': [{'id': str(c['_id']), 'name': c['name']} for c in campus_objs],
                'courses': [{'id': str(c['_id']), 'name': c['name']} for c in course_objs],
                'student_count': student_count
            })
        return jsonify({'success': True, 'data': batch_list}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching batches for campus: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/create', methods=['POST'])
@jwt_required()
def create_batch():
    """Create a new batch and upload student data from an Excel file - SUPER ADMIN ONLY"""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        # Only super admin can create batches
        if not user or user.get('role') != 'superadmin':
            return jsonify({
                'success': False,
                'message': 'Access denied. Only super admin can create batches.'
            }), 403
        
        if 'student_file' not in request.files:
            return jsonify({'success': False, 'message': 'No student file part'}), 400

        file = request.files['student_file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No selected file'}), 400

        batch_name = request.form.get('batch_name')
        campus_id = request.form.get('campus_id')
        course_id = request.form.get('course_id')

        if not all([batch_name, campus_id, course_id]):
            return jsonify({'success': False, 'message': 'Missing batch name, campus ID, or course ID'}), 400

        # Create a new batch
        batch_doc = {
            'name': batch_name,
            'campus_id': ObjectId(campus_id),
            'course_id': ObjectId(course_id),
            'campus_ids': [ObjectId(campus_id)],
            'course_ids': [ObjectId(course_id)],
            'created_at': datetime.now(pytz.utc)
        }
        new_batch_id = mongo_db.batches.insert_one(batch_doc).inserted_id

        # Create batch-course instance
        instance_id = mongo_db.find_or_create_batch_course_instance(new_batch_id, ObjectId(course_id))

        # Process student file
        workbook = openpyxl.load_workbook(file, data_only=True)
        worksheet = workbook.active
        
        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            headers.append(str(cell.value).strip() if cell.value else '')
        
        # Get data rows
        students_data = []
        for row in worksheet.iter_rows(min_row=2):
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = str(cell.value).strip() if cell.value else ''
            students_data.append(row_data)

        created_students = []
        for student in students_data:
            # Generate username and password
            username = str(student['roll_number'])
            password = f"{student['student_name'].split()[0][:4].lower()}{student['roll_number'][-4:]}"
            
            # Use bcrypt for hashing
            password_hash = bcrypt.generate_password_hash(password).decode('utf-8')

            student_doc = {
                'name': student['student_name'],
                'email': student['email_id'],
                'roll_number': str(student['roll_number']),
                'username': username,
                'password_hash': password_hash,
                'role': ROLES['STUDENT'],
                'campus_id': ObjectId(campus_id),
                'course_id': ObjectId(course_id),
                'batch_id': new_batch_id,
                'batch_course_instance_id': instance_id,  # Link to instance
                'is_active': True,
                'created_at': datetime.now(pytz.utc),
                'mfa_enabled': False
            }
            mongo_db.users.insert_one(student_doc)
            
            # Create student profile with instance link
            student_profile = {
                'user_id': student_doc['_id'],
                'name': student['student_name'],
                'roll_number': str(student['roll_number']),
                'email': student['email_id'],
                'campus_id': ObjectId(campus_id),
                'course_id': ObjectId(course_id),
                'batch_id': new_batch_id,
                'batch_course_instance_id': instance_id,  # Link to instance
                'created_at': datetime.now(pytz.utc)
            }
            mongo_db.students.insert_one(student_profile)
            
            created_students.append({
                "student_name": student['student_name'],
                "email_id": student['email_id'],
                "roll_number": str(student['roll_number']),
                "username": username,
                "password": password
            })

            # Send welcome email
            try:
                html_content = render_template(
                    'student_credentials.html',
                    params={
                        'name': student['student_name'],
                        'username': username,
                        'email': student['email_id'],
                        'password': password,
                        'login_url': "https://pydah-studyedge.vercel.app/login"
                    }
                )
                send_email(
                    to_email=student['email_id'],
                    to_name=student['student_name'],
                    subject="Welcome to VERSANT - Your Student Credentials",
                    html_content=html_content
                )
            except Exception as e:
                print(f"Failed to send welcome email to {student['email_id']}: {e}")

        return jsonify({
            'success': True,
            'message': 'Batch and students created successfully',
            'data': {
                'batch_id': str(new_batch_id),
                'instance_id': str(instance_id),
                'created_students': created_students
            }
        }), 201
    except Exception as e:
        return jsonify({'success': False, 'message': f'An error occurred: {str(e)}'}), 500

@batch_management_bp.route('/<batch_id>', methods=['PUT'])
@jwt_required()
def edit_batch(batch_id):
    try:
        data = request.get_json()
        name = data.get('name')

        if not name:
            return jsonify({'success': False, 'message': 'Batch name is required.'}), 400

        result = mongo_db.batches.update_one(
            {'_id': ObjectId(batch_id)},
            {'$set': {'name': name}}
        )

        if result.matched_count == 0:
            return jsonify({'success': False, 'message': 'Batch not found.'}), 404

        return jsonify({'success': True, 'message': 'Batch updated successfully.'}), 200
    except Exception as e:
        current_app.logger.error(f"Error updating batch: {e}")
        return jsonify({'success': False, 'message': 'An error occurred while updating the batch.'}), 500

@batch_management_bp.route('/<batch_id>', methods=['DELETE'])
@jwt_required()
def delete_batch(batch_id):
    try:
        batch_obj_id = ObjectId(batch_id)
        
        # Find all students in this batch to get their user_ids
        students_to_delete = list(mongo_db.students.find({'batch_id': batch_obj_id}))
        user_ids_to_delete = [s['user_id'] for s in students_to_delete]
        
        # Delete associated users
        if user_ids_to_delete:
            mongo_db.users.delete_many({'_id': {'$in': user_ids_to_delete}})
        
        # Delete student records
        mongo_db.students.delete_many({'batch_id': batch_obj_id})
        
        # Finally, delete the batch
        result = mongo_db.batches.delete_one({'_id': batch_obj_id})
        
        if result.deleted_count == 0:
            return jsonify({'success': False, 'message': 'Batch not found or already deleted.'}), 404
            
        return jsonify({'success': True, 'message': 'Batch and all associated students have been deleted.'}), 200
        
    except Exception as e:
        current_app.logger.error(f"Error deleting batch: {e}")
        return jsonify({'success': False, 'message': f'An error occurred: {str(e)}'}), 500

@batch_management_bp.route('/campuses', methods=['GET'])
@jwt_required()
def get_campuses():
    campuses = list(mongo_db.campuses.find())
    return jsonify({'success': True, 'data': [{'id': str(c['_id']), 'name': c['name']} for c in campuses]}), 200

@batch_management_bp.route('/courses', methods=['GET'])
@jwt_required()
def get_courses():
    campus_ids = request.args.getlist('campus_ids')
    if not campus_ids:
        return jsonify({'success': False, 'message': 'campus_ids required'}), 400
    try:
        campus_obj_ids = [ObjectId(cid) for cid in campus_ids]
        
        pipeline = [
            {'$match': {'campus_id': {'$in': campus_obj_ids}}},
            {
                '$lookup': {
                    'from': 'campuses',
                    'localField': 'campus_id',
                    'foreignField': '_id',
                    'as': 'campus_info'
                }
            },
            {
                '$unwind': '$campus_info'
            }
        ]
        courses = list(mongo_db.courses.aggregate(pipeline))

        courses_data = [{
            'id': str(c['_id']),
            'name': c['name'],
            'campus_id': str(c['campus_id']),
            'campus_name': c['campus_info']['name']
        } for c in courses]
        
        return jsonify({'success': True, 'data': courses_data}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching courses by campus: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

def _parse_student_file(file):
    filename = secure_filename(file.filename)
    if not (filename.endswith('.csv') or filename.endswith('.xlsx')):
        raise ValueError('Please upload a valid CSV or Excel file.')
    
    try:
        if filename.endswith('.csv'):
            # Read CSV file
            content = file.read().decode('utf-8-sig')
            csv_reader = csv.DictReader(io.StringIO(content))
            rows = list(csv_reader)
            print(f"CSV parsing: Found {len(rows)} rows, headers: {list(rows[0].keys()) if rows else []}")
        else:
            # Read Excel file
            workbook = openpyxl.load_workbook(file, data_only=True)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                headers.append(str(cell.value).strip() if cell.value else '')
            
            # Get data rows
            rows = []
            for row in worksheet.iter_rows(min_row=2):
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = str(cell.value).strip() if cell.value else ''
                rows.append(row_data)
            print(f"Excel parsing: Found {len(rows)} rows, headers: {headers}")
    except Exception as e:
        print(f"File parsing error: {e}")
        raise ValueError(f"Error reading file: {e}")

    return rows

@batch_management_bp.route('/test-file-parse', methods=['POST'])
@jwt_required()
def test_file_parse():
    """Test endpoint to debug file parsing"""
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({'success': False, 'message': 'No file provided'}), 400

        rows = _parse_student_file(file)
        
        return jsonify({
            'success': True,
            'data': {
                'row_count': len(rows),
                'columns': list(rows[0].keys()) if rows else [],
                'first_row': rows[0] if rows else None,
                'sample_rows': rows[:3] if len(rows) > 3 else rows
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/validate-student-upload', methods=['POST'])
@jwt_required()
def validate_student_upload():
    try:
        file = request.files.get('file')
        campus_id = request.form.get('campus_id')

        if not file or not campus_id:
            return jsonify({'success': False, 'message': 'A file and campus ID are required.'}), 400

        rows = _parse_student_file(file)

        if not rows:
            return jsonify({'success': False, 'message': 'File is empty or invalid.'}), 400

        # Get column names from first row
        columns = list(rows[0].keys()) if rows else []
        
        required_fields = ['Campus Name', 'Course Name', 'Student Name', 'Roll Number', 'Email']
        missing_fields = [field for field in required_fields if field not in columns]
        if missing_fields:
            return jsonify({'success': False, 'message': f"Invalid file structure. Missing columns: {', '.join(missing_fields)}"}), 400

        # Fetch existing data for validation
        existing_roll_numbers = set(s['roll_number'] for s in mongo_db.students.find({}, {'roll_number': 1}))
        existing_emails = set(u['email'] for u in mongo_db.users.find({}, {'email': 1}))
        existing_mobile_numbers = set(u.get('mobile_number', '') for u in mongo_db.users.find({'mobile_number': {'$exists': True, '$ne': ''}}, {'mobile_number': 1}))
        
        # Get campus info for validation
        campus = mongo_db.campuses.find_one({'_id': ObjectId(campus_id)})
        if not campus:
            return jsonify({'success': False, 'message': 'Invalid campus ID'}), 400
        
        campus_courses = list(mongo_db.courses.find({'campus_id': ObjectId(campus_id)}, {'name': 1}))
        valid_course_names = {course['name'] for course in campus_courses}

        preview_data = []
        for index, row in enumerate(rows):
            student_data = {
                'campus_name': str(row.get('Campus Name', '')).strip(),
                'course_name': str(row.get('Course Name', '')).strip(),
                'student_name': str(row.get('Student Name', '')).strip(),
                'roll_number': str(row.get('Roll Number', '')).strip(),
                'email': str(row.get('Email', '')).strip().lower(),
                'mobile_number': str(row.get('Mobile Number', '')).strip(),
            }
            
            errors = []
            if not all([student_data['campus_name'], student_data['course_name'], student_data['student_name'], student_data['roll_number'], student_data['email']]):
                errors.append('Missing required fields.')
            if student_data['roll_number'] in existing_roll_numbers:
                errors.append('Roll number already exists.')
            if student_data['email'] in existing_emails:
                errors.append('Email already exists.')
            if student_data['mobile_number'] and student_data['mobile_number'] in existing_mobile_numbers:
                errors.append('Mobile number already exists.')
            if student_data['campus_name'] != campus['name']:
                errors.append(f"Campus '{student_data['campus_name']}' doesn't match selected campus '{campus['name']}'.")
            if student_data['course_name'] not in valid_course_names:
                errors.append(f"Course '{student_data['course_name']}' not found in this campus.")

            student_data['errors'] = errors
            preview_data.append(student_data)
        
        return jsonify({'success': True, 'data': preview_data})

    except ValueError as ve:
        return jsonify({'success': False, 'message': str(ve)}), 400
    except Exception as e:
        current_app.logger.error(f"Error validating student upload: {e}")
        return jsonify({'success': False, 'message': f'An unexpected error occurred: {e}'}), 500

@batch_management_bp.route('/upload-students', methods=['POST'])
@jwt_required()
def upload_students_to_batch():
    try:
        # Accept file upload and form data
        file = request.files.get('file')
        batch_id = request.form.get('batch_id')
        course_ids = request.form.getlist('course_ids')  # Accept multiple course IDs
        user_id = get_jwt_identity()  # Get current user ID for progress updates
        
        # Debug logging
        current_app.logger.info(f"Upload request - batch_id: {batch_id}, course_ids: {course_ids}, filename: {file.filename if file else 'None'}")
        
        if not file or not batch_id or not course_ids:
            return jsonify({'success': False, 'message': 'File, batch ID, and at least one course ID are required.'}), 400

        rows = _parse_student_file(file)
        current_app.logger.info(f"Parsed {len(rows)} rows from file")
        
        if not rows:
            return jsonify({'success': False, 'message': 'File is empty or invalid.'}), 400

        # Validate columns - support both formats
        columns = list(rows[0].keys()) if rows else []
        current_app.logger.info(f"File columns: {columns}")
        
        required_fields_v1 = ['Student Name', 'Roll Number', 'Email', 'Mobile Number']
        required_fields_v2 = ['Group', 'Roll Number', 'Student Name', 'Email', 'Mobile Number']
        
        missing_fields_v1 = [field for field in required_fields_v1 if field not in columns]
        missing_fields_v2 = [field for field in required_fields_v2 if field not in columns]
        
        current_app.logger.info(f"Missing v1 fields: {missing_fields_v1}, Missing v2 fields: {missing_fields_v2}")
        
        if missing_fields_v1 and missing_fields_v2:
            return jsonify({'success': False, 'message': f"Invalid file structure. Expected either: {', '.join(required_fields_v1)} OR {', '.join(required_fields_v2)}. Found columns: {', '.join(columns)}"}), 400

        # Determine file format
        is_v2_format = 'Group' in columns
        current_app.logger.info(f"Using format v2: {is_v2_format}")

        # Fetch batch and campus info
        batch = mongo_db.batches.find_one({'_id': ObjectId(batch_id)})
        if not batch:
            return jsonify({'success': False, 'message': 'Batch not found.'}), 404
        campus_ids = batch.get('campus_ids', [])
        if not campus_ids:
            return jsonify({'success': False, 'message': 'Batch is missing campus info.'}), 400
        campus_id = campus_ids[0]  # Assume single campus per batch for now

        # Validate course IDs
        valid_course_ids = set(str(cid) for cid in batch.get('course_ids', []))
        for cid in course_ids:
            if cid not in valid_course_ids:
                return jsonify({'success': False, 'message': f'Course ID {cid} is not valid for this batch.'}), 400

        # Fetch existing data for validation
        existing_roll_numbers = set(s['roll_number'] for s in mongo_db.students.find({}, {'roll_number': 1}))
        existing_emails = set(u['email'] for u in mongo_db.users.find({}, {'email': 1}))
        existing_mobile_numbers = set(u.get('mobile_number', '') for u in mongo_db.users.find({'mobile_number': {'$exists': True, '$ne': ''}}, {'mobile_number': 1}))

        created_students = []
        errors = []
        uploaded_emails = []  # Track emails for verification
        total_students = len(rows)
        
        # Send initial progress update
        socketio.emit('upload_progress', {
            'user_id': user_id,
            'status': 'started',
            'total': total_students,
            'processed': 0,
            'percentage': 0,
            'message': 'Starting student upload...'
        }, room=str(user_id))
        
        for index, row in enumerate(rows):
            if is_v2_format:
                student_name = str(row.get('Student Name', '')).strip()
                roll_number = str(row.get('Roll Number', '')).strip()
                email = str(row.get('Email', '')).strip().lower()
                mobile_number = str(row.get('Mobile Number', '')).strip()
                group_name = str(row.get('Group', '')).strip()
                
                # Find course by group name (assuming group name matches course name)
                # Try exact match first, then case-insensitive match
                course = mongo_db.courses.find_one({'name': group_name, '_id': {'$in': [ObjectId(cid) for cid in course_ids]}})
                if not course:
                    # Try case-insensitive match
                    course = mongo_db.courses.find_one({
                        'name': {'$regex': f'^{group_name}$', '$options': 'i'}, 
                        '_id': {'$in': [ObjectId(cid) for cid in course_ids]}
                    })
                
                if not course:
                    # Get available courses for better error message
                    available_courses = list(mongo_db.courses.find({'_id': {'$in': [ObjectId(cid) for cid in course_ids]}}, {'name': 1}))
                    available_names = [c['name'] for c in available_courses]
                    errors.append(f"{student_name}: Course/Group '{group_name}' not found in this batch. Available courses: {', '.join(available_names)}")
                    continue
                course_id = str(course['_id'])
            else:
                student_name = str(row.get('Student Name', '')).strip()
                roll_number = str(row.get('Roll Number', '')).strip()
                email = str(row.get('Email', '')).strip().lower()
                mobile_number = str(row.get('Mobile Number', '')).strip()
                course_id = course_ids[0]  # Use first course for v1 format

            # Validation
            errs = []
            if not all([student_name, roll_number, email]):
                errs.append('Missing required fields.')
            if roll_number in existing_roll_numbers:
                errs.append('Roll number already exists.')
            if email in existing_emails:
                errs.append('Email already exists.')
            if mobile_number and mobile_number in existing_mobile_numbers:
                errs.append('Mobile number already exists.')
            if errs:
                errors.append(f"{student_name or roll_number or email}: {', '.join(errs)}")
                continue

            try:
                username = roll_number
                password = f"{student_name.split()[0][:4].lower()}{roll_number[-4:]}"
                password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
                
                user_doc = {
                    'username': username,
                    'email': email,
                    'password_hash': password_hash,
                    'role': 'student',
                    'name': student_name,
                    'mobile_number': mobile_number,
                    'campus_id': campus_id,
                    'course_id': ObjectId(course_id),
                    'batch_id': ObjectId(batch_id),
                    'is_active': True,
                    'created_at': datetime.now(pytz.utc),
                    'mfa_enabled': False
                }
                
                # Insert user first
                user_result = mongo_db.users.insert_one(user_doc)
                if not user_result.inserted_id:
                    errors.append(f"{student_name}: Failed to create user account.")
                    continue
                
                user_id = user_result.inserted_id
                
                student_doc = {
                    'user_id': user_id,
                    'name': student_name,
                    'roll_number': roll_number,
                    'email': email,
                    'mobile_number': mobile_number,
                    'campus_id': campus_id,
                    'course_id': ObjectId(course_id),
                    'batch_id': ObjectId(batch_id),
                    'created_at': datetime.now(pytz.utc)
                }
                
                # Insert student profile
                student_result = mongo_db.students.insert_one(student_doc)
                if not student_result.inserted_id:
                    # Rollback user creation if student creation fails
                    mongo_db.users.delete_one({'_id': user_id})
                    errors.append(f"{student_name}: Failed to create student profile.")
                    continue
                
                created_students.append({
                    'name': student_name,
                    'email': email,
                    'username': username,
                    'password': password
                })
                uploaded_emails.append(email)
                
                # Update existing sets to prevent duplicates within the same upload
                existing_roll_numbers.add(roll_number)
                existing_emails.add(email)
                if mobile_number:
                    existing_mobile_numbers.add(mobile_number)
                
                # Send welcome email (non-blocking - don't fail the whole process if email fails)
                email_sent = False
                email_error = None
                try:
                    html_content = render_template(
                        'student_credentials.html',
                        params={
                            'name': student_name,
                            'username': username,
                            'email': email,
                            'password': password,
                            'login_url': "https://pydah-studyedge.vercel.app/login"
                        }
                    )
                    send_email(
                        to_email=email,
                        to_name=student_name,
                        subject="Welcome to Study Edge - Your Student Credentials",
                        html_content=html_content
                    )
                    email_sent = True
                except Exception as e:
                    email_error = str(e)
                    # Don't add to errors array - just log it
                    current_app.logger.error(f"Failed to send email to {email}: {e}")
                
                # Send progress update after student creation (regardless of email status)
                percentage = int(((index + 1) / total_students) * 100)
                if email_sent:
                    socketio.emit('upload_progress', {
                        'user_id': user_id,
                        'status': 'processing',
                        'total': total_students,
                        'processed': index + 1,
                        'percentage': percentage,
                        'message': f'Student created and email sent to {student_name} ({email})',
                        'current_student': {
                            'name': student_name,
                            'email': email,
                            'username': username
                        }
                    }, room=str(user_id))
                else:
                    socketio.emit('upload_progress', {
                        'user_id': user_id,
                        'status': 'processing',
                        'total': total_students,
                        'processed': index + 1,
                        'percentage': percentage,
                        'message': f'Student created successfully for {student_name} ({email}) - Email sending failed',
                        'current_student': {
                            'name': student_name,
                            'email': email,
                            'username': username
                        },
                        'email_warning': True,
                        'email_error': email_error
                    }, room=str(user_id))
                    
            except Exception as e:
                errors.append(f"{student_name}: Database error - {str(e)}")
                continue

        # Verify upload success
        verification_results = []
        if uploaded_emails:
            students_verified = list(mongo_db.students.find({'email': {'$in': uploaded_emails}, 'batch_id': ObjectId(batch_id)}))
            users_verified = list(mongo_db.users.find({'email': {'$in': uploaded_emails}, 'batch_id': ObjectId(batch_id)}))
            
            for email in uploaded_emails:
                student_exists = any(s['email'] == email for s in students_verified)
                user_exists = any(u['email'] == email for u in users_verified)
                verification_results.append({
                    'email': email,
                    'student_profile_exists': student_exists,
                    'user_account_exists': user_exists,
                    'fully_uploaded': student_exists and user_exists
                })

        # Send completion progress update
        if errors:
            socketio.emit('upload_progress', {
                'user_id': user_id,
                'status': 'completed_with_errors',
                'total': total_students,
                'processed': total_students,
                'percentage': 100,
                'message': f'Upload completed with {len(errors)} errors. {len(created_students)} students uploaded successfully.',
                'errors': errors,
                'created_count': len(created_students)
            }, room=str(user_id))
            
            return jsonify({
                'success': bool(created_students), 
                'message': f"Upload completed with {len(errors)} errors.", 
                'data': {
                    'created_students': created_students,
                    'verification_results': verification_results
                }, 
                'errors': errors
            }), 207
        
        # Send success completion update
        socketio.emit('upload_progress', {
            'user_id': user_id,
            'status': 'completed',
            'total': total_students,
            'processed': total_students,
            'percentage': 100,
            'message': f'Successfully uploaded {len(created_students)} students to the batch!',
            'created_count': len(created_students)
        }, room=str(user_id))
        
        return jsonify({
            'success': True, 
            'message': f"Successfully uploaded {len(created_students)} students to the batch.", 
            'data': {
                'created_students': created_students,
                'verification_results': verification_results
            }
        }), 201

    except Exception as e:
        current_app.logger.error(f"Error uploading students to batch: {e}")
        import traceback
        current_app.logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'message': f'Failed to upload students. Please check your file format and try again. Error: {str(e)}'}), 500

@batch_management_bp.route('/batch/<batch_id>/students', methods=['GET'])
@jwt_required()
@require_permission(module='batch_management')
def get_batch_students(batch_id):
    """Get all students and detailed info for a specific batch."""
    try:
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        batch = mongo_db.batches.find_one({'_id': ObjectId(batch_id)})
        if not batch:
            return jsonify({'success': False, 'message': 'Batch not found'}), 404

        # Check if user has access to this batch
        if user.get('role') in ['campus_admin', 'course_admin']:
            user_campus_id = user.get('campus_id')
            batch_campus_ids = batch.get('campus_ids', [])
            
            # Handle both old and new batch structures
            if not batch_campus_ids and batch.get('campus_id'):
                batch_campus_ids = [batch.get('campus_id')]
            
            if not user_campus_id or ObjectId(user_campus_id) not in batch_campus_ids:
                return jsonify({'success': False, 'message': 'Access denied. You do not have permission to view this batch.'}), 403

        # Fetch campus and course names for the batch header
        campus_ids = batch.get('campus_ids', [])
        course_ids = batch.get('course_ids', [])
        
        campuses = list(mongo_db.campuses.find({'_id': {'$in': campus_ids}}))
        courses = list(mongo_db.courses.find({'_id': {'$in': course_ids}}))

        batch_info = {
            'id': str(batch['_id']),
            'name': batch.get('name'),
            'campus_name': ', '.join([c['name'] for c in campuses]),
            'course_name': ', '.join([c['name'] for c in courses]),
            'course_ids': [str(c['_id']) for c in courses],
        }

        # Fetch students with populated info using the new method
        students_with_details = mongo_db.get_students_by_batch(batch_id)

        return jsonify({
            'success': True,
            'data': students_with_details,
            'batch_info': batch_info
        }), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching batch students: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/students/filtered', methods=['GET'])
@jwt_required()
@require_permission(module='batch_management')
def get_filtered_students():
    """Get filtered students with pagination"""
    try:
        current_app.logger.info("get_filtered_students endpoint called")
        print("get_filtered_students endpoint called")
        current_user_id = get_jwt_identity()
        user = mongo_db.find_user_by_id(current_user_id)
        
        # Get query parameters
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        search = request.args.get('search', '')
        campus_id = request.args.get('campus_id', '')
        course_id = request.args.get('course_id', '')
        batch_id = request.args.get('batch_id', '')
        
        # Build query
        query = {'role': 'student'}
        
        if search:
            query['$or'] = [
                {'name': {'$regex': search, '$options': 'i'}},
                {'email': {'$regex': search, '$options': 'i'}}
            ]
        
        if campus_id:
            query['campus_id'] = ObjectId(campus_id)
        
        if course_id:
            query['course_id'] = ObjectId(course_id)
        
        if batch_id:
            query['batch_id'] = ObjectId(batch_id)
        
        # Super admin can see all students, others only their campus
        if user.get('role') != 'superadmin':
            user_campus_id = user.get('campus_id')
            if user_campus_id:
                query['campus_id'] = ObjectId(user_campus_id)
        
        # Get total count
        total = mongo_db.users.count_documents(query)
        
        # Get paginated results
        skip = (page - 1) * limit
        students = list(mongo_db.users.find(query).skip(skip).limit(limit))
        
        # Get additional student details
        student_details = []
        for student in students:
            # Get student profile
            student_profile = mongo_db.students.find_one({'user_id': student['_id']})
            
            # Get campus and course names
            campus = mongo_db.campuses.find_one({'_id': student.get('campus_id')})
            course = mongo_db.courses.find_one({'_id': student.get('course_id')})
            batch = mongo_db.batches.find_one({'_id': student.get('batch_id')})
            
            student_details.append({
                '_id': str(student['_id']),
                'name': student.get('name', ''),
                'email': student.get('email', ''),
                'roll_number': student_profile.get('roll_number', '') if student_profile else '',
                'mobile_number': student_profile.get('mobile_number', '') if student_profile else '',
                'campus_name': campus.get('name', '') if campus else '',
                'course_name': course.get('name', '') if course else '',
                'batch_name': batch.get('name', '') if batch else '',
                'is_active': student.get('is_active', True),
                'created_at': student.get('created_at')
            })
        
        return jsonify({
            'success': True,
            'data': student_details,
            'pagination': {
                'page': page,
                'limit': limit,
                'total': total,
                'has_more': (page * limit) < total
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching filtered students: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/student/<student_id>', methods=['GET'])
@jwt_required()
def get_student_details(student_id):
    try:
        student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        user = mongo_db.users.find_one({'_id': student['user_id']})
        campus = mongo_db.campuses.find_one({'_id': student['campus_id']})
        course = mongo_db.courses.find_one({'_id': student['course_id']})
        batch = mongo_db.batches.find_one({'_id': student['batch_id']})

        student_details = {
            'id': str(student['_id']),
            'name': student['name'],
            'roll_number': student['roll_number'],
            'email': student['email'],
            'mobile_number': student['mobile_number'],
            'campus_name': campus['name'] if campus else 'N/A',
            'course_name': course['name'] if course else 'N/A',
            'batch_name': batch['name'] if batch else 'N/A',
            'username': user.get('username', 'N/A')
        }
        return jsonify({'success': True, 'data': student_details}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching student details: {e}")
        return jsonify({'success': False, 'message': 'An error occurred fetching student details.'}), 500

@batch_management_bp.route('/student/<student_id>', methods=['PUT'])
@jwt_required()
def update_student(student_id):
    try:
        data = request.json
        student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        # Uniqueness check for email and mobile_number (across all users except this one)
        email = data.get('email', student['email'])
        mobile_number = data.get('mobile_number', student['mobile_number'])
        user = mongo_db.users.find_one({'_id': student['user_id']})
        if not user:
            return jsonify({'success': False, 'message': 'User not found'}), 404

        existing_user = mongo_db.users.find_one({
            '_id': {'$ne': user['_id']},
            '$or': [
                {'email': email},
                {'mobile_number': mobile_number}
            ]
        })
        if existing_user:
            if existing_user['email'] == email:
                return jsonify({'success': False, 'message': 'Email already exists'}), 400
            return jsonify({'success': False, 'message': 'Mobile number already exists'}), 400

        # Update student collection
        student_update = {
            'name': data.get('name', student['name']),
            'roll_number': data.get('roll_number', student['roll_number']),
            'email': email,
            'mobile_number': mobile_number,
        }
        mongo_db.students.update_one({'_id': ObjectId(student_id)}, {'$set': student_update})

        # Update user collection
        user_update = {
            'name': data.get('name', student['name']),
            'email': email,
            'username': data.get('roll_number', student['roll_number']),
            'mobile_number': mobile_number
        }
        mongo_db.users.update_one({'_id': student['user_id']}, {'$set': user_update})

        return jsonify({'success': True, 'message': 'Student updated successfully'}), 200
    except Exception as e:
        current_app.logger.error(f"Error updating student: {e}")
        return jsonify({'success': False, 'message': 'An error occurred updating the student.'}), 500

@batch_management_bp.route('/student/<student_id>', methods=['DELETE'])
@jwt_required()
def delete_student(student_id):
    try:
        student = mongo_db.students.find_one_and_delete({'_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404

        # Also delete the associated user account
        mongo_db.users.delete_one({'_id': student['user_id']})

        return jsonify({'success': True, 'message': 'Student deleted successfully'}), 200
    except Exception as e:
        current_app.logger.error(f"Error deleting student: {e}")
        return jsonify({'success': False, 'message': 'An error occurred deleting the student.'}), 500

@batch_management_bp.route('/student/<student_id>/authorize-level', methods=['POST'])
@jwt_required()
def authorize_student_level(student_id):
    try:
        data = request.json
        level = data.get('level')
        if not level:
            return jsonify({'success': False, 'message': 'Level is required'}), 400
        # Ensure authorized_levels exists
        student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found.'}), 404
        if 'authorized_levels' not in student:
            mongo_db.students.update_one({'_id': ObjectId(student_id)}, {'$set': {'authorized_levels': []}})
        # Add the level to authorized_levels
        result = mongo_db.students.update_one({'_id': ObjectId(student_id)}, {'$addToSet': {'authorized_levels': level}})
        student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
        if result.modified_count == 0:
            return jsonify({'success': False, 'message': f"Level '{level}' was already authorized for student.", 'authorized_levels': student.get('authorized_levels', [])}), 200
        return jsonify({'success': True, 'message': f"Level '{level}' authorized for student.", 'authorized_levels': student.get('authorized_levels', [])}), 200
    except Exception as e:
        current_app.logger.error(f"Error authorizing level: {e}")
        return jsonify({'success': False, 'message': 'An error occurred authorizing the level.'}), 500

@batch_management_bp.route('/student/<student_id>/lock-level', methods=['POST'])
@jwt_required()
def lock_student_level(student_id):
    try:
        data = request.json
        level = data.get('level')
        if not level:
            return jsonify({'success': False, 'message': 'Level is required'}), 400

        # Find student by user_id or _id
        student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
        if not student:
            student = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found.'}), 404

        # Remove level from authorized_levels
        mongo_db.students.update_one({'_id': student['_id']}, {'$pull': {'authorized_levels': level}})
        student = mongo_db.students.find_one({'_id': student['_id']})

        # Emit real-time event to the student
        socketio.emit('level_access_changed', {'student_id': str(student['_id']), 'level': level, 'action': 'locked'}, room=str(student['_id']))

        return jsonify({'success': True, 'message': f"Level '{level}' locked for student.", 'authorized_levels': student.get('authorized_levels', [])}), 200
    except Exception as e:
        current_app.logger.error(f"Error locking level: {e}")
        return jsonify({'success': False, 'message': 'An error occurred locking the level.'}), 500

@batch_management_bp.route('/student/<student_id>/authorize-module', methods=['POST'])
@batch_management_bp.route('/student/<student_id>/authorize-module/', methods=['POST'])
@jwt_required()
def authorize_student_module(student_id):
    print(f"DEBUG: authorize_student_module route hit - student_id: {student_id}")
    print(f"DEBUG: Request method: {request.method}")
    print(f"DEBUG: Request URL: {request.url}")
    print(f"DEBUG: Request headers: {dict(request.headers)}")
    print(f"DEBUG: Request JSON: {request.json}")
    try:
        data = request.json
        module = data.get('module')
        if not module:
            return jsonify({'success': False, 'message': 'Module is required'}), 400

        # Find all levels for this module
        from config.constants import LEVELS
        module_levels = [level_id for level_id, level in LEVELS.items() if (level.get('module_id') if isinstance(level, dict) else None) == module]
        if not module_levels:
            return jsonify({'success': False, 'message': 'No levels found for this module.'}), 404

        # Ensure authorized_levels exists
        print(f"DEBUG: Looking for student with ID: {student_id}")
        print(f"DEBUG: Student ID type: {type(student_id)}")
        try:
            obj_id = ObjectId(student_id)
            print(f"DEBUG: Converted to ObjectId: {obj_id}")
        except Exception as e:
            print(f"DEBUG: Failed to convert to ObjectId: {e}")
            return jsonify({'success': False, 'message': f'Invalid student_id: {student_id}'}), 400
        student = mongo_db.students.find_one({'_id': obj_id})
        print(f"DEBUG: Student found: {student is not None}")
        if not student:
            print(f"DEBUG: No student found with ID: {student_id}")
            # Fallback: try user_id
            student = mongo_db.students.find_one({'user_id': obj_id})
            if student:
                print(f"DEBUG: Found student by user_id fallback: {student}")
            else:
                # Let's also check if there are any students in the database
                all_students = list(mongo_db.students.find({}, {'_id': 1, 'name': 1, 'email': 1}))
                print(f"DEBUG: Total students in database: {len(all_students)}")
                print(f"DEBUG: Sample student IDs: {[str(s['_id']) for s in all_students[:5]]}")
                return jsonify({'success': False, 'message': 'Student not found.'}), 404
        if 'authorized_levels' not in student:
            mongo_db.students.update_one({'_id': student['_id']}, {'$set': {'authorized_levels': []}})

        # Add all levels to authorized_levels
        mongo_db.students.update_one({'_id': student['_id']}, {'$addToSet': {'authorized_levels': {'$each': module_levels}}})
        student = mongo_db.students.find_one({'_id': student['_id']})

        # Emit real-time event to the student
        socketio.emit('module_access_changed', {'student_id': str(student['_id']), 'module': module, 'action': 'unlocked'}, room=str(student['_id']))

        return jsonify({'success': True, 'message': f"Module '{module}' authorized for student.", 'authorized_levels': student.get('authorized_levels', [])}), 200
    except Exception as e:
        current_app.logger.error(f"Error authorizing module: {e}")
        return jsonify({'success': False, 'message': 'An error occurred authorizing the module.'}), 500

@batch_management_bp.route('/student/<student_id>/lock-module', methods=['POST'])
@jwt_required()
def lock_student_module(student_id):
    try:
        data = request.json
        module = data.get('module')
        if not module:
            return jsonify({'success': False, 'message': 'Module is required'}), 400

        from config.constants import LEVELS
        module_levels = [level_id for level_id, level in LEVELS.items() if level.get('module_id') == module or level.get('module') == module]
        if not module_levels:
            return jsonify({'success': False, 'message': 'No levels found for this module.'}), 404

        # Remove all levels from authorized_levels
        student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
        if not student:
            student = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found.'}), 404
        mongo_db.students.update_one({'_id': student['_id']}, {'$pull': {'authorized_levels': {'$in': module_levels}}})
        student = mongo_db.students.find_one({'_id': student['_id']})

        # Emit real-time event to the student
        socketio.emit('module_access_changed', {'student_id': str(student['_id']), 'module': module, 'action': 'locked'}, room=str(student['_id']))

        return jsonify({'success': True, 'message': f"Module '{module}' locked for student.", 'authorized_levels': student.get('authorized_levels', [])}), 200
    except Exception as e:
        current_app.logger.error(f"Error locking module: {e}")
        return jsonify({'success': False, 'message': 'An error occurred locking the module.'}), 500

@batch_management_bp.route('/create-with-students', methods=['POST'])
@jwt_required()
def create_batch_with_students():
    try:
        data = request.get_json()
        name = data.get('name')
        campus_ids = [ObjectId(cid) for cid in data.get('campus_ids', [])]
        course_ids = [ObjectId(cid) for cid in data.get('course_ids', [])]
        students_data = data.get('students', [])

        if not name or not campus_ids or not students_data:
            return jsonify({'success': False, 'message': 'Batch name, campus, and student data are required.'}), 400
        
        if mongo_db.batches.find_one({'name': name}):
            return jsonify({'success': False, 'message': 'A batch with this name already exists.'}), 409

        # 1. Create the batch
        batch_doc = {
            'name': name,
            'campus_ids': campus_ids,
            'course_ids': course_ids,
            'created_at': datetime.now(pytz.utc)
        }
        batch_id = mongo_db.batches.insert_one(batch_doc).inserted_id

        # 2. Create students and users
        created_students_details = []
        errors = []
        
        for student in students_data:
            try:
                # Find campus by name from the uploaded file
                campus = mongo_db.campuses.find_one({'name': student['campus_name']})
                if not campus:
                    errors.append(f"Campus '{student['campus_name']}' not found for student '{student.get('student_name', 'N/A')}'.")
                    continue
                
                # Find course by name and campus_id
                course = mongo_db.courses.find_one({
                    'name': student['course_name'],
                    'campus_id': campus['_id']
                })
                if not course:
                    errors.append(f"Course '{student['course_name']}' not found in campus '{student['campus_name']}' for student '{student.get('student_name', 'N/A')}'.")
                    continue

                # Check for existing user with same roll number, email, or mobile number
                existing_user = mongo_db.users.find_one({
                    '$or': [
                        {'username': student['roll_number']},
                        {'email': student['email']}
                    ]
                })
                if existing_user:
                    errors.append(f"Student with roll number '{student['roll_number']}' or email '{student['email']}' already exists.")
                    continue

                # Check for duplicate mobile number if provided
                if student.get('mobile_number') and mongo_db.users.find_one({'mobile_number': student['mobile_number']}):
                    errors.append(f"Student with mobile number '{student['mobile_number']}' already exists.")
                    continue

                username = student['roll_number']
                password = f"{student['student_name'].split()[0][:4].lower()}{student['roll_number'][-4:]}"
                password_hash = bcrypt.generate_password_hash(password).decode('utf-8')

                user_doc = {
                    'username': username,
                    'email': student['email'],
                    'password_hash': password_hash,
                    'role': ROLES['STUDENT'],
                    'name': student['student_name'],
                    'mobile_number': student.get('mobile_number', ''),
                    'campus_id': campus['_id'],
                    'course_id': course['_id'],
                    'batch_id': batch_id,
                    'is_active': True,
                    'created_at': datetime.now(pytz.utc),
                    'mfa_enabled': False
                }
                user_id = mongo_db.users.insert_one(user_doc).inserted_id

                student_doc = {
                    'user_id': user_id,
                    'name': student['student_name'],
                    'roll_number': student['roll_number'],
                    'email': student['email'],
                    'mobile_number': student.get('mobile_number', ''),
                    'campus_id': campus['_id'],
                    'course_id': course['_id'],
                    'batch_id': batch_id,
                    'created_at': datetime.now(pytz.utc)
                }
                mongo_db.students.insert_one(student_doc)

                created_students_details.append({
                    "student_name": student['student_name'],
                    "email": student['email'],
                    "username": username,
                    "password": password
                })

            except Exception as student_error:
                errors.append(f"An error occurred for student {student.get('student_name', 'N/A')}: {str(student_error)}")

        # 3. Send emails with progress updates
        current_user_id = get_jwt_identity()
        total_emails = len(created_students_details)
        
        # Send initial progress update
        socketio.emit('upload_progress', {
            'user_id': current_user_id,
            'status': 'sending_emails',
            'total': total_emails,
            'processed': 0,
            'percentage': 0,
            'message': 'Sending welcome emails to students...'
        }, room=str(current_user_id))
        
        for index, student_details in enumerate(created_students_details):
             email_sent = False
             email_error = None
             try:
                html_content = render_template('student_credentials.html', params={
                    'name': student_details['student_name'],
                    'username': student_details['username'],
                    'email': student_details['email'],
                    'password': student_details['password'],
                    'login_url': "https://pydah-studyedge.vercel.app/login"
                })
                send_email(to_email=student_details['email'], to_name=student_details['student_name'], subject="Welcome to VERSANT - Your Student Credentials", html_content=html_content)
                email_sent = True
                
             except Exception as email_error:
                email_error = str(email_error)
                current_app.logger.error(f"Failed to send welcome email to {student_details['email']}: {email_error}")
                # Don't add to errors array - just log it
                
             # Send progress update after email attempt (regardless of success/failure)
             percentage = int(((index + 1) / total_emails) * 100)
             if email_sent:
                 socketio.emit('upload_progress', {
                     'user_id': current_user_id,
                     'status': 'sending_emails',
                     'total': total_emails,
                     'processed': index + 1,
                     'percentage': percentage,
                     'message': f'Email sent to {student_details["student_name"]} ({student_details["email"]})',
                     'current_student': {
                         'name': student_details['student_name'],
                         'email': student_details['email'],
                         'username': student_details['username']
                     }
                 }, room=str(current_user_id))
             else:
                 socketio.emit('upload_progress', {
                     'user_id': current_user_id,
                     'status': 'sending_emails',
                     'total': total_emails,
                     'processed': index + 1,
                     'percentage': percentage,
                     'message': f'Email sending failed for {student_details["student_name"]} ({student_details["email"]}) - Student created successfully',
                     'current_student': {
                         'name': student_details['student_name'],
                         'email': student_details['email'],
                         'username': student_details['username']
                     },
                     'email_warning': True,
                     'email_error': email_error
                 }, room=str(current_user_id))
        
        if errors:
            return jsonify({
                'success': False, 
                'message': f"Batch '{name}' created, but some students could not be added.", 
                'data': {'batch_id': str(batch_id), 'created_students': created_students_details}, 
                'errors': errors
            }), 207
        
        return jsonify({
            'success': True, 
            'message': f"Batch '{name}' created successfully with {len(created_students_details)} students.", 
            'data': {'batch_id': str(batch_id), 'created_students': created_students_details}
        }), 201

    except Exception as e:
        current_app.logger.error(f"Error creating batch with students: {e}")
        return jsonify({'success': False, 'message': f'An unexpected error occurred: {str(e)}'}), 500

@batch_management_bp.route('/<batch_id>/add-students', methods=['POST'])
@jwt_required()
def add_students_to_batch(batch_id):
    try:
        # Support both file upload (preferred) and JSON (legacy)
        if 'student_file' in request.files:
            file = request.files['student_file']
            batch_obj_id = ObjectId(batch_id)
            batch = mongo_db.batches.find_one({'_id': batch_obj_id})
            if not batch:
                return jsonify({'success': False, 'message': 'Batch not found.'}), 404
            campus_ids = batch.get('campus_ids', [])
            course_ids = batch.get('course_ids', [])
            if not campus_ids or not course_ids:
                return jsonify({'success': False, 'message': 'Batch is missing campus or course info.'}), 400
            campus_id = campus_ids[0]  # Assume single campus per batch for now
            # Parse and validate file
            rows = _parse_student_file(file)
            if not rows:
                return jsonify({'success': False, 'message': 'File is empty or invalid.'}), 400
            # Get campus and course info
            campus = mongo_db.campuses.find_one({'_id': campus_id})
            valid_course_names = set(c['name'] for c in mongo_db.courses.find({'_id': {'$in': course_ids}}))
            # Fetch existing data for validation
            existing_roll_numbers = set(s['roll_number'] for s in mongo_db.students.find({}, {'roll_number': 1}))
            existing_emails = set(u['email'] for u in mongo_db.users.find({}, {'email': 1}))
            existing_mobile_numbers = set(u.get('mobile_number', '') for u in mongo_db.users.find({'mobile_number': {'$exists': True, '$ne': ''}}, {'mobile_number': 1}))
            preview_data = []
            for row in rows:
                student_data = {
                    'campus_name': str(row.get('Campus Name', '')).strip(),
                    'course_name': str(row.get('Course Name', '')).strip(),
                    'student_name': str(row.get('Student Name', '')).strip(),
                    'roll_number': str(row.get('Roll Number', '')).strip(),
                    'email': str(row.get('Email', '')).strip().lower(),
                    'mobile_number': str(row.get('Mobile Number', '')).strip(),
                }
                errors = []
                if not all([student_data['campus_name'], student_data['course_name'], student_data['student_name'], student_data['roll_number'], student_data['email']]):
                    errors.append('Missing required fields.')
                if student_data['roll_number'] in existing_roll_numbers:
                    errors.append('Roll number already exists.')
                if student_data['email'] in existing_emails:
                    errors.append('Email already exists.')
                if student_data['mobile_number'] and student_data['mobile_number'] in existing_mobile_numbers:
                    errors.append('Mobile number already exists.')
                if student_data['campus_name'] != campus['name']:
                    errors.append(f"Campus '{student_data['campus_name']}' doesn't match batch campus '{campus['name']}'.")
                if student_data['course_name'] not in valid_course_names:
                    errors.append(f"Course '{student_data['course_name']}' not found in this batch.")
                student_data['errors'] = errors
                preview_data.append(student_data)
            # Only add students with no errors
            created_students_details = []
            errors = []
            uploaded_emails = []  # Track emails for verification
            
            for student in preview_data:
                if student['errors']:
                    errors.append(f"{student['student_name']}: {', '.join(student['errors'])}")
                    continue
                try:
                    course = mongo_db.courses.find_one({'name': student['course_name'], '_id': {'$in': course_ids}})
                    username = student['roll_number']
                    password = f"{student['student_name'].split()[0][:4].lower()}{student['roll_number'][-4:]}"
                    password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
                    
                    user_doc = {
                        'username': username,
                        'email': student['email'],
                        'password_hash': password_hash,
                        'role': ROLES['STUDENT'],
                        'name': student['student_name'],
                        'mobile_number': student.get('mobile_number', ''),
                        'campus_id': campus_id,
                        'course_id': course['_id'],
                        'batch_id': batch_obj_id,
                        'is_active': True,
                        'created_at': datetime.now(pytz.utc),
                        'mfa_enabled': False
                    }
                    
                    # Insert user first
                    user_result = mongo_db.users.insert_one(user_doc)
                    if not user_result.inserted_id:
                        errors.append(f"{student['student_name']}: Failed to create user account.")
                        continue
                    
                    user_id = user_result.inserted_id
                    
                    student_doc = {
                        'user_id': user_id,
                        'name': student['student_name'],
                        'roll_number': student['roll_number'],
                        'email': student['email'],
                        'mobile_number': student.get('mobile_number', ''),
                        'campus_id': campus_id,
                        'course_id': course['_id'],
                        'batch_id': batch_obj_id,
                        'created_at': datetime.now(pytz.utc)
                    }
                    
                    # Insert student profile
                    student_result = mongo_db.students.insert_one(student_doc)
                    if not student_result.inserted_id:
                        # Rollback user creation if student creation fails
                        mongo_db.users.delete_one({'_id': user_id})
                        errors.append(f"{student['student_name']}: Failed to create student profile.")
                        continue
                    
                    created_students_details.append({
                        "student_name": student['student_name'],
                        "email": student['email'],
                        "username": username,
                        "password": password
                    })
                    uploaded_emails.append(student['email'])
                    
                    # Send welcome email
                    try:
                        html_content = render_template('student_credentials.html', params={
                            'name': student['student_name'],
                            'username': username,
                            'email': student['email'],
                            'password': password,
                            'login_url': "https://pydah-studyedge.vercel.app/login"
                        })
                        send_email(to_email=student['email'], to_name=student['student_name'], subject="Welcome to Study Edge - Your Student Credentials", html_content=html_content)
                    except Exception as e:
                        errors.append(f"Failed to send email to {student['email']}: {e}")
                        
                except Exception as student_error:
                    errors.append(f"An error occurred for student {student.get('student_name', 'N/A')}: {str(student_error)}")
            
            # Verify upload success
            verification_results = []
            if uploaded_emails:
                students_verified = list(mongo_db.students.find({'email': {'$in': uploaded_emails}, 'batch_id': batch_obj_id}))
                users_verified = list(mongo_db.users.find({'email': {'$in': uploaded_emails}, 'batch_id': batch_obj_id}))
                
                for email in uploaded_emails:
                    student_exists = any(s['email'] == email for s in students_verified)
                    user_exists = any(u['email'] == email for u in users_verified)
                    verification_results.append({
                        'email': email,
                        'student_profile_exists': student_exists,
                        'user_account_exists': user_exists,
                        'fully_uploaded': student_exists and user_exists
                    })
            
            if errors:
                return jsonify({
                    'success': bool(created_students_details), 
                    'message': f"Process completed with {len(errors)} errors.", 
                    'data': {
                        'created_students': created_students_details,
                        'verification_results': verification_results
                    }, 
                    'errors': errors
                }), 207
            
            return jsonify({
                'success': True, 
                'message': f"Successfully added {len(created_students_details)} students to the batch.", 
                'data': {
                    'created_students': created_students_details,
                    'verification_results': verification_results
                }
            }), 201
        # Fallback: legacy JSON method
        data = request.get_json()
        students_data = data.get('students', []) if data else []
        if not students_data:
            return jsonify({'success': False, 'message': 'Student data are required.'}), 400
        batch_obj_id = ObjectId(batch_id)
        batch = mongo_db.batches.find_one({'_id': batch_obj_id})
        if not batch:
            return jsonify({'success': False, 'message': 'Batch not found.'}), 404
        valid_campus_ids = batch.get('campus_ids', [])
        valid_course_ids = batch.get('course_ids', [])
        created_students_details = []
        errors = []
        for student in students_data:
            try:
                campus = mongo_db.campuses.find_one({'name': student['campus_name']})
                if not campus or campus['_id'] not in valid_campus_ids:
                    errors.append(f"Campus '{student['campus_name']}' is not valid for this batch for student '{student.get('student_name', 'N/A')}'.")
                    continue
                course = mongo_db.courses.find_one({'name': student['course_name'], 'campus_id': campus['_id']})
                if not course or course['_id'] not in valid_course_ids:
                    errors.append(f"Course '{student['course_name']}' is not valid for this batch for student '{student.get('student_name', 'N/A')}'.")
                    continue
                # Check for existing user with same roll number, email, or mobile number
                existing_user = mongo_db.users.find_one({
                    '$or': [
                        {'username': student['roll_number']},
                        {'email': student['email']}
                    ]
                })
                if existing_user:
                    errors.append(f"Student with roll number '{student['roll_number']}' or email '{student['email']}' already exists.")
                    continue
                # Check for duplicate mobile number if provided
                if student.get('mobile_number') and mongo_db.users.find_one({'mobile_number': student['mobile_number']}):
                    errors.append(f"Student with mobile number '{student['mobile_number']}' already exists.")
                    continue
                username = student['roll_number']
                password = f"{student['student_name'].split()[0][:4].lower()}{student['roll_number'][-4:]}"
                password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
                user_doc = {
                    'username': username,
                    'email': student['email'],
                    'password_hash': password_hash,
                    'role': ROLES['STUDENT'],
                    'name': student['student_name'],
                    'mobile_number': student.get('mobile_number', ''),
                    'campus_id': campus['_id'],
                    'course_id': course['_id'],
                    'batch_id': batch_obj_id,
                    'is_active': True,
                    'created_at': datetime.now(pytz.utc),
                    'mfa_enabled': False
                }
                user_id = mongo_db.users.insert_one(user_doc).inserted_id
                student_doc = {
                    'user_id': user_id,
                    'name': student['student_name'],
                    'roll_number': student['roll_number'],
                    'email': student['email'],
                    'mobile_number': student.get('mobile_number', ''),
                    'campus_id': campus['_id'],
                    'course_id': course['_id'],
                    'batch_id': batch_obj_id,
                    'created_at': datetime.now(pytz.utc)
                }
                mongo_db.students.insert_one(student_doc)
                created_students_details.append({
                    "student_name": student['student_name'],
                    "email": student['email'],
                    "username": username,
                    "password": password
                })
                # Send welcome email
                try:
                    html_content = render_template('student_credentials.html', params={
                        'name': student['student_name'],
                        'username': username,
                        'email': student['email'],
                        'password': password,
                        'login_url': "https://pydah-studyedge.vercel.app/login"
                    })
                    send_email(to_email=student['email'], to_name=student['student_name'], subject="Welcome to VERSANT - Your Student Credentials", html_content=html_content)
                except Exception as e:
                    errors.append(f"Failed to send email to {student['email']}: {e}")
            except Exception as student_error:
                errors.append(f"An error occurred for student {student.get('student_name', 'N/A')}: {str(student_error)}")
        if errors:
            return jsonify({'success': bool(created_students_details), 'message': f"Process completed with {len(errors)} errors.", 'data': {'created_students': created_students_details}, 'errors': errors}), 207
        return jsonify({'success': True, 'message': f"Successfully added {len(created_students_details)} students to the batch.", 'data': {'created_students': created_students_details}}), 201
    except Exception as e:
        current_app.logger.error(f"Error adding students to batch: {e}")
        return jsonify({'success': False, 'message': f'An unexpected error occurred: {str(e)}'}), 500

@batch_management_bp.route('/add-student', methods=['POST'])
@jwt_required()
def add_single_student():
    try:
        data = request.get_json()
        batch_id = data.get('batch_id')
        course_id = data.get('course_id')
        name = data.get('name')
        roll_number = data.get('roll_number')
        email = data.get('email')
        mobile_number = data.get('mobile_number')

        if not all([batch_id, course_id, name, roll_number, email]):
            return jsonify({'success': False, 'message': 'All fields are required.'}), 400

        batch = mongo_db.batches.find_one({'_id': ObjectId(batch_id)})
        if not batch:
            return jsonify({'success': False, 'message': 'Batch not found.'}), 404
        campus_ids = batch.get('campus_ids', [])
        if not campus_ids:
            return jsonify({'success': False, 'message': 'Batch is missing campus info.'}), 400
        campus_id = campus_ids[0]  # Assume single campus per batch for now

        # Validate course_id
        valid_course_ids = set(str(cid) for cid in batch.get('course_ids', []))
        if course_id not in valid_course_ids:
            return jsonify({'success': False, 'message': 'Course is not valid for this batch.'}), 400

        # Check for existing user
        if mongo_db.users.find_one({'username': roll_number}):
            return jsonify({'success': False, 'message': 'Roll number already exists.'}), 400
        if mongo_db.users.find_one({'email': email}):
            return jsonify({'success': False, 'message': 'Email already exists.'}), 400
        if mobile_number and mongo_db.users.find_one({'mobile_number': mobile_number}):
            return jsonify({'success': False, 'message': 'Mobile number already exists.'}), 400

        # Create user and student
        password = f"{name.split()[0][:4].lower()}{roll_number[-4:]}"
        password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
        user_doc = {
            'username': roll_number,
            'email': email,
            'password_hash': password_hash,
            'role': 'student',
            'name': name,
            'mobile_number': mobile_number,
            'campus_id': campus_id,
            'course_id': ObjectId(course_id),
            'batch_id': ObjectId(batch_id),
            'is_active': True,
            'created_at': datetime.now(pytz.utc),
            'mfa_enabled': False
        }
        user_id = mongo_db.users.insert_one(user_doc).inserted_id
        student_doc = {
            'user_id': user_id,
            'name': name,
            'roll_number': roll_number,
            'email': email,
            'mobile_number': mobile_number,
            'campus_id': campus_id,
            'course_id': ObjectId(course_id),
            'batch_id': ObjectId(batch_id),
            'created_at': datetime.now(pytz.utc)
        }
        mongo_db.students.insert_one(student_doc)
        # Send welcome email
        try:
            html_content = render_template(
                'student_credentials.html',
                params={
                    'name': name,
                    'username': roll_number,
                    'email': email,
                    'password': password,
                    'login_url': "https://pydah-studyedge.vercel.app/login"
                }
            )
            send_email(
                to_email=email,
                to_name=name,
                subject="Welcome to VERSANT - Your Student Credentials",
                html_content=html_content
            )
        except Exception as e:
            return jsonify({'success': True, 'message': 'Student added, but failed to send email.', 'created_students': [{
                'student_name': name,
                'username': roll_number,
                'password': password,
                'email': email
            }], 'email_error': str(e)}), 200
        return jsonify({'success': True, 'message': 'Student added successfully!', 'created_students': [{
            'student_name': name,
            'username': roll_number,
            'password': password,
            'email': email
        }]}), 200
    except Exception as e:
        current_app.logger.error(f"Error adding single student: {str(e)}")
        return jsonify({'success': False, 'message': f'An unexpected server error occurred: {str(e)}'}), 500

@batch_management_bp.route('/student/<student_id>/access-status', methods=['GET'])
@jwt_required()
def get_student_access_status(student_id):
    try:
        from config.constants import MODULES, LEVELS
        student = mongo_db.students.find_one({'_id': ObjectId(student_id)})
        # Default logic if student not found or no authorized_levels
        default_grammar_unlocked = ['GRAMMAR_NOUN']
        modules_status = []
        if not student or not student.get('authorized_levels'):
            for module_id, module_name in MODULES.items():
                levels = [
                    {
                        'level_id': level_id,
                        'level_name': level['name'] if isinstance(level, dict) else level,
                        'unlocked': (
                            (module_id == 'GRAMMAR' and level_id == 'GRAMMAR_NOUN') or
                            (module_id == 'VOCABULARY')
                        )
                    }
                    for level_id, level in LEVELS.items()
                    if (level.get('module_id') if isinstance(level, dict) else None) == module_id
                ]
                modules_status.append({
                    'module_id': module_id,
                    'module_name': module_name,
                    'unlocked': (
                        module_id == 'GRAMMAR' or module_id == 'VOCABULARY'
                    ),
                    'levels': levels
                })
            return jsonify({'success': True, 'data': modules_status}), 200
        # If student has authorized_levels, use them
        authorized_levels = set(student.get('authorized_levels', []))
        for module_id, module_name in MODULES.items():
            levels = [
                {
                    'level_id': level_id,
                    'level_name': level['name'] if isinstance(level, dict) else level,
                    'unlocked': level_id in authorized_levels
                }
                for level_id, level in LEVELS.items()
                if (level.get('module_id') if isinstance(level, dict) else None) == module_id
            ]
            modules_status.append({
                'module_id': module_id,
                'module_name': module_name,
                'unlocked': all(l['unlocked'] for l in levels) if levels else False,
                'levels': levels
            })
        return jsonify({'success': True, 'data': modules_status}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching access status: {e}")
        return jsonify({'success': False, 'message': 'An error occurred fetching access status.'}), 500

# --- BATCH-COURSE INSTANCE MANAGEMENT ---

@batch_management_bp.route('/instances', methods=['GET'])
@jwt_required()
def get_batch_course_instances():
    """Get all batch-course instances with details"""
    try:
        pipeline = [
            {
                '$lookup': {
                    'from': 'batches',
                    'localField': 'batch_id',
                    'foreignField': '_id',
                    'as': 'batch'
                }
            },
            {
                '$lookup': {
                    'from': 'courses',
                    'localField': 'course_id',
                    'foreignField': '_id',
                    'as': 'course'
                }
            },
            {
                '$lookup': {
                    'from': 'campuses',
                    'localField': 'batch.campus_ids',
                    'foreignField': '_id',
                    'as': 'campuses'
                }
            },
            {
                '$unwind': '$batch'
            },
            {
                '$unwind': '$course'
            },
            {
                '$project': {
                    '_id': 1,
                    'batch_name': '$batch.name',
                    'course_name': '$course.name',
                    'campus_names': '$campuses.name',
                    'student_count': {'$size': {'$ifNull': ['$students', []]}},
                    'created_at': 1
                }
            },
            {'$sort': {'created_at': -1}}
        ]
        
        instances = list(mongo_db.db.batch_course_instances.aggregate(pipeline))
        
        # Convert ObjectIds to strings
        for instance in instances:
            instance['_id'] = str(instance['_id'])
            if 'created_at' in instance:
                instance['created_at'] = safe_isoformat(instance['created_at'])
        
        return jsonify({'success': True, 'data': instances}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching batch-course instances: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/instances/<instance_id>', methods=['GET'])
@jwt_required()
def get_batch_course_instance_details(instance_id):
    """Get detailed information about a specific batch-course instance"""
    try:
        instance = mongo_db.db.batch_course_instances.find_one({'_id': ObjectId(instance_id)})
        if not instance:
            return jsonify({'success': False, 'message': 'Instance not found'}), 404
        
        # Get batch details
        batch = mongo_db.batches.find_one({'_id': instance['batch_id']})
        course = mongo_db.courses.find_one({'_id': instance['course_id']})
        
        # Get students in this instance
        students = list(mongo_db.students.find({'batch_course_instance_id': ObjectId(instance_id)}))
        
        # Get test results for this instance
        test_results = list(mongo_db.student_test_attempts.find({'batch_course_instance_id': ObjectId(instance_id)}))
        
        instance_data = {
            'id': str(instance['_id']),
            'batch': {
                'id': str(batch['_id']),
                'name': batch['name']
            } if batch else None,
            'course': {
                'id': str(course['_id']),
                'name': course['name']
            } if course else None,
            'student_count': len(students),
            'test_results_count': len(test_results),
            'created_at': safe_isoformat(instance.get('created_at')) if instance.get('created_at') else None
        }
        
        return jsonify({'success': True, 'data': instance_data}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching instance details: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/instances/<instance_id>/students', methods=['GET'])
@jwt_required()
def get_instance_students(instance_id):
    """Get all students in a specific batch-course instance"""
    try:
        students = list(mongo_db.students.find({'batch_course_instance_id': ObjectId(instance_id)}))
        
        student_list = []
        for student in students:
            # Get user details
            user = mongo_db.users.find_one({'_id': student['user_id']})
            student_list.append({
                'id': str(student['_id']),
                'name': student['name'],
                'roll_number': student['roll_number'],
                'email': student['email'],
                'username': user['username'] if user else '',
                'is_active': user['is_active'] if user else True,
                'created_at': safe_isoformat(student['created_at']) if student.get('created_at') else None
            })
        
        return jsonify({'success': True, 'data': student_list}), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching instance students: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/instances/<instance_id>/upload-students', methods=['POST'])
@jwt_required()
def upload_students_to_instance(instance_id):
    """Upload students to a specific batch-course instance"""
    try:
        # Verify instance exists
        instance = mongo_db.db.batch_course_instances.find_one({'_id': ObjectId(instance_id)})
        if not instance:
            return jsonify({'success': False, 'message': 'Instance not found'}), 404
        
        # Get batch and course details
        batch = mongo_db.batches.find_one({'_id': instance['batch_id']})
        course = mongo_db.courses.find_one({'_id': instance['course_id']})
        
        if not batch or not course:
            return jsonify({'success': False, 'message': 'Batch or course not found'}), 404
        
        # Handle file upload
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        # Parse student file
        rows = _parse_student_file(file)
        if not rows:
            return jsonify({'success': False, 'message': 'File is empty or invalid'}), 400
        
        # Validate columns
        columns = list(rows[0].keys()) if rows else []
        required_fields = ['Student Name', 'Roll Number', 'Email', 'Mobile Number']
        missing_fields = [field for field in required_fields if field not in columns]
        if missing_fields:
            return jsonify({'success': False, 'message': f"Invalid file structure. Missing columns: {', '.join(missing_fields)}"}), 400
        
        # Fetch existing data for validation
        existing_roll_numbers = set(s['roll_number'] for s in mongo_db.students.find({}, {'roll_number': 1}))
        existing_emails = set(u['email'] for u in mongo_db.users.find({}, {'email': 1}))
        
        created_students = []
        errors = []
        
        for row in rows:
            student_name = str(row.get('Student Name', '')).strip()
            roll_number = str(row.get('Roll Number', '')).strip()
            email = str(row.get('Email', '')).strip().lower()
            mobile_number = str(row.get('Mobile Number', '')).strip()
            
            # Validation
            if not all([student_name, roll_number, email]):
                errors.append(f"{student_name or roll_number or email}: Missing required fields")
                continue
            
            if roll_number in existing_roll_numbers:
                errors.append(f"{student_name}: Roll number already exists")
                continue
            
            if email in existing_emails:
                errors.append(f"{student_name}: Email already exists")
                continue
            
            try:
                # Create user account
                username = roll_number
                password = f"{student_name.split()[0][:4].lower()}{roll_number[-4:]}"
                password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
                
                user_doc = {
                    'username': username,
                    'email': email,
                    'password_hash': password_hash,
                    'role': 'student',
                    'name': student_name,
                    'mobile_number': mobile_number,
                    'campus_id': batch['campus_ids'][0] if batch.get('campus_ids') else None,
                    'course_id': course['_id'],
                    'batch_id': batch['_id'],
                    'batch_course_instance_id': ObjectId(instance_id),
                    'is_active': True,
                    'created_at': datetime.now(pytz.utc),
                    'mfa_enabled': False
                }
                
                user_id = mongo_db.users.insert_one(user_doc).inserted_id
                
                # Create student profile
                student_doc = {
                    'user_id': user_id,
                    'name': student_name,
                    'roll_number': roll_number,
                    'email': email,
                    'mobile_number': mobile_number,
                    'campus_id': batch['campus_ids'][0] if batch.get('campus_ids') else None,
                    'course_id': course['_id'],
                    'batch_id': batch['_id'],
                    'batch_course_instance_id': ObjectId(instance_id),
                    'created_at': datetime.now(pytz.utc)
                }
                
                mongo_db.students.insert_one(student_doc)
                
                created_students.append({
                    'name': student_name,
                    'email': email,
                    'username': username,
                    'password': password
                })
                
                # Update existing sets
                existing_roll_numbers.add(roll_number)
                existing_emails.add(email)
                
            except Exception as e:
                errors.append(f"{student_name}: {str(e)}")
        
        return jsonify({
            'success': True,
            'message': f'Successfully created {len(created_students)} students',
            'data': {
                'created_students': created_students,
                'errors': errors
            }
        }), 201
        
    except Exception as e:
        current_app.logger.error(f"Error uploading students to instance: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/<batch_id>/verify-students', methods=['POST'])
@jwt_required()
def verify_students_upload(batch_id):
    """Verify that students were actually uploaded to the batch"""
    try:
        data = request.get_json()
        student_emails = data.get('student_emails', [])
        student_roll_numbers = data.get('student_roll_numbers', [])
        
        if not student_emails and not student_roll_numbers:
            return jsonify({'success': False, 'message': 'No student emails or roll numbers provided for verification.'}), 400
        
        batch_obj_id = ObjectId(batch_id)
        batch = mongo_db.batches.find_one({'_id': batch_obj_id})
        if not batch:
            return jsonify({'success': False, 'message': 'Batch not found.'}), 404
        
        # Check for students in the batch
        query = {'batch_id': batch_obj_id}
        if student_emails:
            query['email'] = {'$in': student_emails}
        elif student_roll_numbers:
            query['roll_number'] = {'$in': student_roll_numbers}
        
        students = list(mongo_db.students.find(query))
        users = list(mongo_db.users.find({'batch_id': batch_obj_id, 'role': 'student'}))
        
        # Create lookup dictionaries
        student_lookup = {s['email']: s for s in students}
        user_lookup = {u['email']: u for u in users}
        
        verification_results = []
        for email in student_emails:
            student_exists = email in student_lookup
            user_exists = email in user_lookup
            
            verification_results.append({
                'email': email,
                'student_profile_exists': student_exists,
                'user_account_exists': user_exists,
                'fully_uploaded': student_exists and user_exists,
                'student_id': str(student_lookup[email]['_id']) if student_exists else None,
                'user_id': str(user_lookup[email]['_id']) if user_exists else None
            })
        
        total_students = len(student_emails)
        successful_uploads = sum(1 for r in verification_results if r['fully_uploaded'])
        failed_uploads = total_students - successful_uploads
        
        return jsonify({
            'success': True,
            'data': {
                'verification_results': verification_results,
                'summary': {
                    'total_students': total_students,
                    'successful_uploads': successful_uploads,
                    'failed_uploads': failed_uploads,
                    'success_rate': (successful_uploads / total_students * 100) if total_students > 0 else 0
                }
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Error verifying students upload: {e}")
        return jsonify({'success': False, 'message': f'An unexpected error occurred: {str(e)}'}), 500

@batch_management_bp.route('/<batch_id>/cleanup-failed-students', methods=['POST'])
@jwt_required()
def cleanup_failed_students(batch_id):
    """Clean up any orphaned user accounts or student profiles"""
    try:
        data = request.get_json()
        student_emails = data.get('student_emails', [])
        
        if not student_emails:
            return jsonify({'success': False, 'message': 'No student emails provided for cleanup.'}), 400
        
        batch_obj_id = ObjectId(batch_id)
        
        # Find orphaned records (users without corresponding student profiles or vice versa)
        cleanup_results = []
        
        for email in student_emails:
            user = mongo_db.users.find_one({'email': email, 'batch_id': batch_obj_id})
            student = mongo_db.students.find_one({'email': email, 'batch_id': batch_obj_id})
            
            if user and not student:
                # Orphaned user account - delete it
                mongo_db.users.delete_one({'_id': user['_id']})
                cleanup_results.append({
                    'email': email,
                    'action': 'deleted_orphaned_user',
                    'user_id': str(user['_id'])
                })
            elif student and not user:
                # Orphaned student profile - delete it
                mongo_db.students.delete_one({'_id': student['_id']})
                cleanup_results.append({
                    'email': email,
                    'action': 'deleted_orphaned_student',
                    'student_id': str(student['_id'])
                })
            elif not user and not student:
                cleanup_results.append({
                    'email': email,
                    'action': 'no_records_found'
                })
        
        return jsonify({
            'success': True,
            'message': f'Cleanup completed. {len(cleanup_results)} records processed.',
            'data': {
                'cleanup_results': cleanup_results
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Error cleaning up failed students: {e}")
        return jsonify({'success': False, 'message': f'An unexpected error occurred: {str(e)}'}), 500

@batch_management_bp.route('/students/<student_id>', methods=['DELETE'])
@jwt_required()
@require_permission(module='batch_management')
def delete_student_management(student_id):
    """Delete a student"""
    try:
        # Delete user account
        mongo_db.users.delete_one({'_id': ObjectId(student_id)})
        
        # Delete student profile
        mongo_db.students.delete_one({'user_id': ObjectId(student_id)})
        
        return jsonify({'success': True, 'message': 'Student deleted successfully'}), 200
        
    except Exception as e:
        current_app.logger.error(f"Error deleting student: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/students/<student_id>/send-credentials', methods=['POST'])
@jwt_required()
@require_permission(module='batch_management')
def send_student_credentials(student_id):
    """Send credentials to student"""
    try:
        student = mongo_db.users.find_one({'_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        
        # Get student profile for roll number
        student_profile = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
        
        # Generate password using consistent pattern: first 4 letters of name + last 4 digits of roll number
        student_name = student.get('name', '')
        roll_number = student_profile.get('roll_number', '') if student_profile else ''
        
        if student_name and roll_number:
            # Extract first 4 letters of first name and last 4 digits of roll number
            first_name = student_name.split()[0] if student_name.split() else student_name
            password = f"{first_name[:4].lower()}{roll_number[-4:]}"
        else:
            # Fallback to random password if name or roll number is missing
            import secrets
            import string
            password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(8))
        
        # Update password
        password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
        mongo_db.users.update_one(
            {'_id': ObjectId(student_id)},
            {'$set': {'password_hash': password_hash}}
        )
        
        # Send email with credentials
        subject = "Your Study Edge Login Credentials"
        html_content = render_template(
            'student_credentials.html',
            params={
                'name': student.get('name', ''),
                'email': student.get('email', ''),
                'username': student.get('username', ''),
                'password': password,
                'roll_number': student_profile.get('roll_number', '') if student_profile else '',
                'login_url': "https://pydah-studyedge.vercel.app/login"
            }
        )
        
        send_email(
            to_email=student.get('email'),
            to_name=student.get('name', 'Student'),
            subject=subject,
            html_content=html_content
        )
        
        return jsonify({'success': True, 'message': 'Credentials sent successfully'}), 200
        
    except Exception as e:
        current_app.logger.error(f"Error sending credentials: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@batch_management_bp.route('/students/<student_id>/credentials', methods=['GET'])
@jwt_required()
@require_permission(module='batch_management')
def download_student_credentials(student_id):
    """Download student credentials as CSV"""
    try:
        student = mongo_db.users.find_one({'_id': ObjectId(student_id)})
        if not student:
            return jsonify({'success': False, 'message': 'Student not found'}), 404
        
        student_profile = mongo_db.students.find_one({'user_id': ObjectId(student_id)})
        
        # Generate password using consistent pattern: first 4 letters of name + last 4 digits of roll number
        student_name = student.get('name', '')
        roll_number = student_profile.get('roll_number', '') if student_profile else ''
        
        if student_name and roll_number:
            # Extract first 4 letters of first name and last 4 digits of roll number
            first_name = student_name.split()[0] if student_name.split() else student_name
            password = f"{first_name[:4].lower()}{roll_number[-4:]}"
        else:
            # Fallback to random password if name or roll number is missing
            import secrets
            import string
            password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(8))
        
        # Update password
        password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
        mongo_db.users.update_one(
            {'_id': ObjectId(student_id)},
            {'$set': {'password_hash': password_hash}}
        )
        
        # Create CSV content
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Name', 'Email', 'Roll Number', 'Username', 'Password'])
        writer.writerow([
            student.get('name', ''),
            student.get('email', ''),
            student_profile.get('roll_number', '') if student_profile else '',
            student.get('username', ''),
            password
        ])
        
        output.seek(0)
        return output.getvalue(), 200, {
            'Content-Type': 'text/csv',
            'Content-Disposition': f'attachment; filename="{student.get("name", "student")}_credentials.csv"'
        }
        
    except Exception as e:
        current_app.logger.error(f"Error downloading credentials: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500